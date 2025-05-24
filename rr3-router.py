#!/usr/bin/env python3
import os
import sys
import json
import shutil
import socket
import subprocess
import tempfile
import threading
import glob
import paramiko, socket, time, re, os, json, yaml, threading, logging, sys, ipaddress, subprocess, secrets, string, webbrowser, argparse, gzip, base64, shutil, csv, io # Added csv and io for CSV handling
from typing import List, Dict, Any # Re-add typing imports
from datetime import datetime, timezone, timedelta # Re-add datetime imports


# Device type mapping for Netmiko (added by fix script)
def get_netmiko_device_type(device_type):
    """Map generic device types to Netmiko-specific device types"""
    mapping = {
        'router': 'cisco_ios',
        'switch': 'cisco_ios',
        'firewall': 'cisco_asa',
    }
    return mapping.get(device_type.lower(), device_type)
import paramiko
from colorama import Fore, Style, init as colorama_init
from netmiko import ConnectHandler, NetmikoTimeoutException, NetmikoAuthenticationException
from flask import Flask, render_template, render_template_string, redirect, url_for, request, jsonify, send_from_directory, flash, Response
from dotenv import load_dotenv, set_key, find_dotenv
from werkzeug.utils import secure_filename
from jinja2 import DictLoader

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as ReportLabImage
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import colors
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from flask_socketio import SocketIO, emit
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

colorama_init(autoreset=True)

TEMPLATE_BASE_NAME = "base_layout.html"
TEMPLATE_INDEX_NAME = "index_page.html"
TEMPLATE_SETTINGS_NAME = "settings_page.html"
TEMPLATE_VIEW_JSON_NAME = "view_json_page.html"
TEMPLATE_EDIT_INVENTORY_NAME = "edit_inventory_page.html"

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config['UPLOAD_FOLDER'] = 'inventories'
app.config['ALLOWED_EXTENSIONS'] = {'csv'} 
app.config['PORT'] = 5007  # Store port in app config for template access
socketio = SocketIO(app, async_mode=None, cors_allowed_origins="*")

# Global application state variables
ui_logs: List[str] = []
SENSITIVE_STRINGS_TO_REDACT: List[str] = []
APP_CONFIG: Dict[str, str] = {}
interactive_sessions: Dict[str, Dict[str, Any]] = {}

# State management lock to prevent race conditions
state_lock = threading.RLock()

# Centralized state management class
class ApplicationState:
    """Centralized state management to maintain consistent application state.
    
    This class provides thread-safe access to global state variables and ensures
    proper initialization and reset functionality.
    """
    def __init__(self):
        self.reset_state()
    
    def reset_state(self):
        """Reset all state variables to their initial values."""
        with state_lock:
            global AUDIT_PROGRESS, audit_thread, current_audit_progress
            global last_run_summary_data, current_run_failures
            global audit_paused, audit_pause_event, audit_thread_exception
            
            # Initialize or reset progress tracking
            AUDIT_PROGRESS = {
                'status': 'Not started',
                'percentage_complete': 0,
                'current_device': 'None',
                'device_status': 'Idle',
                'device_status_class': 'status-idle',
                'completed_devices': 0,
                'total_devices': 0,
                'current_phase': 'Not started',
                'phase_progress': 0,
                'overall_success_count': 0,
                'overall_warning_count': 0,
                'overall_failure_count': 0,
                'start_time': None,
                'end_time': None,
                'elapsed_time': '00:00:00'
            }
            
            # Reset current progress tracker
            current_audit_progress = AUDIT_PROGRESS.copy()
            
            # Reset summary data
            last_run_summary_data = {
                'successful_devices': [],
                'failed_devices': [],
                'warnings': [],
                'report_path': None
            }
            
            # Reset failures tracking
            current_run_failures = []
            
            # Reset audit control flags
            audit_paused = False
            audit_thread_exception = None
            
            # Create a new event for pause control
            audit_pause_event = threading.Event()
            audit_pause_event.set()  # Not paused by default
            
            # Clear audit thread reference
            audit_thread = None
            
            # Clear UI logs if requested
            global ui_logs
            ui_logs.clear()
            
            self.emit_state_update()
    
    def update_progress(self, progress_updates):
        """Update the audit progress with thread safety."""
        with state_lock:
            global AUDIT_PROGRESS, current_audit_progress
            
            for key, value in progress_updates.items():
                if key in AUDIT_PROGRESS:
                    AUDIT_PROGRESS[key] = value
                    current_audit_progress[key] = value
            
            self.emit_state_update()
    
    def emit_state_update(self):
        """Emit state updates to all connected clients."""
        try:
            global AUDIT_PROGRESS
            # Handle serialization internally to avoid dependency issues
            json_safe_progress = self._prepare_progress_for_json(AUDIT_PROGRESS)
            socketio.emit('progress_update', json_safe_progress, namespace='/')
            
            # Get audit_paused status safely
            audit_paused_status = False
            if 'audit_paused' in globals():
                audit_paused_status = audit_paused
                
            socketio.emit('state_update', {'audit_paused': audit_paused_status}, namespace='/')
        except Exception as e:
            print(f"Error emitting state update: {e}")
            
    def _prepare_progress_for_json(self, progress_data):
        """Prepare progress data for JSON serialization (internal helper)."""
        if not progress_data:
            return {}
            
        # Create a copy to avoid modifying the original
        serializable_data = {}
        
        for key, value in progress_data.items():
            # Handle datetime objects
            if isinstance(value, datetime):
                serializable_data[key] = value.strftime("%Y-%m-%d %H:%M:%S")
            # Handle other non-serializable objects
            elif hasattr(value, '__dict__'):
                serializable_data[key] = str(value)
            # Handle nested dictionaries
            elif isinstance(value, dict):
                serializable_data[key] = self._prepare_progress_for_json(value)
            # All other values are copied as-is
            else:
                serializable_data[key] = value
                
        return serializable_data

# Create global state manager instance
app_state = ApplicationState()

# Task Lifecycle States
class TaskState:
    """Enumeration of possible task states for proper lifecycle management."""
    IDLE = 'idle'
    STARTING = 'starting'
    RUNNING = 'running'
    PAUSING = 'pausing'
    PAUSED = 'paused'
    RESUMING = 'resuming'
    STOPPING = 'stopping'
    COMPLETED = 'completed'
    FAILED = 'failed'

class TaskLifecycleManager:
    """Manages the complete lifecycle of audit tasks with proper state transitions.
    
    This class ensures that resources are properly managed throughout the task lifecycle,
    transitions between states are handled correctly, and provides hooks for monitoring
    and debugging state changes.
    
    Benefits:
    - Prevents resource leaks by using context managers
    - Ensures consistent state transitions
    - Provides audit logging of all state changes
    - Handles cancellation and timeouts properly
    - Allows proper recovery from errors
    """
    def __init__(self):
        """Initialize a new task lifecycle manager."""
        self.state = TaskState.IDLE
        self.task_id = None
        self.start_time = None
        self.end_time = None
        self.current_operation = None
        self.state_history = []
        self.resources = []
        self._lock = threading.RLock()
        self._state_callbacks = {}
        self._operation_timeout = 60  # Default timeout in seconds

    def register_state_callback(self, state, callback):
        """Register a callback to be executed when entering a specific state."""
        if state not in self._state_callbacks:
            self._state_callbacks[state] = []
        self._state_callbacks[state].append(callback)
        
    def transition_to(self, new_state, reason=None):
        """Transition to a new state with proper logging and callbacks."""
        with self._lock:
            old_state = self.state
            timestamp = datetime.now()
            
            # Validate the state transition
            if not self._is_valid_transition(old_state, new_state):
                log_msg = f"Invalid state transition attempted: {old_state} → {new_state}"
                print(f"[ERROR] {log_msg}")
                log_to_ui_and_console(log_msg, log_level="ERROR")
                return False
                
            # Log the transition
            transition_record = {
                'from_state': old_state,
                'to_state': new_state,
                'timestamp': timestamp,
                'reason': reason
            }
            self.state_history.append(transition_record)
            
            # Update the state
            self.state = new_state
            
            # Log the transition
            log_msg = f"Task state changed: {old_state} → {new_state}"
            if reason:
                log_msg += f" (Reason: {reason})"
            print(f"[INFO] {log_msg}")
            log_to_ui_and_console(log_msg, log_level="INFO")
            
            # Execute callbacks for this state
            if new_state in self._state_callbacks:
                for callback in self._state_callbacks[new_state]:
                    try:
                        callback(old_state, new_state, reason)
                    except Exception as e:
                        print(f"[ERROR] Error in state callback: {e}")
            
            # Perform state-specific actions
            if new_state == TaskState.STARTING:
                self.start_time = timestamp
                self.task_id = f"task_{int(timestamp.timestamp())}"
            elif new_state in [TaskState.COMPLETED, TaskState.FAILED]:
                self.end_time = timestamp
                self._release_all_resources()
            
            # Update the UI with the new state
            app_state.update_progress({'status': new_state})
            
            return True
    
    def _is_valid_transition(self, from_state, to_state):
        """Check if a state transition is valid based on the state machine rules."""
        # Define valid transitions for each state
        valid_transitions = {
            TaskState.IDLE: [TaskState.STARTING],
            TaskState.STARTING: [TaskState.RUNNING, TaskState.FAILED],
            TaskState.RUNNING: [TaskState.PAUSING, TaskState.STOPPING, TaskState.COMPLETED, TaskState.FAILED],
            TaskState.PAUSING: [TaskState.PAUSED, TaskState.FAILED],
            TaskState.PAUSED: [TaskState.RESUMING, TaskState.STOPPING, TaskState.FAILED],
            TaskState.RESUMING: [TaskState.RUNNING, TaskState.FAILED],
            TaskState.STOPPING: [TaskState.IDLE, TaskState.FAILED],
            TaskState.COMPLETED: [TaskState.IDLE],
            TaskState.FAILED: [TaskState.IDLE]
        }
        
        return to_state in valid_transitions.get(from_state, [])
    
    def register_resource(self, resource, close_method='close'):
        """Register a resource to be automatically managed and closed when task ends."""
        with self._lock:
            self.resources.append((resource, close_method))
    
    def _release_all_resources(self):
        """Release all registered resources, ensuring proper cleanup."""
        with self._lock:
            resources_to_close = list(self.resources)
            self.resources = []
            
        # Close resources outside the lock to prevent deadlocks
        for resource, close_method in resources_to_close:
            try:
                if callable(close_method):
                    close_method(resource)
                else:
                    method = getattr(resource, close_method, None)
                    if method and callable(method):
                        method()
                    else:
                        print(f"[WARNING] Cannot close resource: {resource}, method '{close_method}' not found")
            except Exception as e:
                print(f"[ERROR] Error closing resource {resource}: {e}")
    
    def start_operation(self, operation_name, timeout=None):
        """Start a new operation with optional timeout."""
        self.current_operation = operation_name
        timeout = timeout or self._operation_timeout
        
        log_msg = f"Starting operation: {operation_name} (timeout: {timeout}s)"
        print(f"[DEBUG] {log_msg}")
        
        # Return a context manager for the operation
        return self.OperationContext(self, operation_name, timeout)
    
    class OperationContext:
        """Context manager for task operations to ensure proper cleanup."""
        def __init__(self, manager, operation_name, timeout):
            self.manager = manager
            self.operation_name = operation_name
            self.timeout = timeout
            self.start_time = datetime.now()
            self.timer = None
            
        def __enter__(self):
            # Set up the timeout if specified
            if self.timeout:
                self.timer = threading.Timer(self.timeout, self._handle_timeout)
                self.timer.daemon = True
                self.timer.start()
            return self
        
        def __exit__(self, exc_type, exc_val, exc_tb):
            # Cancel the timeout timer
            if self.timer:
                self.timer.cancel()
            
            # Log completion or failure
            duration = (datetime.now() - self.start_time).total_seconds()
            if exc_type is None:
                print(f"[DEBUG] Operation completed: {self.operation_name} ({duration:.2f}s)")
            else:
                error_msg = f"Operation failed: {self.operation_name} - {exc_type.__name__}: {exc_val}"
                print(f"[ERROR] {error_msg}")
                log_to_ui_and_console(error_msg, log_level="ERROR")
                
                # Transition to FAILED state if an exception occurred
                self.manager.transition_to(TaskState.FAILED, reason=f"{exc_type.__name__}: {exc_val}")
                
                # Don't suppress the exception
                return False
        
        def _handle_timeout(self):
            """Handle operation timeout by logging and potentially cancelling."""
            timeout_msg = f"Operation timed out: {self.operation_name} (timeout: {self.timeout}s)"
            print(f"[ERROR] {timeout_msg}")
            log_to_ui_and_console(timeout_msg, log_level="ERROR")
            
            # Transition to FAILED state
            self.manager.transition_to(TaskState.FAILED, reason=f"Timeout after {self.timeout}s")

# Create global task lifecycle manager
task_manager = TaskLifecycleManager()

def strip_ansi(text: str) -> str:
    return re.sub(r'\x1B(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])', '', text)

def sanitize_log_message(msg: str) -> str:
    sanitized_msg = str(msg)
    if APP_CONFIG:
        for key in ["JUMP_PASSWORD", "DEVICE_PASSWORD", "DEVICE_ENABLE"]:
            value = APP_CONFIG.get(key)
            if value and len(value) > 2:
                sanitized_msg = sanitized_msg.replace(value, "*******")
    sanitized_msg = re.sub(r'(password|secret)\s*[:=]\s*([^\s,;"\']+)', r'\1: ********', sanitized_msg, flags=re.IGNORECASE)
    sanitized_msg = re.sub(r'(password|secret)\s+([^\s,;"\']+)', r'\1 ********', sanitized_msg, flags=re.IGNORECASE)
    return sanitized_msg

def log_to_ui_and_console(msg, console_only=False, is_sensitive=False, end="\n", log_level="INFO", phase=None, device=None, **kw):
    """Enhanced logging function with log levels and context.
    
    Args:
        msg: The message to log
        console_only: If True, only print to console (legacy parameter, kept for compatibility)
        is_sensitive: If True, sanitize sensitive information
        end: String appended after the message (typically '\n' or '\r')
        log_level: Log level ("INFO", "WARNING", "ERROR", "DEBUG", "SUCCESS")
        phase: Current audit phase (for context)
        device: Current device being processed (for context)
        **kw: Additional keyword arguments for print
    """
    timestamp = datetime.now().strftime("%H:%M:%S")
    
    # Format log level prefix for console
    level_prefix = ""
    if log_level == "WARNING":
        level_prefix = f"{Fore.YELLOW}[WARNING]{Style.RESET_ALL} "
    elif log_level == "ERROR":
        level_prefix = f"{Fore.RED}[ERROR]{Style.RESET_ALL} "
    elif log_level == "DEBUG":
        level_prefix = f"{Fore.CYAN}[DEBUG]{Style.RESET_ALL} "
    elif log_level == "SUCCESS":
        level_prefix = f"{Fore.GREEN}[SUCCESS]{Style.RESET_ALL} "
    elif log_level == "INFO":
        level_prefix = ""
    
    # Format additional context information
    context_info = ""
    if phase:
        context_info += f"[PHASE:{phase}] "
    if device:
        context_info += f"[DEVICE:{device}] "
    
    # Process message for console
    processed_msg_console = str(msg)
    if is_sensitive:
        processed_msg_console = sanitize_log_message(processed_msg_console)
    # Combine all parts for console output
    console_output = f"[{timestamp}] {level_prefix}{context_info}{processed_msg_console}"
    print(console_output, end=end, flush=True, **kw)
    
    # For UI logs, preserve the exact same output as console (minus ANSI color codes)
    # This ensures what's shown in the browser matches what's in the terminal
    ui_console_output = strip_ansi(console_output)
    cleaned_progress_marker = " कार्य प्रगति पर है "
    
    if end == '\r':
        if ui_logs and cleaned_progress_marker in ui_logs[-1]:
            ui_logs[-1] = ui_console_output + cleaned_progress_marker
        else:
            ui_logs.append(ui_console_output + cleaned_progress_marker)
    else:
        if ui_logs and cleaned_progress_marker in ui_logs[-1] and not ui_console_output.startswith(f"[{timestamp}] [#"):
            ui_logs[-1] = ui_console_output
        else:
            ui_logs.append(ui_console_output)
            
    # Make sure the UI is updated with the latest logs
    if socketio:
        try:
            # Send log update with namespace to ensure proper routing
            socketio.emit('log_update', {'logs': ui_logs[-100:]}, namespace='/')
            
            # Debug log event emission
            print(f"Debug: Emitted log_update event with {len(ui_logs[-100:])} logs")
            
            # Also emit a progress update to keep the UI in sync
            # This ensures the dashboard stays updated with each log message
            global AUDIT_PROGRESS
            # Convert datetime objects to strings for JSON serialization
            json_safe_progress = prepare_progress_for_json(AUDIT_PROGRESS)
            socketio.emit('progress_update', json_safe_progress, namespace='/')
        except Exception as e:
            print(f"Error emitting update: {e}")  # Log errors but continue execution

def sync_progress_with_ui():
    """Synchronizes the audit progress state with the UI components"""
    global AUDIT_PROGRESS, current_audit_progress
    
    # Ensure UI always has the latest progress data
    try:
        # Print the debug information to help diagnose serialization issues
        print(f"[DEBUG] Synchronizing UI with audit progress\n")
        
        # Update the UI state
        try:
            # Convert datetime objects to strings for JSON serialization
            json_safe_progress = prepare_progress_for_json(AUDIT_PROGRESS)
            socketio.emit('progress_update', json_safe_progress, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
    except Exception as sync_err:
        print(f"[WARNING] Error in sync_progress_with_ui: {sync_err}")

def close_all_ssh_connections():
    """Thoroughly close all SSH connections and channels in the application.
    
    This function searches for SSH connections in global variables and attempts
    to close them properly to avoid resource leaks.
    
    Returns:
        int: Number of connections successfully closed
    """
    connections_closed = 0
    
    # Connections can exist in multiple places, so we try all possible locations
    try:
        # 1. Check global scope for SSH clients and channels
        global_connections = [obj for obj in globals().values() 
                          if isinstance(obj, paramiko.SSHClient) or 
                             isinstance(obj, paramiko.Channel)]
        
        # 2. Check for connections in interactive_sessions dictionary
        session_connections = []
        global interactive_sessions
        for sid, session in interactive_sessions.items():
            if 'client' in session and isinstance(session['client'], paramiko.SSHClient):
                session_connections.append(session['client'])
            if 'channel' in session and isinstance(session['channel'], paramiko.Channel):
                session_connections.append(session['channel'])
        
        # 3. Combine all found connections
        all_connections = global_connections + session_connections
        
        # Log what we found
        if all_connections:
            print(f"[DEBUG] Found {len(all_connections)} SSH connections/channels to close")
        
        # 4. Close each connection with proper error handling
        for conn in all_connections:
            try:
                conn_type = type(conn).__name__
                print(f"[DEBUG] Closing connection: {conn_type}")
                conn.close()
                connections_closed += 1
                # Log each connection closure individually for better tracking
                print(f"[DEBUG] Successfully closed {conn_type} connection")
            except Exception as close_err:
                print(f"[WARNING] Error closing connection: {close_err}")
        
        return connections_closed
    except Exception as e:
        print(f"[ERROR] Error in close_all_ssh_connections: {e}")
        return connections_closed

def execute_verbose_command(ssh_client, command, timeout=30, device_name="Unknown", is_sensitive=False):
    """Execute a command with verbose logging and return the output.
    
    Args:
        ssh_client: Paramiko SSH client
        command: Command to execute
        timeout: Command execution timeout in seconds
        device_name: Name of the device for logging
        is_sensitive: Whether the command contains sensitive information
        
    Returns:
        Tuple of (success, stdout_data, stderr_data, exit_status)
    """
    start_time = time.time()
    cmd_to_log = sanitize_log_message(command) if is_sensitive else command
    stdin, stdout, stderr = None, None, None
    channel = None
    
    # Log command execution start with detailed information
    log_to_ui_and_console(
        f"Executing command on {device_name} with timeout {timeout}s: {cmd_to_log}", 
        log_level="INFO", 
        phase="COMMAND_EXEC", 
        device=device_name,
        is_sensitive=is_sensitive
    )
    
    try:
        # Track when the command was sent
        command_sent_time = time.time()
        
        # Use a shorter timeout for the command to prevent hanging
        stdin, stdout, stderr = ssh_client.exec_command(command, timeout=timeout)
        
        # Get the channel and set a timeout for receiving exit status
        channel = stdout.channel
        channel.settimeout(timeout + 5)  # Add 5 seconds buffer to the timeout
        
        # Log that command was sent successfully
        log_to_ui_and_console(
            f"Command sent successfully, waiting for output (timeout: {timeout}s)",
            log_level="DEBUG", 
            phase="COMMAND_EXEC", 
            device=device_name
        )
        
        # Get exit status with timeout protection
        exit_status = stdout.channel.recv_exit_status()
        
        # Read output and error streams after getting exit status
        stdout_data = stdout.read().decode('utf-8', errors='replace').strip()
        stderr_data = stderr.read().decode('utf-8', errors='replace').strip()
        
        elapsed_time = time.time() - start_time
        
        # Log stdout data with appropriate truncation
        if stdout_data:
            stdout_lines = stdout_data.splitlines()
            stdout_size = len(stdout_data)
            stdout_line_count = len(stdout_lines)
            
            # Determine if we need to truncate the output
            truncated = stdout_size > 500
            if truncated:
                display_lines = stdout_lines[:10]  # Show first 10 lines
                display_output = "\n".join(display_lines)
                log_to_ui_and_console(
                    f"Command output ({stdout_line_count} lines, {stdout_size} bytes, truncated):\n{display_output}\n[...truncated...]\n",
                    log_level="DEBUG", 
                    phase="COMMAND_EXEC", 
                    device=device_name
                )
            else:
                log_to_ui_and_console(
                    f"Command output ({stdout_line_count} lines, {stdout_size} bytes):\n{stdout_data}",
                    log_level="DEBUG", 
                    phase="COMMAND_EXEC", 
                    device=device_name
                )
        else:
            log_to_ui_and_console(
                "Command produced no output",
                log_level="DEBUG", 
                phase="COMMAND_EXEC", 
                device=device_name
            )
        
        # Log stderr data if any
        if stderr_data:
            log_to_ui_and_console(
                f"Command produced error output:\n{stderr_data}",
                log_level="WARNING", 
                phase="COMMAND_EXEC", 
                device=device_name
            )
        
        # Log completion status with timing information
        command_status = "successful" if exit_status == 0 else f"failed with status {exit_status}"
        log_to_ui_and_console(
            f"Command execution {command_status} (elapsed: {elapsed_time:.2f}s)",
            log_level="SUCCESS" if exit_status == 0 else "WARNING", 
            phase="COMMAND_EXEC", 
            device=device_name
        )
        
        return True, stdout_data, stderr_data, exit_status
    except Exception as e:
        elapsed_time = time.time() - start_time
        error_type = type(e).__name__
        
        # Log failure with detailed error information
        log_to_ui_and_console(
            f"Command execution failed: {error_type}: {e} (elapsed: {elapsed_time:.2f}s)",
            log_level="ERROR", 
            phase="COMMAND_EXEC", 
            device=device_name
        )
        
        return False, "", f"{error_type}: {e}", -1

def ping_local(host: str) -> bool:
    """Ping a host from the local machine."""
    # Check if host is empty or None
    if not host:
        log_to_ui_and_console(f"Invalid host address: empty", log_level="ERROR", phase="CONNECTIVITY")
        return False
        
    # Check if host is localhost or loopback
    if host in ['localhost', '127.0.0.1']:
        log_to_ui_and_console(f"Host is localhost/loopback, considering as reachable", log_level="DEBUG", phase="CONNECTIVITY")
        return True
        
    is_windows = sys.platform.startswith('win')
    param_count = '-n' if is_windows else '-c'
    param_timeout_opt = '-w' if is_windows else '-W'
    timeout_val = "1000" if is_windows else "1"
    command = ["ping", param_count, "1", param_timeout_opt, timeout_val, host]
    
    # Log the exact command being executed
    command_str = ' '.join(command)
    log_to_ui_and_console(f"Executing: {command_str}", log_level="DEBUG", phase="CONNECTIVITY")
    
    try:
        # Shorter timeout to avoid long waits
        process_timeout = 2
        log_to_ui_and_console(f"Pinging host {host} with timeout {process_timeout}s...", log_level="INFO", phase="CONNECTIVITY")
        
        start_time = time.time()
        result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=process_timeout, text=True)
        elapsed_time = time.time() - start_time
        
        success = result.returncode == 0
        if success:
            # Extract ping response time if available
            response_time = "unknown"
            if result.stdout:
                time_match = re.search(r'time=([\d\.]+)\s*ms', result.stdout)
                response_time = time_match.group(1) + 'ms' if time_match else "unknown"
                
            log_to_ui_and_console(
                f"Ping to {host} successful (response time: {response_time}, process time: {elapsed_time:.2f}s)", 
                log_level="SUCCESS", 
                phase="CONNECTIVITY"
            )
            # Add more detailed output if in debug mode
            if result.stdout and len(result.stdout.strip()) > 0:
                log_to_ui_and_console(f"Ping output:\n{result.stdout}", log_level="DEBUG", phase="CONNECTIVITY")
        else:
            log_to_ui_and_console(
                f"Ping to {host} failed with return code {result.returncode} (elapsed: {elapsed_time:.2f}s)", 
                log_level="WARNING", 
                phase="CONNECTIVITY"
            )
            if result.stderr and len(result.stderr.strip()) > 0:
                log_to_ui_and_console(f"Ping error output:\n{result.stderr}", log_level="DEBUG", phase="CONNECTIVITY")
        return success
    except subprocess.TimeoutExpired:
        log_to_ui_and_console(f"Local ping to {host} timed out after {process_timeout}s", log_level="WARNING", phase="CONNECTIVITY")
        return False

def ping_remote(ssh: paramiko.SSHClient, ip: str) -> bool:
    """
Ping an IP address from the jump host via SSH.
    
Args:
    ssh: The SSH client connected to the jump host
    ip: IP address to ping
        
Returns:
    bool: True if ping was successful, False otherwise
    """
    # Start timestamp for performance tracking
    start_time = time.time()
    
    # Get ping path from config or use default
    ping_executable = APP_CONFIG.get("JUMP_PING_PATH", "/bin/ping")
    if not ping_executable:
        ping_executable = "ping"
        log_to_ui_and_console(
            f"Warning: JUMP_PING_PATH not set. Defaulting to 'ping'", 
            log_level="WARNING",
            phase="REMOTE_PING"
        )
    
    # Log detailed information about the remote ping operation
    log_to_ui_and_console(
        f"Initiating remote ping from jump host to {ip}", 
        log_level="INFO",
        phase="REMOTE_PING"
    )
    
    # Update activity timestamp if available in the current scope
    try:
        if 'update_activity_timestamp' in locals() or 'update_activity_timestamp' in globals():
            update_activity_timestamp()
    except Exception:
        pass
    
    # Construct ping command with more reliable parameters
    # Use more aggressive timeouts to prevent hanging
    command = f"{ping_executable} -c 2 -W 2 {ip}"
    log_to_ui_and_console(
        f"Remote ping command: {command}", 
        log_level="DEBUG",
        phase="REMOTE_PING"
    )
    
    # Initialize variables for better error handling
    stdin, stdout, stderr = None, None, None
    channel = None
    
    # Set up retry logic
    max_attempts = 2
    current_attempt = 1
    
    while current_attempt <= max_attempts:
        if current_attempt > 1:
            log_to_ui_and_console(
                f"Retry attempt {current_attempt}/{max_attempts} for ping to {ip}",
                log_level="INFO",
                phase="REMOTE_PING"
            )
            
        try:
            # Use a shorter timeout for the command to prevent hanging
            log_to_ui_and_console(
                f"Executing ping command (attempt {current_attempt}/{max_attempts})", 
                log_level="DEBUG",
                phase="CONNECTIVITY"
            )
            
            # Execute the command with much shorter explicit timeout to prevent hanging
            stdin, stdout, stderr = ssh.exec_command(command, timeout=5)
            
            # Get the channel and set a timeout for receiving exit status
            channel = stdout.channel
            channel.settimeout(6)  # Slightly longer than exec_command timeout
            
            # Log that we're waiting for ping response
            log_to_ui_and_console(
                f"Pinging host {ip} (attempt {current_attempt}/{max_attempts})...", 
                log_level="INFO",
                phase="CONNECTIVITY"
            )
            
            # Update activity timestamp if available
            try:
                if 'update_activity_timestamp' in locals() or 'update_activity_timestamp' in globals():
                    update_activity_timestamp()
            except Exception:
                pass
            
            # Get exit status with timeout protection
            exit_status = stdout.channel.recv_exit_status()
            
            # Read outputs for better diagnostics
            stdout_data = stdout.read().decode(errors='ignore').strip()
            stderr_data = stderr.read().decode(errors='ignore').strip()
            
            # If we got here, we have a result (success or failure), so break the retry loop
            break
            
        except socket.timeout as e:
            # Handle timeout specifically
            elapsed_time = time.time() - start_time
            log_to_ui_and_console(
                f"Ping attempt {current_attempt}/{max_attempts} to {ip} timed out after {elapsed_time:.2f}s", 
                log_level="WARNING",
                phase="REMOTE_PING"
            )
            
            # Clean up resources before retry
            for resource in [stdout, stderr, stdin, channel]:
                if resource:
                    try:
                        if hasattr(resource, 'close') and callable(resource.close):
                            resource.close()
                    except Exception as close_error:
                        pass
            
            # If this was the last attempt, return failure
            if current_attempt == max_attempts:
                log_to_ui_and_console(
                    f"All ping attempts to {ip} timed out", 
                    log_level="ERROR",
                    phase="REMOTE_PING"
                )
                return False
            
            # Otherwise, increment attempt counter and retry
            current_attempt += 1
            continue
            
        except Exception as e:
            # Handle other exceptions
            elapsed_time = time.time() - start_time
            log_to_ui_and_console(
                f"Ping attempt {current_attempt}/{max_attempts} to {ip} failed with error: {str(e)}", 
                log_level="ERROR",
                phase="REMOTE_PING"
            )
            
            # Clean up resources before retry
            for resource in [stdout, stderr, stdin, channel]:
                if resource:
                    try:
                        if hasattr(resource, 'close') and callable(resource.close):
                            resource.close()
                    except Exception as close_error:
                        pass
            
            # If this was the last attempt, return failure
            if current_attempt == max_attempts:
                return False
            
            # Otherwise, increment attempt counter and retry
            current_attempt += 1
            continue
        
    # Calculate elapsed time
    elapsed_time = time.time() - start_time
    
    # Extract ping statistics if available
    packet_loss = "unknown"
    avg_time = "unknown"
    if stdout_data:
        # Try to parse packet loss
        loss_match = re.search(r'(\d+)% packet loss', stdout_data)
        if loss_match:
            packet_loss = f"{loss_match.group(1)}%"
            
        # Try to parse average time
        rtt_match = re.search(r'min/avg/max/.+\s+=\s+[\d\.]+/([\d\.]+)/', stdout_data)
        if rtt_match:
            avg_time = f"{rtt_match.group(1)}ms"
    
    # Log detailed results
    result = exit_status == 0
    if result:
        log_to_ui_and_console(
            f"Remote ping to {ip} successful (packet loss: {packet_loss}, avg time: {avg_time}, elapsed: {elapsed_time:.2f}s)", 
            log_level="SUCCESS",
            phase="REMOTE_PING"
        )
        
        # Update activity timestamp if available
        try:
            if 'update_activity_timestamp' in locals() or 'update_activity_timestamp' in globals():
                update_activity_timestamp()
        except Exception:
            pass
    else:
        log_to_ui_and_console(
            f"Remote ping to {ip} failed (exit status: {exit_status}, elapsed: {elapsed_time:.2f}s)", 
            log_level="WARNING",
            phase="REMOTE_PING"
        )
        
        # Log error output if any
        if stderr_data:
            log_to_ui_and_console(
                f"Remote ping error output:\n{stderr_data}", 
                log_level="DEBUG",
                phase="REMOTE_PING"
            )
        
        # Log stdout for debugging
        if stdout_data:
            log_to_ui_and_console(
                f"Remote ping stdout:\n{stdout_data}", 
                log_level="DEBUG",
                phase="REMOTE_PING"
            )
    
    # Ensure all resources are properly closed
    try:
        for resource in [stdout, stderr, stdin, channel]:
            if resource:
                try:
                    if hasattr(resource, 'close') and callable(resource.close):
                        resource.close()
                except Exception as close_error:
                    log_to_ui_and_console(
                        f"Error closing resource: {str(close_error)}", 
                        log_level="DEBUG",
                        phase="REMOTE_PING"
                    )
    except Exception as cleanup_error:
        log_to_ui_and_console(
            f"Error during resource cleanup: {str(cleanup_error)}", 
            log_level="DEBUG",
            phase="REMOTE_PING"
        )
    
    return result

def banner_to_log_audit(title: str):
    """Display a prominent banner for a new audit phase.
    
    Args:
        title: Phase title to display
    """
    try: w = shutil.get_terminal_size(fallback=(80, 20)).columns
    except OSError: w = 80
    
    # Format the title for better visibility
    phase_name = title.replace('PHASE ', '').split('–')[0].strip() if '–' in title else title
    phase_description = title.split('–')[1].strip() if '–' in title else ''
    
    # Create the banner with colored output for console
    log_message_raw = f"\n{Fore.CYAN}{title} {'─' * (w - len(title) - 2)}{Style.RESET_ALL}"
    print(log_message_raw)  # Print directly to console
    
    # Use the new structured logging for UI
    log_to_ui_and_console(
        f"Starting: {phase_description if phase_description else title}",
        log_level="INFO",
        phase=phase_name
    )
    
    # Add a detailed description if available
    if phase_description:
        log_to_ui_and_console(
            f"Phase details: {phase_description}",
            log_level="DEBUG",
            phase=phase_name
        )

def bar_audit(pct: float, width=30) -> str:
    done = int(width * pct / 100); return f"[{'#' * done}{'.' * (width - done)}] {pct:5.1f}%"

def mark_audit(ok: bool) -> str:
    return f"{Fore.GREEN}✔{Style.RESET_ALL}" if ok else f"{Fore.RED}✖{Style.RESET_ALL}"

DOTENV_PATH = find_dotenv()
if not DOTENV_PATH:
    DOTENV_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
    if not os.path.exists(DOTENV_PATH):
        with open(DOTENV_PATH, "w") as f_dotenv:
            f_dotenv.write(
                "# Auto-generated .env file\n"
                "JUMP_HOST=172.16.39.128\nJUMP_USERNAME=root\nJUMP_PASSWORD=eve\n"
                "JUMP_PING_PATH=/bin/ping\nDEVICE_USERNAME=cisco\nDEVICE_PASSWORD=cisco\n"
                "DEVICE_ENABLE=cisco\nACTIVE_INVENTORY_FILE=inventory-list-v00.yml\n"
                "ACTIVE_INVENTORY_FORMAT=yaml\n" # New setting for inventory format
            )

def load_app_config():
    global APP_CONFIG, SENSITIVE_STRINGS_TO_REDACT
    load_dotenv(DOTENV_PATH, override=True)
    APP_CONFIG = {
        "JUMP_HOST": os.getenv("JUMP_HOST", "172.16.39.128"),
        "JUMP_USERNAME": os.getenv("JUMP_USERNAME", "root"),
        "JUMP_PASSWORD": os.getenv("JUMP_PASSWORD", "eve"),
        "JUMP_PING_PATH": os.getenv("JUMP_PING_PATH", "/bin/ping"),
        "DEVICE_USERNAME": os.getenv("DEVICE_USERNAME", "cisco"),
        "DEVICE_PASSWORD": os.getenv("DEVICE_PASSWORD", "cisco"),
        "DEVICE_ENABLE": os.getenv("DEVICE_ENABLE", "cisco"),
        "ACTIVE_INVENTORY_FILE": os.getenv("ACTIVE_INVENTORY_FILE", "inventory-list-v00.yml"),
        "ACTIVE_INVENTORY_FORMAT": os.getenv("ACTIVE_INVENTORY_FORMAT", "yaml").lower() # Default to yaml
    }
    SENSITIVE_STRINGS_TO_REDACT.clear()
    for key_config in ["JUMP_PASSWORD", "DEVICE_PASSWORD", "DEVICE_ENABLE"]:
        if APP_CONFIG[key_config] is None: APP_CONFIG[key_config] = ""
        elif APP_CONFIG[key_config]: SENSITIVE_STRINGS_TO_REDACT.append(APP_CONFIG[key_config])
    return APP_CONFIG
load_app_config()

DEFAULT_YAML_INVENTORY_FILENAME = "inventory-list-v00.yml"
DEFAULT_CSV_INVENTORY_FILENAME = "inventory-list-v00.csv" # New default
ACTIVE_INVENTORY_DATA: Dict[str, Any] = {} # This will always be a dict internally

# Progress tracking for audits
AUDIT_PROGRESS = {
    'status': 'idle',  # idle, running, paused, completed
    'current_device': None,
    'total_devices': 0,
    'completed_devices': 0,
    'start_time': None,
    'end_time': None,
    'current_device_start_time': None,
    'device_times': {},  # hostname -> duration in seconds
    'estimated_completion_time': None,
    'device_statuses': {}  # hostname -> status (success, warning, failure, in_progress)
}
AUDIT_SHOULD_PAUSE = False
AUDIT_SHOULD_STOP = False

# Progress tracking functions
def prepare_progress_for_json(progress_data):
    """Convert datetime objects to ISO format strings for JSON serialization and ensure all required fields are present"""
    json_safe_progress = {}
    
    # Copy all items except datetime objects
    for key, value in progress_data.items():
        if isinstance(value, datetime):
            json_safe_progress[key] = value.isoformat()
        elif isinstance(value, dict):
            # Handle nested dictionaries with possible datetime values
            json_safe_progress[key] = {}
            for nested_key, nested_value in value.items():
                if isinstance(nested_value, datetime):
                    json_safe_progress[key][nested_key] = nested_value.isoformat()
                elif isinstance(nested_value, dict):
                    # Handle deeply nested dictionaries (e.g., device_times with start/end times)
                    json_safe_progress[key][nested_key] = {}
                    for deep_key, deep_value in nested_value.items():
                        if isinstance(deep_value, datetime):
                            json_safe_progress[key][nested_key][deep_key] = deep_value.isoformat()
                        else:
                            json_safe_progress[key][nested_key][deep_key] = deep_value
                else:
                    json_safe_progress[key][nested_key] = nested_value
        else:
            json_safe_progress[key] = value
    
    # Get the current status to determine appropriate defaults
    current_status = json_safe_progress.get('status', 'idle')
    
    # Set appropriate status message based on current status
    status_message = 'Not running'
    if current_status == 'running':
        status_message = 'Audit in progress'
    elif current_status == 'completed':
        status_message = 'Audit completed'
    elif current_status == 'paused':
        status_message = 'Audit paused'
    elif current_status == 'idle':
        status_message = 'Ready'
    
    # Ensure all necessary fields exist with default values if not present
    required_fields = {
        # Status fields
        'status': current_status,
        'status_message': status_message,
        'current_phase': 'Not started' if current_status == 'idle' else ('Completed' if current_status == 'completed' else 'In progress'),
        
        # Device tracking
        'current_device': 'None',
        'total_devices': 0,
        'completed_devices': 0,
        
        # Timing information
        'start_time': None,
        'end_time': None,
        'current_device_start_time': None,
        'estimated_completion_time': None,
        
        # Status counters
        'overall_success_count': 0,
        'overall_warning_count': 0,
        'overall_failure_count': 0,
        'in_progress_count': 0,
        
        # Data structures
        'device_statuses': {},
        'device_times': {},
        
        # Calculated metrics
        'progress_percent': 0.0,
        'avg_device_processing_time': 0.0
    }
    
    for field, default_value in required_fields.items():
        if field not in json_safe_progress:
            json_safe_progress[field] = default_value
    
    # Add timestamp to help with debugging
    json_safe_progress['last_update_time'] = datetime.now().isoformat()
    
    return json_safe_progress

def sync_progress_with_ui():
    """Synchronize AUDIT_PROGRESS with current_audit_progress for UI updates
    and ensure all tracking data is consistently available in all audit states"""
    global AUDIT_PROGRESS, current_audit_progress
    
    # Get current audit status to determine appropriate defaults
    current_status = AUDIT_PROGRESS.get('status', 'idle')
    
    # Set appropriate status message based on current status
    status_message = 'Ready'
    if current_status == 'running':
        status_message = 'Audit in progress'
    elif current_status == 'completed':
        status_message = 'Audit completed'
    elif current_status == 'paused':
        status_message = 'Audit paused'
    elif current_status == 'error':
        status_message = 'Audit error'
    
    # Make sure AUDIT_PROGRESS has all required fields with state-appropriate defaults
    required_fields = {
        'status': current_status,
        'status_message': status_message,
        'current_device': None if current_status in ['idle', 'completed'] else AUDIT_PROGRESS.get('current_device'),
        'total_devices': AUDIT_PROGRESS.get('total_devices', 0),
        'completed_devices': AUDIT_PROGRESS.get('completed_devices', 0),
        'overall_success_count': AUDIT_PROGRESS.get('overall_success_count', 0),
        'overall_warning_count': AUDIT_PROGRESS.get('overall_warning_count', 0),
        'overall_failure_count': AUDIT_PROGRESS.get('overall_failure_count', 0),
        'current_phase': 'Not started' if current_status == 'idle' else 
                        ('Completed' if current_status == 'completed' else 
                        ('Paused' if current_status == 'paused' else 'In progress'))
    }
    
    for field, default_value in required_fields.items():
        if field not in AUDIT_PROGRESS:
            AUDIT_PROGRESS[field] = default_value
    
    # Update legacy current_audit_progress structure comprehensively
    mapping = [
        # (AUDIT_PROGRESS key, current_audit_progress key)
        ('status', 'status'),
        ('current_device', 'current_device_hostname'),
        ('completed_devices', 'devices_processed_count'),
        ('total_devices', 'total_devices_to_process'),
        ('status_message', 'status_message'),
        ('current_phase', 'current_phase'),
        ('overall_success_count', 'overall_success_count'),
        ('overall_warning_count', 'overall_warning_count'),
        ('overall_failure_count', 'overall_failure_count')
    ]
    
    for audit_key, ui_key in mapping:
        current_audit_progress[ui_key] = AUDIT_PROGRESS.get(audit_key, required_fields.get(audit_key))
    
    # Calculate percentage complete
    if AUDIT_PROGRESS.get('total_devices', 0) > 0:
        percentage = (AUDIT_PROGRESS.get('completed_devices', 0) / AUDIT_PROGRESS.get('total_devices', 0)) * 100
        current_audit_progress['percentage_complete'] = round(percentage, 1)
    else:
        current_audit_progress['percentage_complete'] = 0
        
    # Ensure timing information is properly set based on audit state
    if current_status == 'idle':
        # Reset timing information for idle state
        AUDIT_PROGRESS['start_time'] = None
        AUDIT_PROGRESS['end_time'] = None
        AUDIT_PROGRESS['current_device_start_time'] = None
        AUDIT_PROGRESS['estimated_completion_time'] = None
    elif current_status == 'completed' and 'end_time' not in AUDIT_PROGRESS:
        # Ensure completed audits have an end time
        AUDIT_PROGRESS['end_time'] = datetime.now()
    
    # Update device statuses visual indicators (for the UI)
    current_audit_progress['device_status_class'] = 'status-idle'
    if AUDIT_PROGRESS['status'] == 'running':
        if AUDIT_PROGRESS['current_device']:
            current_audit_progress['device_status_class'] = 'status-inprogress'
    
    # Detailed logging for debugging
    key_metrics = ['status', 'current_device', 'completed_devices', 'total_devices', 
                   'overall_success_count', 'overall_warning_count', 'overall_failure_count']
    for key in key_metrics:
        print(f"Progress sync: {key}={AUDIT_PROGRESS.get(key)}")
    
    return current_audit_progress  # Return for convenience
    
    # Update the UI with both legacy and enhanced progress formats (matching /get_audit_progress endpoint)
    try:
        # Calculate enhanced progress data (same format as in /get_audit_progress)
        progress_percent = 0
        if AUDIT_PROGRESS.get('total_devices', 0) > 0:
            progress_percent = (AUDIT_PROGRESS.get('completed_devices', 0) / AUDIT_PROGRESS.get('total_devices', 0)) * 100
        
        enhanced_progress = {
            'percentage': progress_percent,
            'status': AUDIT_PROGRESS.get('status', 'idle'),
            'current_device': AUDIT_PROGRESS.get('current_device'),
            'total_devices': AUDIT_PROGRESS.get('total_devices', 0),
            'completed_devices': AUDIT_PROGRESS.get('completed_devices', 0),
            'start_time': AUDIT_PROGRESS.get('start_time'),
            'end_time': AUDIT_PROGRESS.get('end_time'),
            'current_device_start_time': AUDIT_PROGRESS.get('current_device_start_time'),
            'estimated_completion_time': AUDIT_PROGRESS.get('estimated_completion_time'),
            'status_counts': {
                'success': AUDIT_PROGRESS.get('overall_success_count', 0),
                'warning': AUDIT_PROGRESS.get('overall_warning_count', 0),
                'failure': AUDIT_PROGRESS.get('overall_failure_count', 0),
                'in_progress': 1 if AUDIT_PROGRESS.get('status') == 'running' else 0
            },
            'device_statuses': AUDIT_PROGRESS.get('device_statuses', {})
        }
        
        # Emit with the same structure as /get_audit_progress
        data_to_send = {
            'progress': prepare_progress_for_json(AUDIT_PROGRESS),  # Legacy progress format
            'enhanced_progress': enhanced_progress  # New enhanced progress format
        }
        
        socketio.emit('progress_update', data_to_send, namespace='/')
    except Exception as e:
        print(f"Error emitting progress update: {e}")

def update_audit_progress(device=None, status=None, completed=False):
    """Update the audit progress tracking with device status and emit to frontend"""
    global AUDIT_PROGRESS, current_audit_progress
    current_time = datetime.now()
    
    # Initialize required fields if not present
    required_fields = {
        'status': 'running', 
        'device_statuses': {},
        'device_times': {},
        'completed_devices': 0,
        'total_devices': AUDIT_PROGRESS.get('total_devices', 0),
        'overall_success_count': 0,
        'overall_warning_count': 0,
        'overall_failure_count': 0,
        'current_device_start_time': None,
        'estimated_completion_time': None
    }
    
    for field, default_value in required_fields.items():
        if field not in AUDIT_PROGRESS:
            AUDIT_PROGRESS[field] = default_value
    
    # Update device status if provided
    if device:
        # Initialize device status if not already in tracking dict
        if device not in AUDIT_PROGRESS['device_statuses']:
            AUDIT_PROGRESS['device_statuses'][device] = 'pending'
            print(f"Initializing device '{device}' with status 'pending'")
            
        # Track device status changes
        if status:
            previous_status = AUDIT_PROGRESS['device_statuses'].get(device)
            print(f"Updating device '{device}' status: {previous_status} -> {status} (completed={completed})")
            
            # Update the actual device status in our tracking dict
            AUDIT_PROGRESS['device_statuses'][device] = status
                
            # Only increment counters if this device has a new terminal status
            # and is completing its audit process (different from previous status)
            if completed and previous_status != status:
                # Update status counts
                if status == 'success':
                    AUDIT_PROGRESS['overall_success_count'] += 1
                    print(f"Incremented success count to {AUDIT_PROGRESS['overall_success_count']}")
                elif status == 'warning':
                    AUDIT_PROGRESS['overall_warning_count'] += 1
                    print(f"Incremented warning count to {AUDIT_PROGRESS['overall_warning_count']}")
                elif status == 'failure':
                    AUDIT_PROGRESS['overall_failure_count'] += 1
                    print(f"Incremented failure count to {AUDIT_PROGRESS['overall_failure_count']}")
        
        # Update current device being processed
        AUDIT_PROGRESS['current_device'] = device
        current_audit_progress['current_device_hostname'] = device
        print(f"Current device set to: {device}")
        
        # Track time for this device if it's the first time we're seeing it
        if 'device_times' not in AUDIT_PROGRESS:
            AUDIT_PROGRESS['device_times'] = {}
            
        if device not in AUDIT_PROGRESS['device_times']:
            AUDIT_PROGRESS['device_times'][device] = {'start': current_time}
            AUDIT_PROGRESS['current_device_start_time'] = current_time
            print(f"Device '{device}' start time recorded: {current_time}")
    
    # Update completion status
    if completed:
        # Increment completed devices counter
        AUDIT_PROGRESS['completed_devices'] += 1
        current_audit_progress['devices_processed_count'] = AUDIT_PROGRESS['completed_devices']
        print(f"Device completed. Total completed: {AUDIT_PROGRESS['completed_devices']}/{AUDIT_PROGRESS['total_devices']}")
        
        # Record completion time for this device
        if device and 'device_times' in AUDIT_PROGRESS and device in AUDIT_PROGRESS['device_times']:
            AUDIT_PROGRESS['device_times'][device]['end'] = current_time
            print(f"Device '{device}' end time recorded: {current_time}")
        
        # Update percentage complete
        if AUDIT_PROGRESS.get('total_devices', 0) > 0:
            percentage = (AUDIT_PROGRESS['completed_devices'] / AUDIT_PROGRESS['total_devices']) * 100
            current_audit_progress['percentage_complete'] = round(percentage, 1)
            print(f"Progress: {percentage:.1f}%")
        
        # If all devices are completed, mark the audit as completed
        if AUDIT_PROGRESS['completed_devices'] >= AUDIT_PROGRESS['total_devices']:
            AUDIT_PROGRESS['status'] = 'completed'
            AUDIT_PROGRESS['end_time'] = current_time
            print(f"All devices completed. Audit marked as completed at {current_time}")
    
    # Calculate estimated completion time and other metrics
    if AUDIT_PROGRESS.get('completed_devices', 0) > 0 and AUDIT_PROGRESS.get('start_time'):
        try:
            # Calculate ETA
            elapsed_time = current_time - AUDIT_PROGRESS['start_time']
            avg_time_per_device = elapsed_time / AUDIT_PROGRESS['completed_devices']
            remaining_devices = AUDIT_PROGRESS['total_devices'] - AUDIT_PROGRESS['completed_devices']
            estimated_remaining_time = avg_time_per_device * remaining_devices
            estimated_completion_time = current_time + estimated_remaining_time
            AUDIT_PROGRESS['estimated_completion_time'] = estimated_completion_time
            print(f"ETA calculated: {estimated_completion_time}")
            
            # Calculate progress percentage with 1 decimal point precision
            if AUDIT_PROGRESS['total_devices'] > 0:
                progress_percent = (AUDIT_PROGRESS['completed_devices'] / AUDIT_PROGRESS['total_devices']) * 100
                AUDIT_PROGRESS['progress_percent'] = round(progress_percent, 1)
                print(f"Progress: {AUDIT_PROGRESS['progress_percent']}%")
            
            # Calculate device processing statistics
            if AUDIT_PROGRESS['device_times']:
                device_times_list = []
                for device, times in AUDIT_PROGRESS['device_times'].items():
                    if 'start' in times and 'end' in times:
                        duration = (times['end'] - times['start']).total_seconds()
                        device_times_list.append(duration)
                
                if device_times_list:
                    avg_duration = sum(device_times_list) / len(device_times_list)
                    AUDIT_PROGRESS['avg_device_processing_time'] = avg_duration
                    print(f"Average device processing time: {avg_duration:.2f} seconds")
        except Exception as calc_error:
            print(f"Error calculating metrics: {calc_error}")
    
    # Sync with UI data structure
    sync_progress_with_ui()
            
    # Always emit progress update to all clients
    try:
        # Convert datetime objects to strings for JSON serialization
        json_safe_progress = prepare_progress_for_json(AUDIT_PROGRESS)
        print(f"Emitting progress update via socketio: success={json_safe_progress.get('overall_success_count', 0)}, " 
              f"warning={json_safe_progress.get('overall_warning_count', 0)}, " 
              f"failure={json_safe_progress.get('overall_failure_count', 0)}")
        socketio.emit('progress_update', json_safe_progress, namespace='/')
    except Exception as e:
        print(f"Error emitting progress update: {e}")
        # Continue despite the error
    
        # Calculate estimated completion time
        if AUDIT_PROGRESS['completed_devices'] > 0 and AUDIT_PROGRESS['start_time']:
            try:
                elapsed = (current_time - AUDIT_PROGRESS['start_time']).total_seconds()
                avg_time_per_device = elapsed / AUDIT_PROGRESS['completed_devices']
                remaining_devices = AUDIT_PROGRESS['total_devices'] - AUDIT_PROGRESS['completed_devices']
                estimated_remaining_seconds = avg_time_per_device * remaining_devices
                
                AUDIT_PROGRESS['estimated_completion_time'] = current_time + timedelta(seconds=estimated_remaining_seconds)
            except Exception as e:
                log_to_ui_and_console(f"Error calculating estimated completion time: {e}")
                # Set a default estimated completion time
                AUDIT_PROGRESS['estimated_completion_time'] = current_time + timedelta(minutes=5)
    
    # Check if all devices are completed
    if AUDIT_PROGRESS['completed_devices'] >= AUDIT_PROGRESS['total_devices']:
        AUDIT_PROGRESS['status'] = 'completed'
        AUDIT_PROGRESS['end_time'] = current_time

def start_audit_progress(device_count):
    """Initialize the audit progress tracking and broadcast initial state"""
    global AUDIT_PROGRESS, current_audit_progress
    current_time = datetime.now()
    
    # Initialize AUDIT_PROGRESS structure with all required fields
    AUDIT_PROGRESS = {
        # Status fields
        'status': 'running',
        'status_message': 'Initializing audit...',
        'current_phase': 'Initialization',
        
        # Device counting
        'total_devices': device_count,
        'completed_devices': 0,
        'current_device': None,
        
        # Timing information
        'start_time': current_time,
        'end_time': None,
        'current_device_start_time': None,
        'estimated_completion_time': None,
        
        # Tracking dictionaries
        'device_times': {},
        'device_statuses': {},
        
        # Status counters
        'overall_success_count': 0,
        'overall_warning_count': 0,
        'overall_failure_count': 0,
        'in_progress_count': 0
    }
    
    print(f"Audit progress initialized with {device_count} devices at {current_time}")
    
    # Initialize current_audit_progress for UI compatibility
    current_audit_progress['current_device_hostname'] = 'None'
    current_audit_progress['devices_processed_count'] = 0
    current_audit_progress['total_devices_to_process'] = device_count
    current_audit_progress['percentage_complete'] = 0
    current_audit_progress['status'] = 'running'
    current_audit_progress['status_message'] = 'Initializing audit...'
    current_audit_progress['overall_success_count'] = 0
    current_audit_progress['overall_warning_count'] = 0
    current_audit_progress['overall_failure_count'] = 0
    
    # Emit initial state to all clients
    try:
        json_safe_progress = prepare_progress_for_json(AUDIT_PROGRESS)
        print(f"Emitting initial audit progress state via socketio")
        socketio.emit('progress_update', json_safe_progress, namespace='/')
    except Exception as e:
        print(f"Error emitting initial progress: {e}")
        # Continue despite the error
    
    # Initialize current_audit_progress for UI
    current_audit_progress.update({
        "status_message": "Initializing...", 
        "devices_processed_count": 0, 
        "total_devices_to_process": device_count, 
        "percentage_complete": 0, 
        "current_phase": "Initialization", 
        "current_device_hostname": "N/A"
    })
    
    # Emit initial progress to all clients
    try:
        # Convert datetime objects to strings for JSON serialization
        json_safe_progress = prepare_progress_for_json(AUDIT_PROGRESS)
        socketio.emit('progress_update', json_safe_progress, namespace='/')
    except Exception as e:
        print(f"Error emitting initial progress update: {e}")

def pause_audit_progress():
    """Pause the audit progress tracking"""
    global AUDIT_PROGRESS, current_audit_progress
    AUDIT_PROGRESS['status'] = 'paused'
    AUDIT_PROGRESS['status_message'] = 'Audit Paused'
    current_audit_progress['status_message'] = 'Audit Paused'
    
    # Emit progress update
    try:
        socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
    except Exception as e:
        print(f"Error emitting progress update: {e}")

def stop_audit_progress():
    """Stop the audit progress tracking and reset all state
    
    This function handles comprehensive cleanup of all audit-related resources and state,
    ensuring that the application is returned to a clean initial state ready for a new audit.
    """
    # Access all relevant global variables
    global AUDIT_PROGRESS, current_audit_progress, audit_pause_event, audit_thread
    global watchdog_active, watchdog_thread, jump_client, _AUDIT_RUNNING, audit_status
    global chart_data, current_run_failures, last_run_summary_data, audit_paused, detailed_reports_manifest
    
    # Store the current status for return value
    audit_status = AUDIT_PROGRESS.get('status', 'unknown')
    
    # Log the audit completion
    log_to_ui_and_console(f"Audit stopping with status: {audit_status}", log_level="INFO", phase="STOP")
    
    # Reset audit progress tracking - create completely fresh state
    AUDIT_PROGRESS = {
        'status': 'idle',
        'start_time': None,
        'end_time': None,
        'device_statuses': {},
        'completed_devices': 0,
        'total_devices': 0,
        'percentage': 0,
        'current_device': None,
        'current_device_start_time': None,
        'estimated_completion_time': None,
        'status_counts': {'success': 0, 'failure': 0, 'warning': 0, 'skipped': 0}
    }
    
    # Reset current_audit_progress structure
    current_audit_progress = {
        "status": "idle",
        "current_phase": "Ready",
        "status_message": "Audit stopped and reset",
        "percentage_complete": 0,
        "current_device": None,
        "overall_success_count": 0,
        "overall_warning_count": 0,
        "overall_failure_count": 0,
        "in_progress_count": 0,
        "completed_devices": 0,
        "total_devices": 0,
        "start_time": None,
        "end_time": None,
        "last_update_time": datetime.now(),
        "estimated_completion_time": None,
        "avg_device_processing_time": None,
        "device_times": {},
        "device_statuses": {}
    }
    
    # Reset global running flag and audit status
    _AUDIT_RUNNING = False
    audit_status = 'idle'  # Explicitly set audit_status to 'idle'
    
    # Initialize additional global state variables
    chart_data = {}
    current_run_failures = {}
    
    # Preserve the detailed_reports_manifest if it contains report references
    # This ensures report download buttons remain available after stopping an audit
    has_summary_reports = False
    if 'detailed_reports_manifest' in globals() and detailed_reports_manifest:
        # Check if there are summary reports available
        has_summary_reports = any(key in detailed_reports_manifest for key in 
                                ['summary_file', 'pdf_summary_file', 'excel_summary_file'])
        
    # Only reset the reports manifest if there are no summary reports
    if not has_summary_reports:
        # Clear the manifest but keep its existing structure
        if 'detailed_reports_manifest' in globals() and detailed_reports_manifest:
            # Preserve only the summary report entries
            preserved_keys = {}
            for key in ['summary_file', 'pdf_summary_file', 'excel_summary_file']:
                if key in detailed_reports_manifest:
                    preserved_keys[key] = detailed_reports_manifest[key]
            
            # Clear and restore preserved keys
            detailed_reports_manifest.clear()
            for key, value in preserved_keys.items():
                detailed_reports_manifest[key] = value
    
    # Initialize last_run_summary_data but preserve completion_time if it exists
    completion_time = None
    if 'last_run_summary_data' in globals() and last_run_summary_data and 'completion_time' in last_run_summary_data:
        completion_time = last_run_summary_data['completion_time']
    
    last_run_summary_data = {
        'total_devices': 0,
        'success_count': 0,
        'warning_count': 0,
        'failure_count': 0,
        'completion_time': completion_time,  # Preserve completion time
        'icmp_reachable': 0,
        'icmp_unreachable': 0
    }
    
    audit_paused = False
    
    # Create a fresh audit pause event
    if audit_pause_event:
        # Clear any existing event
        try:
            audit_pause_event.set()
        except Exception:
            pass
    # Create a new event object
    audit_pause_event = threading.Event()
    
    # Comprehensive cleanup of resources
    try:
        # Stop the watchdog
        watchdog_active = False
        if 'watchdog_thread' in globals() and watchdog_thread and hasattr(watchdog_thread, 'is_alive') and watchdog_thread.is_alive():
            log_to_ui_and_console("Stopping watchdog thread", log_level="DEBUG", phase="CLEANUP")
            # Let the thread exit naturally (it's a daemon thread)
        
        # Wait for audit thread to complete if it exists
        if 'audit_thread' in globals() and audit_thread and hasattr(audit_thread, 'is_alive') and audit_thread.is_alive():
            log_to_ui_and_console("Waiting for audit thread to complete", log_level="INFO", phase="CLEANUP")
            # We don't join here to avoid blocking, but we've set _AUDIT_RUNNING to False
        
        # Safely close all SSH connections with better error handling
        if 'jump_client' in globals() and jump_client:
            log_to_ui_and_console("Closing SSH connection to jump host", log_level="INFO", phase="CLEANUP")
            try:
                # First check if _transport exists and is iterable
                if hasattr(jump_client, '_transport'):
                    if isinstance(jump_client._transport, list):
                        transports = jump_client._transport
                    else:
                        # If it's not a list, try to handle it as a single transport
                        transports = [jump_client._transport]
                        
                    # Ensure all channels are closed
                    for transport in transports:
                        if transport and hasattr(transport, 'is_active') and transport.is_active():
                            if hasattr(transport, '_channels') and isinstance(transport._channels, dict):
                                for chan in transport._channels.values():
                                    if chan and hasattr(chan, 'closed') and not chan.closed:
                                        try:
                                            chan.close()
                                            log_to_ui_and_console(f"Closed channel {chan}", log_level="DEBUG", phase="CLEANUP")
                                        except Exception as chan_error:
                                            log_to_ui_and_console(f"Error closing channel: {chan_error}", log_level="DEBUG", phase="CLEANUP")
                
                # Close the client connection if close method exists
                if hasattr(jump_client, 'close'):
                    jump_client.close()
                    jump_client = None
                    log_to_ui_and_console("SSH connection closed successfully", log_level="SUCCESS", phase="CLEANUP")
            except Exception as close_error:
                log_to_ui_and_console(f"Error closing SSH connection: {close_error}", log_level="WARNING", phase="CLEANUP")
    except Exception as cleanup_error:
        log_to_ui_and_console(f"Error during cleanup: {cleanup_error}", log_level="ERROR", phase="CLEANUP")
    
    # Sync progress with UI
    sync_progress_with_ui()
    
    # Emit final progress update
    try:
        socketio.emit('progress_update', prepare_progress_for_json(AUDIT_PROGRESS), namespace='/')
    except Exception as e:
        print(f"Error emitting final progress update: {e}")
    
    log_to_ui_and_console("Audit has been fully stopped and reset", log_level="SUCCESS", phase="RESET")
    
    # Return the original audit status
    return audit_status

def resume_audit_progress():
    """Resume the audit progress tracking"""
    global AUDIT_PROGRESS, current_audit_progress
    AUDIT_PROGRESS['status'] = 'running'
    AUDIT_PROGRESS['status_message'] = 'Audit Resumed'
    current_audit_progress['status_message'] = 'Audit Resumed'
    
    # Emit progress update
    try:
        socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
    except Exception as e:
        print(f"Error emitting progress update: {e}")

def get_inventory_path(filename: str = None, file_format: str = None) -> str:
    current_upload_folder = app.config.get('UPLOAD_FOLDER', 'inventories')
    if not os.path.isabs(current_upload_folder):
        current_upload_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), current_upload_folder)
    
    # Default to CSV inventory file
    DEFAULT_CSV_INVENTORY_FILENAME = "inventory-list.csv"
    
    if filename is None:
        filename = APP_CONFIG.get("ACTIVE_INVENTORY_FILE", DEFAULT_CSV_INVENTORY_FILENAME)
        if file_format is None: # If format not given, use CSV
            file_format = APP_CONFIG.get("ACTIVE_INVENTORY_FORMAT", "csv")
            # Ensure filename has .csv extension
            if not filename.endswith(".csv"):
                filename = f"{os.path.splitext(filename)[0]}.csv"
    
    return os.path.join(current_upload_folder, filename)


def create_default_inventory_if_missing():
    upload_folder_path = app.config.get('UPLOAD_FOLDER', 'inventories')
    if not os.path.isabs(upload_folder_path):
        upload_folder_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), upload_folder_path)
    os.makedirs(upload_folder_path, exist_ok=True)

    # Default CSV only
    default_csv_path = get_inventory_path(DEFAULT_CSV_INVENTORY_FILENAME, "csv")
    if not os.path.exists(default_csv_path):
        default_csv_data = [
            ['hostname', 'ip', 'device_type', 'username', 'password', 'secret', 'ios_version', 'notes'],
            ['R0', '172.16.39.100', 'cisco_ios', 'cisco', 'cisco', 'cisco', '15.1', 'Default entry'],
            ['R1', '172.16.39.101', 'cisco_ios', 'cisco', 'cisco', 'cisco', '15.1', 'Default entry'],
            ['R2', '172.16.39.102', 'cisco_ios', 'cisco', 'cisco', 'cisco', '15.1', 'Default entry']
        ]
        with open(default_csv_path, 'w', newline='') as f_csv:
            writer = csv.writer(f_csv)
            writer.writerows(default_csv_data)
        log_to_ui_and_console(f"Created default CSV inventory: {default_csv_path}", console_only=True)
        
        # Set the default CSV as active if no active inventory is set
        if not APP_CONFIG.get("ACTIVE_INVENTORY_FILE") or not os.path.exists(get_inventory_path()):
            set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FILE", DEFAULT_CSV_INVENTORY_FILENAME)
            set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FORMAT", "csv")
            APP_CONFIG["ACTIVE_INVENTORY_FILE"] = DEFAULT_CSV_INVENTORY_FILENAME
            APP_CONFIG["ACTIVE_INVENTORY_FORMAT"] = "csv"
            log_to_ui_and_console(f"Set default CSV inventory as active: {DEFAULT_CSV_INVENTORY_FILENAME}", console_only=True)


def list_inventory_files() -> List[str]:
    upload_folder_path = app.config.get('UPLOAD_FOLDER', 'inventories')
    if not os.path.isabs(upload_folder_path):
        upload_folder_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), upload_folder_path)
    
    # Only list CSV files
    inv_files = glob.glob(os.path.join(upload_folder_path, "*.csv"))
    
    return sorted(list(set([os.path.basename(f_name) for f_name in inv_files])))



def get_next_inventory_version(base_name_pattern=r"inventory-list-v(\d+)\.") -> int:
    existing_inventories = list_inventory_files()
    if not existing_inventories: return 0
    max_version = -1
    for f_name_ver in existing_inventories:
        match = re.search(base_name_pattern, f_name_ver) # More generic pattern
        if match:
            try: max_version = max(max_version, int(match.group(1)))
            except ValueError: continue
    return max_version + 1

# Validation for the internal Python dictionary format
def validate_inventory_data_dict(data: Dict[str, Any]) -> tuple[bool, str]:
    """Validates the structure of inventory data (as loaded from YAML or converted from CSV).
       Returns (is_valid, message).
    """
    if not isinstance(data, dict):
        return False, f"Inventory data is not a dictionary. Type: {type(data)}"
    
    if "routers" not in data or not isinstance(data["routers"], dict):
        return False, "Inventory must contain a 'routers' key (dictionary)."
    
    # Check each router entry
    for router_name, router_details in data["routers"].items():
        if not isinstance(router_details, dict):
            return False, f"Router '{router_name}' details must be a dictionary. Found: {type(router_details)}"
        
        # Check for required fields in router details
        if "ip" not in router_details or not router_details["ip"]:
            return False, f"Router '{router_name}' is missing required 'ip' field or it's empty."
        
        if "device_type" not in router_details or not router_details["device_type"]:
            return False, f"Router '{router_name}' is missing required 'device_type' field or it's empty."
    
    return True, "Inventory data structure is valid."

def validate_csv_data_list(data_list: list, headers: list) -> tuple[bool, str]:
    """Validates a list of dictionaries (parsed from CSV).
    Checks for required headers and non-empty values for key fields.
    Returns (is_valid, message).
    """
    if not isinstance(data_list, list):
        return False, "Invalid CSV data format: should be a list of rows."
    
    # Allow empty CSV (no headers, no data)
    if not headers:
        if not data_list:
            return True, "CSV is empty (no headers, no data)."
        return False, "CSV has data but no headers."

    required_headers = ['hostname', 'ip', 'device_type']
    missing_headers = [h for h in required_headers if h not in headers]
    if missing_headers:
        return False, f"CSV is missing required headers: {', '.join(missing_headers)}."

    if not data_list:
        return True, "CSV has headers but no data rows."

    for i, row_dict in enumerate(data_list):
        if not isinstance(row_dict, dict):
            return False, f"Row {i+1}: Invalid row format, expected a dictionary."
        
        # Check for empty essential fields based on required_headers
        for header in required_headers:
            if not row_dict.get(header, '').strip():
                return False, f"Row {i+1}: Missing or empty value for required field '{header}'."
        
        # Basic IP format check
        ip_val = row_dict.get('ip','').strip()
        try:
            ipaddress.ip_address(ip_val)
        except ValueError:
            return False, f"Row {i+1}: Invalid IP address format '{ip_val}'."

    return True, "CSV data is valid."

# Helper validation functions for CSV data
def is_valid_hostname(hostname):
    """Validates if a hostname has a valid format."""
    if not hostname or len(hostname) > 255:
        return False
    # Basic hostname validation - alphanumeric chars, hyphens, and dots
    return bool(re.match(r'^[a-zA-Z0-9][-a-zA-Z0-9\.]*[a-zA-Z0-9]$', hostname))

def is_valid_ip(ip):
    """Validates if an IP address has a valid format."""
    try:
        ipaddress.ip_address(ip)
        return True
    except ValueError:
        # If not a valid IP, check if it's a valid hostname (for DNS names)
        return is_valid_hostname(ip)

def validate_csv_data_list(data_list: List[Dict], headers: List[str]) -> tuple[bool, str]:
    """Validates a list of CSV data rows directly without converting to YAML.
    
    Args:
        data_list: List of dictionaries representing rows from CSV
        headers: List of column headers from CSV
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if not data_list:
        return False, "CSV data is empty"
    
    if not headers:
        return False, "CSV headers are missing"
    
    # Check for required headers
    required_headers = ['hostname', 'ip', 'device_type']
    missing_headers = [h for h in required_headers if h not in headers]
    
    if missing_headers:
        return False, f"CSV is missing required headers: {', '.join(missing_headers)}"
    
    # Validate each row
    for i, row in enumerate(data_list):
        row_num = i + 1  # 1-based indexing for user-friendly messages
        
        # Check for required fields
        missing = [h for h in required_headers if not row.get(h,'').strip()]
        if missing:
            return False, f"Row {row_num} missing required fields: {', '.join(missing)}"
        
        # Validate hostname format
        hostname = row.get('hostname', '').strip()
        if not is_valid_hostname(hostname):
            return False, f"Row {row_num}: Invalid hostname format: {hostname}"
        
        # Validate IP address format
        ip = row.get('ip', '').strip()
        if not is_valid_ip(ip):
            return False, f"Row {row_num}: Invalid IP address format: {ip}"
    
    return True, ""

def validate_csv_inventory_row(row: Dict, row_num: int) -> tuple[bool, str, Dict[str, Any]]:
    """Validates a single row from CSV inventory and extracts router data.
    
    Args:
        row: Dictionary representing a single row from CSV
        row_num: Row number for error reporting
        
    Returns:
        Tuple of (is_valid, error_message, extracted_data)
    """
    if not row:
        return False, f"Row {row_num} is empty", {}
        
    # Check for required headers
    required_headers = ['hostname', 'ip', 'device_type']
    missing = [h for h in required_headers if not row.get(h,'').strip()]
    
    if missing:
        return False, f"Row {row_num} missing required fields: {', '.join(missing)}", {}
    
    # Extract data
    hostname = row.get('hostname', '').strip()
    ip = row.get('ip', '').strip()
    device_type = row.get('device_type', '').strip()
    
    # Basic validation
    if not is_valid_hostname(hostname):
        return False, f"Row {row_num}: Invalid hostname format: {hostname}", {}
    
    if not is_valid_ip(ip):
        return False, f"Row {row_num}: Invalid IP address format: {ip}", {}
        
    # Create router entry
    router_data = {
        'ip': ip,
        'device_type': device_type
    }
    
    # Add optional fields if present
    for field in row:
        if field not in ['hostname'] and row[field].strip():
            router_data[field] = row[field].strip()
    
    return True, "", {hostname: router_data}


# Helper function to convert router dictionary to list of dictionaries for CSV operations
def convert_router_dict_to_csv_list(router_dict: Dict[str, Any]) -> list[dict]:
    """Converts the internal router dictionary to a list of dictionaries for CSV operations."""
    csv_list = []
    if not isinstance(router_dict, dict) or "routers" not in router_dict:
        return csv_list
    
    for hostname, details in router_dict.get("routers", {}).items():
        # Create a row dictionary with hostname
        row = {'hostname': hostname}
        # Add all other details
        for key, value in details.items():
            row[key] = value
        csv_list.append(row)
    
    return csv_list


# CSV-related functions for inventory handling
def read_csv_data_from_str(csv_string):
    """Parses a CSV string into a list of dictionaries (data) and a list of headers.
       Returns (None, None) on critical error or if CSV is fundamentally unparsable.
       Returns ([], headers) for empty CSV with only headers.
       Returns ([], []) for completely empty CSV string.
    """
    if not csv_string.strip():
        return [], [] # Empty string, empty data, empty headers
    try:
        # Use io.StringIO to treat the string as a file
        f = io.StringIO(csv_string)
        # Sniff to find the dialect (handles various CSV styles)
        try:
            dialect = csv.Sniffer().sniff(f.read(1024)) # Read a sample
            f.seek(0) # Rewind to the beginning of the string buffer
            reader = csv.reader(f, dialect)
        except csv.Error: # If sniffer fails (e.g. too simple, or not CSV)
            f.seek(0)
            # Basic check: if it has commas, assume standard CSV, else treat as single column
            if ',' in f.readline():
                f.seek(0)
                reader = csv.reader(f) # Default dialect
            else: # No commas, might be single column or not CSV at all
                f.seek(0)
                # If it's just one line without commas, it could be a header or a single data point.
                # Or if multiple lines without commas, treat each as a single-column row.
                # This basic reader will try its best.
                reader = csv.reader(f)

        headers = next(reader, None) # Get the header row
        if headers is None:
             # This case means the CSV string was empty or had some issue where headers couldn't be read
            app.logger.warning("CSV string seems to be empty or headers could not be read.")
            return [], [] # No headers, no data

        # Clean headers (strip whitespace)
        cleaned_headers = [header.strip() for header in headers]
        
        # Read data rows into a list of dictionaries
        data = []
        for row_values in reader:
            if not any(field.strip() for field in row_values): # Skip completely blank lines
                continue
            # Create a dict, ensuring row has a value for each header (even if empty)
            # If row is shorter than headers, missing values are empty strings implicitly via get
            # If row is longer, extra values are ignored by zip if headers is shorter.
            # For robustness, we ensure all headers are present in each dict.
            row_dict = {}
            for i, header in enumerate(cleaned_headers):
                row_dict[header] = row_values[i].strip() if i < len(row_values) else ''
            data.append(row_dict)
        
        app.logger.debug(f"Parsed CSV. Headers: {cleaned_headers}, Data rows: {len(data)}")
        return data, cleaned_headers
    except Exception as e:
        app.logger.error(f"Error parsing CSV string: {e}\nString was: '{csv_string[:200]}...'")
        return None, None # Indicate critical parsing failure

def write_csv_data_to_str(csv_data_list_of_dicts, headers):
    """Converts a list of dictionaries and headers into a CSV formatted string."""
    if not isinstance(csv_data_list_of_dicts, list):
        app.logger.error("CSV data for writing must be a list of dictionaries.")
        return "" # Or raise an error
    if not headers and csv_data_list_of_dicts: # Try to infer headers if not provided but data exists
        app.logger.warning("Headers not provided for CSV string generation, attempting to infer.")
        headers = get_csv_headers_from_data(csv_data_list_of_dicts)
    elif not headers and not csv_data_list_of_dicts:
        app.logger.debug("Writing empty CSV string as no data and no headers provided.")
        return "" # No headers, no data, return empty string

    output = io.StringIO()
    try:
        # Use QUOTE_MINIMAL to only quote fields containing the delimiter, quotechar, or lineterminator
        # QUOTE_NONNUMERIC can be useful if you want all non-numeric fields quoted
        writer = csv.DictWriter(output, fieldnames=headers, quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
        writer.writeheader()
        for row_dict in csv_data_list_of_dicts:
            # Ensure row_dict only contains keys present in headers to avoid ValueError
            filtered_row = {header: row_dict.get(header, '') for header in headers}
            writer.writerow(filtered_row)
        return output.getvalue()
    except Exception as e:
        app.logger.error(f"Error writing CSV data to string: {e}")
        return "" # Return empty or error string
    finally:
        output.close()

def get_csv_headers_from_data(csv_data_list_of_dicts):
    """Infers CSV headers from a list of dictionaries.
       Tries to maintain a common order for known fields.
    """
    if not csv_data_list_of_dicts:
        return []
    
    # Collect all unique keys from all dictionaries
    all_keys = set()
    for row_dict in csv_data_list_of_dicts:
        all_keys.update(row_dict.keys())
    
    # Define priority order for common fields
    priority_fields = ['hostname', 'ip', 'device_type', 'username', 'password', 'ios_version', 'notes']
    
    # Start with priority fields (if they exist in the data)
    headers = [field for field in priority_fields if field in all_keys]
    
    # Add any remaining fields not in priority list
    remaining_fields = sorted(list(all_keys - set(headers)))
    headers.extend(remaining_fields)
    
    return headers

def convert_csv_data_to_yaml_dict(csv_string):
    """Converts CSV string data to YAML dictionary format.
    
    Args:
        csv_string: String containing CSV data
        
    Returns:
        Tuple of (is_valid, error_message, yaml_equivalent_data)
    """
    # Parse the CSV string into a list of dictionaries and headers
    parsed_csv_data, parsed_csv_headers = read_csv_data_from_str(csv_string)
    
    # Check if parsing was successful
    if parsed_csv_data is None:
        return False, "CSV content is malformed or unparsable", None
    
    # Validate the CSV data
    is_valid, msg = validate_csv_data_list(parsed_csv_data, parsed_csv_headers)
    if not is_valid and (parsed_csv_data or parsed_csv_headers):
        return False, msg, None
    
    # Convert to YAML dictionary format
    yaml_dict = {"routers": {}}
    
    # Process each row in the CSV data
    for row in parsed_csv_data:
        hostname = row.get('hostname')
        if not hostname:
            continue
            
        # Create router entry
        router_data = {}
        for key, value in row.items():
            if key != 'hostname' and value:  # Skip empty values
                router_data[key] = value
                
        # Add to routers dictionary
        yaml_dict["routers"][hostname] = router_data
    
    return True, "CSV data converted successfully", yaml_dict

def load_active_inventory():
    global ACTIVE_INVENTORY_DATA, APP_CONFIG
    create_default_inventory_if_missing()

    # Default to CSV inventory file
    DEFAULT_CSV_INVENTORY_FILENAME = "inventory-list.csv"
    active_file_basename = APP_CONFIG.get("ACTIVE_INVENTORY_FILE", DEFAULT_CSV_INVENTORY_FILENAME)
    active_format = APP_CONFIG.get("ACTIVE_INVENTORY_FORMAT", "csv").lower()
    
    # Ensure we're using CSV format
    if active_format != "csv":
        log_to_ui_and_console(f"Warning: Active inventory format '{active_format}' is not supported. Switching to CSV format.", console_only=True)
        active_format = "csv"
        set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FORMAT", "csv")
        APP_CONFIG["ACTIVE_INVENTORY_FORMAT"] = "csv"
    
    # Ensure filename has .csv extension
    if not active_file_basename.endswith(".csv"):
        active_file_basename = f"{os.path.splitext(active_file_basename)[0]}.csv"
        log_to_ui_and_console(f"Corrected active inventory filename to CSV: {active_file_basename}", console_only=True)
        set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FILE", active_file_basename)
        APP_CONFIG["ACTIVE_INVENTORY_FILE"] = active_file_basename

    inventory_path = get_inventory_path(active_file_basename, active_format)

    if not os.path.exists(inventory_path):
        log_to_ui_and_console(f"Warning: Active inventory '{active_file_basename}' not found. Creating default CSV inventory.", console_only=True)
        create_default_inventory_if_missing()
        inventory_path = get_inventory_path(DEFAULT_CSV_INVENTORY_FILENAME, "csv")
        if os.path.exists(inventory_path):
            set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FILE", DEFAULT_CSV_INVENTORY_FILENAME)
            APP_CONFIG["ACTIVE_INVENTORY_FILE"] = DEFAULT_CSV_INVENTORY_FILENAME
        else:
            log_to_ui_and_console(f"Critical: Default CSV inventory could not be created. Cannot load inventory.", console_only=True)
            ACTIVE_INVENTORY_DATA = {"routers": {}}
            return

    try:
        with open(inventory_path, 'r', newline='') as f_load_inv:
            csv_content_str = f_load_inv.read()
        
        # Parse CSV string into list of dictionaries and headers
        parsed_data, headers = read_csv_data_from_str(csv_content_str)
        
        if parsed_data is None:
            log_to_ui_and_console(f"Error: Failed to parse CSV data from '{inventory_path}'. Using empty inventory.", console_only=True)
            ACTIVE_INVENTORY_DATA = {"routers": {}}
            return
        
        # Validate the CSV data
        is_valid, msg = validate_csv_data_list(parsed_data, headers)
        
        if not is_valid:
            log_to_ui_and_console(f"Error: CSV inventory '{active_file_basename}' is invalid: {msg}. Using empty inventory.", console_only=True)
            ACTIVE_INVENTORY_DATA = {"routers": {}}
            return
        
        # Convert the list of dictionaries to the router dictionary format
        router_dict = {"routers": {}}
        for row in parsed_data:
            hostname = row.get('hostname', '').strip()
            if not hostname:
                continue
            
            # Create a copy of the row dict for the router details
            router_details = row.copy()
            # Remove hostname from details as it's used as the key
            if 'hostname' in router_details:
                del router_details['hostname']
            
            router_dict["routers"][hostname] = router_details
        
        ACTIVE_INVENTORY_DATA = router_dict
        log_to_ui_and_console(f"Successfully loaded CSV inventory: {active_file_basename} with {len(router_dict['routers'])} routers", console_only=True)
    except Exception as e_load_active:
        log_to_ui_and_console(f"Error loading CSV inventory '{inventory_path}': {e_load_active}. Using empty inventory.", console_only=True)
        ACTIVE_INVENTORY_DATA = {"routers": {}}

app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(os.path.abspath(__file__)), app.config['UPLOAD_FOLDER'])
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
load_active_inventory()

audit_status: str = "Not Run"
BASE_DIR_NAME: str = "ALL-ROUTER-REPORTS"
SUMMARY_FILENAME: str = "summary.txt"
PDF_SUMMARY_FILENAME: str = "audit_summary_report.pdf"
EXCEL_SUMMARY_FILENAME: str = "audit_summary_report.xlsx"
detailed_reports_manifest: Dict[str, Any] = {}
last_run_summary_data: Dict[str, Any] = {
    "total_routers": 0, "icmp_reachable": 0, "ssh_auth_ok": 0, "collected": 0, 
    "with_violations": 0, "failed_icmp": 0, "failed_ssh_auth": 0, "failed_collection": 0,
    "per_router_status": {},
    # Initialize counters for telnet violations
    "total_physical_line_issues": 0,
    "failure_category_explicit_telnet": 0,
    "failure_category_transport_all": 0,
    "failure_category_default_telnet": 0
}
current_run_failures: Dict[str, Any] = {}
audit_lock = threading.Lock()
audit_pause_event = threading.Event(); audit_pause_event.set()
audit_paused: bool = False
audit_thread: Any = None  # Thread for running the audit
current_audit_progress: Dict[str, Any] = {
    "status_message": "Idle", "devices_processed_count": 0, "total_devices_to_process": 0, 
    "percentage_complete": 0, "current_phase": "N/A", "current_device_hostname": "N/A"
}
last_audit_start_time_attempt: Any = None
last_successful_audit_completion_time: Any = None
class AuditError(Exception): pass

def opt_c_collect(net_connect: Any, r_name: str, device_type: str, show_line_output_str: str) -> Dict[str, Any]:
    log_to_ui_and_console(f"  Executing opt_c_collect for {r_name} (device_type: {device_type}). Received {len(show_line_output_str)} chars from 'show line'.", console_only=True)
    parsed_show_line = {}; tty_lines_info = []
    for line_sl in show_line_output_str.splitlines():
        if "TTY" in line_sl: tty_lines_info.append(line_sl.strip())
    parsed_show_line["tty_summary"] = tty_lines_info
    parsed_show_line["raw_output_preview"] = show_line_output_str[:200] + "..." if len(show_line_output_str) > 200 else show_line_output_str
    return {"report_c_info": f"Specific data for Report C from router {r_name}", "device_type_provided": device_type, "parsed_show_line_data": parsed_show_line}

def run_the_audit_logic():
    global audit_status, ui_logs, detailed_reports_manifest, last_run_summary_data, current_run_failures, audit_paused, audit_pause_event
    global current_audit_progress, last_audit_start_time_attempt, last_successful_audit_completion_time
    global AUDIT_PROGRESS, AUDIT_SHOULD_PAUSE, AUDIT_SHOULD_STOP
    
    # Initialize the enhanced progress tracking
    log_to_ui_and_console("Initializing audit process", log_level="INFO", phase="INIT")
    current_config = APP_CONFIG.copy()
    log_to_ui_and_console("Loaded current application configuration", log_level="DEBUG", phase="INIT")
    
    # Handle the new CSV inventory structure
    log_to_ui_and_console("Processing inventory data", log_level="INFO", phase="INIT")
    
    if 'data' in ACTIVE_INVENTORY_DATA and 'headers' in ACTIVE_INVENTORY_DATA:
        # New CSV structure
        csv_data = ACTIVE_INVENTORY_DATA.get('data', [])
        headers = ACTIVE_INVENTORY_DATA.get('headers', [])
        log_to_ui_and_console(
            f"Using CSV inventory structure with {len(csv_data)} devices and {len(headers)} columns", 
            log_level="INFO", 
            phase="INIT"
        )
        # Detailed debug logging of CSV headers
        log_to_ui_and_console(f"CSV headers: {headers}", log_level="DEBUG", phase="INIT")
        
        # Convert CSV data to router dictionary format for compatibility
        _INV_ROUTERS = {}
        valid_devices = 0
        for row in csv_data:
            hostname = row.get('hostname')
            if hostname:
                router_data = {k: v for k, v in row.items() if k != 'hostname'}
                _INV_ROUTERS[hostname] = router_data
                valid_devices += 1
                log_to_ui_and_console(
                    f"Added device: {hostname} with IP: {router_data.get('ip', 'Not specified')}", 
                    log_level="DEBUG", 
                    phase="INIT"
                )
            else:
                log_to_ui_and_console(
                    f"Skipping row with missing hostname: {row}", 
                    log_level="WARNING", 
                    phase="INIT"
                )
        
        log_to_ui_and_console(
            f"Processed {valid_devices} valid devices from CSV inventory", 
            log_level="SUCCESS", 
            phase="INIT"
        )
    else:
        # Old structure
        current_inventory = ACTIVE_INVENTORY_DATA.copy()
        _INV_ROUTERS = current_inventory.get("routers", {})
        # Ensure we're using 'devices' (plural) consistently
        inventory_msg = f"Using legacy inventory structure with {len(_INV_ROUTERS)} devices"
        print(f"DEBUG - About to log: {inventory_msg}")
        log_to_ui_and_console(
            inventory_msg,
            log_level="INFO", 
            phase="INIT"
        )
    
    _JUMP_HOST = current_config["JUMP_HOST"]; _JUMP_USERNAME = current_config["JUMP_USERNAME"]
    _JUMP_PASSWORD = current_config["JUMP_PASSWORD"]; _DEVICE_USERNAME = current_config["DEVICE_USERNAME"]
    _DEVICE_PASSWORD = current_config["DEVICE_PASSWORD"]; _DEVICE_ENABLE = current_config["DEVICE_ENABLE"]
    
    # Initialize both progress tracking systems (legacy and new)
    start_audit_progress(len(_INV_ROUTERS))
    AUDIT_SHOULD_PAUSE = False
    AUDIT_SHOULD_STOP = False
    
    current_audit_progress.update({"status_message": "Initializing...", "devices_processed_count": 0, "total_devices_to_process": len(_INV_ROUTERS), "percentage_complete": 0, "current_phase": "Initialization", "current_device_hostname": "N/A"})
    detailed_reports_manifest.clear()
    last_run_summary_data.update({
        "total_routers": len(_INV_ROUTERS), 
        "icmp_reachable": 0, 
        "ssh_auth_ok": 0, 
        "collected": 0, 
        "with_violations": 0, 
        "failed_icmp": 0, 
        "failed_ssh_auth": 0, 
        "failed_collection": 0, 
        "per_router_status": {r: "Pending" for r in _INV_ROUTERS},
        "total_physical_line_issues": 0,
        "failure_category_explicit_telnet": 0,
        "failure_category_transport_all": 0,
        "failure_category_default_telnet": 0
    })
    current_run_failures.clear(); [current_run_failures.update({r_name_init: None}) for r_name_init in _INV_ROUTERS]
    audit_status = "Running"; current_audit_progress["status_message"] = "Audit Running"
    script_dir = os.path.dirname(os.path.abspath(__file__)); base_report_path = os.path.join(script_dir, BASE_DIR_NAME)
    a_dir = os.path.join(base_report_path, "a-folder"); b_dir = os.path.join(base_report_path, "b-folder")
    c_dir = os.path.join(base_report_path, "c-folder"); summary_txt_file_path = os.path.join(base_report_path, SUMMARY_FILENAME)
    a_dir_name = "a-folder"; b_dir_name = "b-folder"; c_dir_name = "c-folder" # Define these for use later
    for d_path in (base_report_path, a_dir, b_dir, c_dir):
        try: os.makedirs(d_path, exist_ok=True)
        except OSError as e_mkdir: log_to_ui_and_console(f"{Fore.RED}Error creating dir {d_path}: {e_mkdir}{Style.RESET_ALL}"); audit_status = f"Failed: Dir creation error {d_path}"; raise AuditError(audit_status)
    ssh_cfg_file_obj = None; jump_client = None
    try:
        ssh_cfg_file_obj = tempfile.NamedTemporaryFile("w", delete=False)
        ssh_cfg_file_obj.write("Host *\n    KexAlgorithms +diffie-hellman-group1-sha1,diffie-hellman-group14-sha1,diffie-hellman-group-exchange-sha1\n    HostkeyAlgorithms +ssh-dss,ssh-rsa\n    PubkeyAcceptedKeyTypes +ssh-dss,ssh-rsa\n    Ciphers +aes128-cbc,aes128-ctr,aes192-ctr,aes256-ctr,3des-cbc\n    StrictHostKeyChecking no\n    UserKnownHostsFile /dev/null\n")
        ssh_cfg_file_obj.flush(); ssh_cfg_filename = ssh_cfg_file_obj.name
        current_audit_progress["current_phase"] = "Jump Host Connection"
        AUDIT_PROGRESS["current_phase"] = "Jump Host Connection"
        AUDIT_PROGRESS["status_message"] = "Connecting to Jump Host"
        current_audit_progress["status_message"] = "Connecting to Jump Host"
        
        # Emit progress update
        try:
            socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
            
        banner_to_log_audit("PHASE 0 – Jump-host Connectivity")
        
        # Check if jump host is configured
        log_to_ui_and_console("Verifying jump host configuration", log_level="INFO", phase="JUMP_HOST")
        if not _JUMP_HOST:
            # For testing/development, allow localhost as jump host
            _JUMP_HOST = "127.0.0.1"
            log_to_ui_and_console(
                "Jump host IP not configured – using localhost for testing", 
                log_level="WARNING", 
                phase="JUMP_HOST"
            )
            current_audit_progress["status_message"] = "Using localhost as jump host (testing mode)"
            log_to_ui_and_console(
                "This is a test/development mode and may have limited functionality", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
        else:
            log_to_ui_and_console(
                f"Jump host configured as {_JUMP_HOST} with user {_JUMP_USERNAME}", 
                log_level="INFO", 
                phase="JUMP_HOST",
                is_sensitive=True
            )
            current_audit_progress["status_message"] = f"Pinging Jump Host ({_JUMP_HOST})..."
            
        # Ping the jump host with improved error handling
        log_to_ui_and_console(f"Verifying ICMP connectivity to jump host {_JUMP_HOST}", log_level="INFO", phase="JUMP_HOST")
        ping_result = ping_local(_JUMP_HOST)
        if not ping_result:
            log_to_ui_and_console(
                f"Jump host {_JUMP_HOST} ICMP unreachable – proceeding in limited mode", 
                log_level="WARNING", 
                phase="JUMP_HOST"
            )
            log_to_ui_and_console(
                "Some audit functionality may be unavailable without jump host connectivity", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
            current_audit_progress["status_message"] = "Jump host unreachable - limited mode"
        else:
            log_to_ui_and_console(
                f"ICMP connectivity to jump host {_JUMP_HOST} verified successfully", 
                log_level="SUCCESS", 
                phase="JUMP_HOST"
            )
        
        # Prepare SSH connection
        current_audit_progress["status_message"] = f"SSH to Jump Host ({_JUMP_USERNAME}@{_JUMP_HOST})..."
        log_to_ui_and_console(
            f"Initializing SSH connection to jump host {_JUMP_HOST}", 
            log_level="INFO", 
            phase="JUMP_HOST"
        )
        
        # Initialize SSH client with detailed logging
        jump_client = paramiko.SSHClient()
        log_to_ui_and_console("Created SSH client instance", log_level="DEBUG", phase="JUMP_HOST")
        
        jump_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        log_to_ui_and_console(
            "Set AutoAddPolicy for SSH host keys (will accept unknown hosts)", 
            log_level="DEBUG", 
            phase="JUMP_HOST"
        )
        
        # Try to connect with SSH, with better error handling
        ssh_success = False
        log_to_ui_and_console(
            f"Attempting SSH connection to {_JUMP_HOST}:22 with timeout 15s", 
            log_level="INFO", 
            phase="JUMP_HOST"
        )
        
        connection_start_time = time.time()
        try:
            # Increased timeout for better reliability but still reasonable
            log_to_ui_and_console(
                "SSH connection parameters: allow_agent=False, look_for_keys=False, timeout=15s, banner_timeout=20s", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
            
            # Connect with more reliable timeouts
            jump_client.connect(_JUMP_HOST, 22, _JUMP_USERNAME, _JUMP_PASSWORD, 
                               allow_agent=False, look_for_keys=False, 
                               timeout=15, banner_timeout=20, auth_timeout=20)
            
            # Verify connection is working with a simple test command
            log_to_ui_and_console(
                "Testing SSH connection with simple command", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
            
            # Execute a simple command to verify the connection
            stdin, stdout, stderr = jump_client.exec_command("echo Connection test", timeout=10)
            channel = stdout.channel
            channel.settimeout(10)
            exit_status = stdout.channel.recv_exit_status()
            
            if exit_status != 0:
                raise Exception(f"Connection test failed with exit status {exit_status}")
                
            # Read and log test command output
            test_output = stdout.read().decode(errors='ignore').strip()
            log_to_ui_and_console(
                f"Connection test output: {test_output}", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
            
            connection_time = time.time() - connection_start_time
            log_to_ui_and_console(
                f"SSH connection to jump host {_JUMP_HOST} successful and verified (connected in {connection_time:.2f}s)", 
                log_level="SUCCESS", 
                phase="JUMP_HOST"
            )
            ssh_success = True
            
            # Close resources from test command
            for resource in [stdout, stderr, stdin]:
                if resource:
                    try:
                        resource.close()
                    except:
                        pass
                        
        except paramiko.AuthenticationException as e_auth:
            connection_time = time.time() - connection_start_time
            log_to_ui_and_console(
                f"SSH authentication failed for jump host {_JUMP_HOST}: {e_auth} (after {connection_time:.2f}s)", 
                log_level="ERROR", 
                phase="JUMP_HOST",
                is_sensitive=True
            )
            log_to_ui_and_console(
                "Check jump host credentials in settings or environment variables", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
            current_audit_progress["status_message"] = "SSH authentication failed - limited mode"
        except socket.timeout as e_timeout:
            connection_time = time.time() - connection_start_time
            log_to_ui_and_console(
                f"SSH connection to jump host {_JUMP_HOST} timed out after {connection_time:.2f}s: {e_timeout}", 
                log_level="ERROR", 
                phase="JUMP_HOST"
            )
            log_to_ui_and_console(
                "Verify jump host is reachable and SSH service is running on port 22", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
            current_audit_progress["status_message"] = "SSH connection timed out - limited mode"
        except Exception as e_other:
            connection_time = time.time() - connection_start_time
            error_type = type(e_other).__name__
            log_to_ui_and_console(
                f"SSH connection to jump host {_JUMP_HOST} failed with error: {error_type}: {e_other} (after {connection_time:.2f}s)", 
                log_level="ERROR", 
                phase="JUMP_HOST",
                is_sensitive=True
            )
            log_to_ui_and_console(
                "Detailed error information for troubleshooting", 
                log_level="DEBUG", 
                phase="JUMP_HOST"
            )
            current_audit_progress["status_message"] = f"SSH connection error: {error_type} - limited mode"
        # Check for pause
        if not audit_pause_event.is_set(): 
            log_to_ui_and_console(f"{Fore.YELLOW}Audit PAUSED. Waiting...{Style.RESET_ALL}")
            current_audit_progress["status_message"] = "Paused"
            audit_pause_event.wait()
            log_to_ui_and_console(f"{Fore.GREEN}Audit RESUMED.{Style.RESET_ALL}")
            
        # Set up a watchdog timer to detect and recover from stalled operations
        watchdog_active = True
        last_activity_time = time.time()
        
        def watchdog_timer():
            # Implementation of a watchdog to detect and recover from stalled operations
            watchdog_timeout = 30  # Reduced timeout to 30 seconds
            stall_counter = 0
            max_stalls_before_action = 3
            
            # Initialize last_activity_time to current time
            last_activity_time = time.time()
            
            while watchdog_active and not AUDIT_SHOULD_STOP:
                time.sleep(5)  # Check every 5 seconds
                current_time = time.time()
                time_since_activity = current_time - last_activity_time
                
                if time_since_activity > watchdog_timeout:
                    stall_counter += 1
                    log_to_ui_and_console(
                        f"Watchdog detected possible stall: No activity for {time_since_activity:.1f} seconds", 
                        log_level="WARNING", 
                        phase="WATCHDOG"
                    )
                    
                    # Take action after multiple consecutive stalls
                    if stall_counter >= max_stalls_before_action:
                        log_to_ui_and_console(
                            f"Watchdog taking action after {stall_counter} consecutive stalls", 
                            log_level="WARNING", 
                            phase="WATCHDOG"
                        )
                        # Force an activity timestamp update to prevent repeated interventions
                        last_activity_time = time.time()
                        stall_counter = 0
                else:
                    # Reset stall counter when activity resumes
                    stall_counter = 0
        
        # Start the watchdog thread
        watchdog_thread = threading.Thread(target=watchdog_timer, daemon=True)
        watchdog_thread.start()
        
        # Function to update the last activity time
        def update_activity_timestamp():
            nonlocal last_activity_time
            last_activity_time = time.time()
            
        # Move to next phase - Router ICMP Check
        current_audit_progress["current_phase"] = "Router ICMP Check"
        AUDIT_PROGRESS["current_phase"] = "Router ICMP Check"
        AUDIT_PROGRESS["status_message"] = "Checking Router Reachability"
        current_audit_progress["status_message"] = "Checking Router Reachability"
        
        # Emit progress update
        try:
            socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
            
        banner_to_log_audit("PHASE 1 – Router Reachability Check")
        update_activity_timestamp()  # Mark activity after phase change
        
        # Check if we have any routers to process
        num_routers_to_ping = len(_INV_ROUTERS)
        if not _INV_ROUTERS:
            log_to_ui_and_console(f"{Fore.YELLOW}No routers in inventory. Skipping ICMP.{Style.RESET_ALL}")
            current_audit_progress["status_message"] = "No routers in inventory"
            audit_status = "Completed (No Routers)"
            return
            
        # Update progress tracking
        reachability_from_jump = {}
        current_audit_progress["total_devices_to_process"] = num_routers_to_ping
        AUDIT_PROGRESS["total_devices"] = num_routers_to_ping
        AUDIT_PROGRESS["status_message"] = f"Found {num_routers_to_ping} routers in inventory"
        log_to_ui_and_console(f"Found {num_routers_to_ping} routers in inventory to process.")
        
        # Sync progress with UI
        sync_progress_with_ui()
        
        # Emit progress update
        try:
            socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
        
        # Iterate through the routers based on the inventory structure
        router_items = []
        try:
            # Handle the CSV inventory data structure as per memory
            if isinstance(_INV_ROUTERS, dict):
                # Old structure (dictionary-based)
                router_items = list(_INV_ROUTERS.items())
                log_to_ui_and_console(f"Using dictionary-based inventory with {len(router_items)} routers")
            elif isinstance(_INV_ROUTERS, list):
                # New structure (CSV-based) - ACTIVE_INVENTORY_DATA['data'] contains list of dictionaries
                # Convert to format compatible with existing code
                log_to_ui_and_console(f"Using CSV-based inventory with {len(_INV_ROUTERS)} routers")
                
                # Create a compatible structure: [(hostname, data_dict), ...]
                router_items = []
                for router_dict in _INV_ROUTERS:
                    hostname = router_dict.get("hostname", "Unknown")
                    router_items.append((hostname, router_dict))
                
                log_to_ui_and_console(f"Converted {len(router_items)} routers from CSV format")
                update_activity_timestamp()  # Mark activity after conversion
            else:
                # This should not happen with the current code, but just in case
                log_to_ui_and_console(f"Warning: Unexpected inventory structure type: {type(_INV_ROUTERS)}")
                raise ValueError(f"Unsupported inventory data structure: {type(_INV_ROUTERS)}")
        except Exception as e:
            log_to_ui_and_console(f"Error processing inventory structure: {e}")
            audit_status = f"Failed: Error processing inventory structure: {e}"
            return
            
        for i, (r_name, r_data) in enumerate(router_items, 1):
            # Update both progress tracking structures
            current_audit_progress.update({"current_device_hostname": r_name, "devices_processed_count": i, "percentage_complete": (i / num_routers_to_ping) * 100 if num_routers_to_ping > 0 else 0, "status_message": f"Pinging {r_name} ({i}/{num_routers_to_ping})..."})
            
            AUDIT_PROGRESS["current_device"] = r_name
            AUDIT_PROGRESS["status_message"] = f"Pinging {r_name} ({i}/{num_routers_to_ping})..."
            
            # Emit progress update
            try:
                socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
            except Exception as e:
                print(f"Error emitting progress update: {e}")
            
            # Initialize detailed_reports_manifest entry for the device
            # Use r_name as key, hostname might change or not be unique initially
            device_report_folder_slug = r_data.get("hostname", r_name).replace(' ', '_').lower() # Slug for device's own folder if it has one
            detailed_reports_manifest[r_name] = {
                "folder": device_report_folder_slug, # Top-level report folder name for this device
                "status": "Pending ICMP Check",
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "device_ip": r_data.get("ip"),
                "files": {
                    "a": None, # Placeholder for report A
                    "b": None, # Placeholder for report B (main JSON audit data)
                    "c": None  # Placeholder for report C (parsed/violations)
                }
            }

            # Check for pause/stop with enhanced progress tracking
            while AUDIT_SHOULD_PAUSE and not AUDIT_SHOULD_STOP:
                pause_audit_progress()
                log_to_ui_and_console(f"{Fore.YELLOW}Audit PAUSED before pinging {r_name}. Waiting...{Style.RESET_ALL}")
                current_audit_progress["status_message"] = f"Paused before {r_name}"
                time.sleep(1)
                if not audit_pause_event.is_set(): audit_pause_event.wait()
            
            if AUDIT_SHOULD_STOP:
                log_to_ui_and_console(f"{Fore.RED}Audit STOPPED before pinging {r_name}.{Style.RESET_ALL}")
                audit_status = "Stopped"
                return
                
            resume_audit_progress()
            log_to_ui_and_console(f"{Fore.GREEN}Audit RESUMED.{Style.RESET_ALL}")
            current_audit_progress["status_message"] = f"Resuming ping for {r_name}"
            
            # Update enhanced progress tracking for current device
            update_audit_progress(device=r_name)
            update_activity_timestamp()  # Mark activity before IP check
            
            ip = r_data.get("ip")
            if not ip: 
                log_to_ui_and_console(f"{Fore.YELLOW}Skipping {r_name}: IP missing.{Style.RESET_ALL}"); 
                reachability_from_jump[r_name] = False; 
                current_run_failures[r_name] = "NO_IP_DEFINED"; 
                last_run_summary_data["per_router_status"][r_name] = "Skipped (No IP)"; 
                last_run_summary_data["failed_icmp"] += 1
                detailed_reports_manifest[r_name]["status"] = "Skipped (No IP)" # Update status in manifest
                update_activity_timestamp()  # Mark activity after skipping
                continue
            log_to_ui_and_console(f"Pinging {r_name} ({ip})... ", end="")
            
            # Update activity timestamp before ping
            update_activity_timestamp()
            
            # Execute ping with proper error handling and a timeout guard
            ping_start_time = time.time()
            ping_max_time = 15  # Maximum time to wait for ping in seconds
            
            # Use a thread to run the ping with a timeout
            ping_result = [False]  # Use a list to store result from thread
            ping_completed = [False]  # Flag to indicate thread completion
            
            def execute_ping_with_timeout():
                try:
                    ping_result[0] = ping_remote(jump_client, ip)
                    ping_completed[0] = True
                except Exception as e:
                    log_to_ui_and_console(
                        f"Ping thread error for {r_name} ({ip}): {str(e)}",
                        log_level="ERROR",
                        phase="ROUTER_ICMP"
                    )
                    ping_completed[0] = True
            
            # Start ping in a separate thread
            ping_thread = threading.Thread(target=execute_ping_with_timeout)
            ping_thread.daemon = True
            ping_thread.start()
            
            # Wait for ping to complete with timeout
            while not ping_completed[0] and (time.time() - ping_start_time) < ping_max_time:
                time.sleep(0.5)
                # Check if audit should stop during ping
                if AUDIT_SHOULD_STOP:
                    log_to_ui_and_console(
                        f"Audit stopped during ping for {r_name}",
                        log_level="WARNING",
                        phase="ROUTER_ICMP"
                    )
                    break
            
            # If ping didn't complete within timeout, mark as failed
            if not ping_completed[0]:
                log_to_ui_and_console(
                    f"Ping operation timed out after {ping_max_time}s for {r_name} ({ip})",
                    log_level="ERROR",
                    phase="ROUTER_ICMP"
                )
                ping_result[0] = False
            
            # Store the result
            ok = ping_result[0]
            reachability_from_jump[r_name] = ok
            
            # Update activity timestamp after ping
            update_activity_timestamp()
            
            if not ok: 
                current_run_failures[r_name] = "ICMP_FAIL_FROM_JUMP"
                last_run_summary_data["per_router_status"][r_name] = "ICMP Failed"
                last_run_summary_data["failed_icmp"] += 1
                detailed_reports_manifest[r_name]["status"] = "ICMP Failed" # Update status in manifest
                
                # Log detailed failure information
                log_to_ui_and_console(
                    f"ICMP ping failed for {r_name} ({ip}). Skipping further tests for this device.", 
                    log_level="WARNING",
                    phase="ROUTER_ICMP"
                )
                
                try:
                    # Update enhanced progress tracking with failure status
                    update_audit_progress(device=r_name, status="failure", completed=True)
                except Exception as e:
                    log_to_ui_and_console(f"Error updating progress for {r_name}: {e}")
                
                # Update activity timestamp after handling failure
                update_activity_timestamp()
                continue
            else: 
                last_run_summary_data["icmp_reachable"] += 1
                detailed_reports_manifest[r_name]["status"] = "ICMP OK" # Update status in manifest
                update_audit_progress(device=r_name, status="icmp_ok") # Real-time update
                # Not marking as completed yet since we still need to do SSH and collection
            ping_result_mark = strip_ansi(mark_audit(ok))
            if ui_logs and ui_logs[-1].endswith(f"Pinging {r_name} ({ip})... "): ui_logs[-1] = ui_logs[-1] + ping_result_mark
            else: log_to_ui_and_console(f"Ping result for {r_name} ({ip}): {ping_result_mark}", console_only=True)
            print(mark_audit(ok), flush=True)
            if num_routers_to_ping > 0: log_to_ui_and_console(bar_audit(i * 100.0 / num_routers_to_ping), end="\r")
        if sys.stdout.isatty() and num_routers_to_ping > 0: print(); current_audit_progress["percentage_complete"] = 100
        if num_routers_to_ping > 0 and not any(reachability_from_jump.values()):
            msg = "No routers ICMP reachable from jump – aborting."; log_to_ui_and_console(f"{Fore.RED}{msg}{Style.RESET_ALL}")
            for r_name_iter_icmp in _INV_ROUTERS:
                if last_run_summary_data["per_router_status"][r_name_iter_icmp] == "Pending": last_run_summary_data["per_router_status"][r_name_iter_icmp] = "Skipped (No ICMP path)"
            audit_status = f"Failed: {msg}"; current_audit_progress["status_message"] = audit_status; raise AuditError(msg)
        if not audit_pause_event.is_set(): log_to_ui_and_console(f"{Fore.YELLOW}Audit PAUSED. Waiting...{Style.RESET_ALL}"); current_audit_progress["status_message"] = "Paused"; audit_pause_event.wait(); log_to_ui_and_console(f"{Fore.GREEN}Audit RESUMED.{Style.RESET_ALL}")
        current_audit_progress["current_phase"] = "Router SSH Authentication (via Jump)"
        AUDIT_PROGRESS["current_phase"] = "Router SSH Authentication (via Jump)"
        AUDIT_PROGRESS["status_message"] = "Testing SSH Authentication"
        current_audit_progress["status_message"] = "Testing SSH Authentication"
        
        # Emit progress update
        try:
            socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
            
        banner_to_log_audit("PHASE 1.5 – Router SSH Authentication Test (via jump host)")
        ssh_auth_ok = {}; routers_to_test_ssh = [r for r, reachable in reachability_from_jump.items() if reachable]; num_to_test_ssh = len(routers_to_test_ssh)
        current_audit_progress["total_devices_to_process"] = num_to_test_ssh
        for i, r_name_ssh in enumerate(routers_to_test_ssh, 1):
            current_audit_progress.update({"current_device_hostname": r_name_ssh, "devices_processed_count": i, "percentage_complete": (i / num_to_test_ssh) * 100 if num_to_test_ssh > 0 else 0, "status_message": f"Testing SSH Auth to {r_name_ssh} ({i}/{num_to_test_ssh})..."})
            
            # Update status for SSH phase in manifest
            if r_name_ssh in detailed_reports_manifest: # Should always be true if ICMP phase ran
                detailed_reports_manifest[r_name_ssh]["status"] = "Pending SSH Auth"

            if not audit_pause_event.is_set(): log_to_ui_and_console(f"{Fore.YELLOW}Audit PAUSED before SSH to {r_name_ssh}. Waiting...{Style.RESET_ALL}"); current_audit_progress["status_message"] = f"Paused before {r_name_ssh}"; audit_pause_event.wait(); log_to_ui_and_console(f"{Fore.GREEN}Audit RESUMED.{Style.RESET_ALL}"); current_audit_progress["status_message"] = f"Resuming SSH to {r_name_ssh}"
            r_config_ssh = _INV_ROUTERS[r_name_ssh]; ip_ssh = r_config_ssh["ip"]; username_ssh = r_config_ssh.get("username", _DEVICE_USERNAME); password_ssh = r_config_ssh.get("password", _DEVICE_PASSWORD)
            log_to_ui_and_console(f"Testing SSH to {r_name_ssh} ({ip_ssh}) with user '{sanitize_log_message(username_ssh)}'... ", end="", is_sensitive=True)
            chan = None; test_ssh_client = None
            try:
                # Verbose logging for SSH connection attempt
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] Opening direct-tcpip channel to {ip_ssh}:22...")
                chan = jump_client.get_transport().open_channel("direct-tcpip", (ip_ssh, 22), ("127.0.0.1", 0), timeout=10)
                
                if chan is None:
                    log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] Failed to open direct-tcpip channel to device.")
                    raise paramiko.SSHException("Failed to open direct-tcpip channel to device.")
                    
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] Channel opened successfully. Attempting SSH connection...")
                test_ssh_client = paramiko.SSHClient()
                test_ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                
                # Verbose logging for SSH authentication
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] Authenticating with username '{sanitize_log_message(username_ssh)}'...")
                test_ssh_client.connect(
                    ip_ssh, 22, username_ssh, password_ssh, 
                    sock=chan, allow_agent=False, look_for_keys=False, 
                    timeout=15, banner_timeout=20
                )
                
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] SSH authentication successful!")
                ssh_auth_ok[r_name_ssh] = True
                last_run_summary_data["ssh_auth_ok"] += 1
                last_run_summary_data["per_router_status"][r_name_ssh] = "SSH Auth OK"
                update_audit_progress(device=r_name_ssh, status="ssh_auth_ok") # Real-time update
                current_run_failures[r_name_ssh] = None
                print(mark_audit(True), flush=True)
                
                if r_name_ssh in detailed_reports_manifest:
                    detailed_reports_manifest[r_name_ssh]["status"] = "SSH Auth OK"
                    
                if ui_logs and ui_logs[-1].endswith(f"'{sanitize_log_message(username_ssh)}'... "):
                    ui_logs[-1] = ui_logs[-1] + strip_ansi(mark_audit(True))
            except paramiko.AuthenticationException as e_auth:
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] Authentication failed: {e_auth}")
                ssh_auth_ok[r_name_ssh] = False
                current_run_failures[r_name_ssh] = f"SSH_AUTH_FAIL: Authentication failed"
                last_run_summary_data["per_router_status"][r_name_ssh] = f"SSH Auth Failed (Authentication)"
                last_run_summary_data["failed_ssh_auth"] += 1
                update_audit_progress(device=r_name_ssh, status="failure", completed=True) # Real-time update
                print(mark_audit(False), flush=True)
                if r_name_ssh in detailed_reports_manifest:
                    detailed_reports_manifest[r_name_ssh]["status"] = f"SSH Auth Failed (Authentication)"
                if ui_logs and ui_logs[-1].endswith(f"'{sanitize_log_message(username_ssh)}'... "):
                    ui_logs[-1] = ui_logs[-1] + strip_ansi(mark_audit(False)) + sanitize_log_message(f" Error: Authentication failed")
            except socket.timeout as e_timeout:
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] Connection timed out: {e_timeout}")
                ssh_auth_ok[r_name_ssh] = False
                current_run_failures[r_name_ssh] = f"SSH_AUTH_FAIL: Connection timeout"
                last_run_summary_data["per_router_status"][r_name_ssh] = f"SSH Auth Failed (Timeout)"
                last_run_summary_data["failed_ssh_auth"] += 1
                update_audit_progress(device=r_name_ssh, status="failure", completed=True) # Real-time update
                print(mark_audit(False), flush=True)
                if r_name_ssh in detailed_reports_manifest:
                    detailed_reports_manifest[r_name_ssh]["status"] = f"SSH Auth Failed (Timeout)"
                if ui_logs and ui_logs[-1].endswith(f"'{sanitize_log_message(username_ssh)}'... "):
                    ui_logs[-1] = ui_logs[-1] + strip_ansi(mark_audit(False)) + sanitize_log_message(f" Error: Connection timeout")
            except paramiko.SSHException as e_ssh:
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] SSH error: {e_ssh}")
                ssh_auth_ok[r_name_ssh] = False
                current_run_failures[r_name_ssh] = f"SSH_AUTH_FAIL: SSH protocol error"
                last_run_summary_data["per_router_status"][r_name_ssh] = f"SSH Auth Failed (Protocol)"
                last_run_summary_data["failed_ssh_auth"] += 1
                update_audit_progress(device=r_name_ssh, status="failure", completed=True) # Real-time update
                print(mark_audit(False), flush=True)
                if r_name_ssh in detailed_reports_manifest:
                    detailed_reports_manifest[r_name_ssh]["status"] = f"SSH Auth Failed (Protocol)"
                if ui_logs and ui_logs[-1].endswith(f"'{sanitize_log_message(username_ssh)}'... "):
                    ui_logs[-1] = ui_logs[-1] + strip_ansi(mark_audit(False)) + sanitize_log_message(f" Error: SSH protocol error")
            except Exception as e_ssh_auth:
                log_to_ui_and_console(f"[ROUTER:{r_name_ssh}] Unexpected error: {type(e_ssh_auth).__name__}: {e_ssh_auth}")
                ssh_auth_ok[r_name_ssh] = False
                current_run_failures[r_name_ssh] = f"SSH_AUTH_FAIL: {type(e_ssh_auth).__name__}"
                last_run_summary_data["per_router_status"][r_name_ssh] = f"SSH Auth Failed ({type(e_ssh_auth).__name__})"
                last_run_summary_data["failed_ssh_auth"] += 1
                update_audit_progress(device=r_name_ssh, status="failure", completed=True) # Real-time update
                print(mark_audit(False), flush=True)
                if r_name_ssh in detailed_reports_manifest:
                    detailed_reports_manifest[r_name_ssh]["status"] = f"SSH Auth Failed ({type(e_ssh_auth).__name__})"
                if ui_logs and ui_logs[-1].endswith(f"'{sanitize_log_message(username_ssh)}'... "):
                    ui_logs[-1] = ui_logs[-1] + strip_ansi(mark_audit(False)) + sanitize_log_message(f" Error: {type(e_ssh_auth).__name__}")

            finally:
                if test_ssh_client: test_ssh_client.close()
            if num_to_test_ssh > 0: log_to_ui_and_console(bar_audit(i * 100.0 / num_to_test_ssh), end="\r")
        if sys.stdout.isatty() and num_to_test_ssh > 0: print(); current_audit_progress["percentage_complete"] = 100
        ready_routers = [r_ready for r_ready in _INV_ROUTERS if reachability_from_jump.get(r_ready) and ssh_auth_ok.get(r_ready)]
        if num_routers_to_ping > 0 and not ready_routers:
            msg = "No devices passed SSH auth – aborting collection."; log_to_ui_and_console(f"{Fore.RED}{msg}{Style.RESET_ALL}")
            for r_name_iter_ssh_fail in routers_to_test_ssh:
                if not ssh_auth_ok.get(r_name_iter_ssh_fail) and last_run_summary_data["per_router_status"][r_name_iter_ssh_fail] not in ["ICMP Failed", "Skipped (No IP)"]: last_run_summary_data["per_router_status"][r_name_iter_ssh_fail] = "SSH Auth Failed"
            audit_status = f"Failed: {msg}"; current_audit_progress["status_message"] = audit_status; raise AuditError(msg)
        if not audit_pause_event.is_set(): log_to_ui_and_console(f"{Fore.YELLOW}Audit PAUSED. Waiting...{Style.RESET_ALL}"); current_audit_progress["status_message"] = "Paused"; audit_pause_event.wait(); log_to_ui_and_console(f"{Fore.GREEN}Audit RESUMED.{Style.RESET_ALL}")
        current_audit_progress["current_phase"] = "Data Collection & Audit"
        AUDIT_PROGRESS["current_phase"] = "Data Collection & Audit"
        AUDIT_PROGRESS["status_message"] = "Collecting Data from Routers"
        current_audit_progress["status_message"] = "Collecting Data from Routers"
        
        # Emit progress update
        try:
            socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
            
        banner_to_log_audit("PHASE 2 – Data Collection & Audit (Physical Line Telnet Check)")
        num_ready_routers = len(ready_routers); current_audit_progress["total_devices_to_process"] = num_ready_routers; collection_done_count = 0
        for r_name_collect in ready_routers:
            collection_done_count += 1; current_audit_progress.update({"current_device_hostname": r_name_collect, "devices_processed_count": collection_done_count, "percentage_complete": (collection_done_count / num_ready_routers) * 100 if num_ready_routers > 0 else 0, "status_message": f"Collecting from {r_name_collect} ({(collection_done_count)}/{num_ready_routers})..."})
            
            # Update status for Collection phase in manifest
            if r_name_collect in detailed_reports_manifest: # Should always be true
                detailed_reports_manifest[r_name_collect]["status"] = "Pending Data Collection"

            if not audit_pause_event.is_set(): log_to_ui_and_console(f"{Fore.YELLOW}Audit PAUSED before collecting from {r_name_collect}. Waiting...{Style.RESET_ALL}"); current_audit_progress["status_message"] = f"Paused before {r_name_collect}"; audit_pause_event.wait(); log_to_ui_and_console(f"{Fore.GREEN}Audit RESUMED.{Style.RESET_ALL}"); current_audit_progress["status_message"] = f"Resuming collection for {r_name_collect}"
            log_to_ui_and_console(f"{Fore.CYAN}{r_name_collect} → Attempting data collection...{Style.RESET_ALL}")
            r_config_collect = _INV_ROUTERS[r_name_collect]; ip_addr_collect = r_config_collect["ip"]; username_collect = r_config_collect.get("username", _DEVICE_USERNAME); password_collect = r_config_collect.get("password", _DEVICE_PASSWORD); secret_collect = r_config_collect.get("secret", _DEVICE_ENABLE); device_type_collect = r_config_collect.get("device_type", "cisco_ios")
            net_connect = None
            try:
                # First try with Netmiko
                try:
                    conn_chan = jump_client.get_transport().open_channel("direct-tcpip", (ip_addr_collect, 22), ("127.0.0.1", 0), timeout=10)
                    if conn_chan is None: raise paramiko.SSHException("Failed to open Netmiko direct-tcpip channel for data collection.")
                    
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Attempting connection with Netmiko...")
                    netmiko_params = {'device_type': device_type_collect, 'ip': ip_addr_collect, 'username': username_collect, 'password': password_collect, 'secret': secret_collect, 'sock': conn_chan, 'ssh_config_file': ssh_cfg_filename, 'fast_cli': False, 'global_delay_factor': 2, 'conn_timeout': 30, 'banner_timeout': 30, 'auth_timeout': 30}
                    net_connect = ConnectHandler(**netmiko_params)
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Netmiko connection successful")
                except (NetmikoTimeoutException, NetmikoAuthenticationException) as e_netmiko:
                    # If Netmiko fails, try with Paramiko
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Netmiko connection failed: {type(e_netmiko).__name__} - {e_netmiko}. Trying with Paramiko...")
                    
                    try:
                        # Create a new channel for Paramiko
                        conn_chan = jump_client.get_transport().open_channel("direct-tcpip", (ip_addr_collect, 22), ("127.0.0.1", 0), timeout=10)
                        if conn_chan is None: raise paramiko.SSHException("Failed to open Paramiko direct-tcpip channel for data collection.")
                        
                        # Set up Paramiko client
                        router_client = paramiko.SSHClient()
                        router_client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                        
                        # Connect with Paramiko
                        router_client.connect(
                            hostname=ip_addr_collect,
                            username=username_collect,
                            password=password_collect,
                            sock=conn_chan,
                            timeout=30,
                            allow_agent=False,
                            look_for_keys=False
                        )
                        
                        # Create a custom wrapper to make Paramiko behave like Netmiko
                        class ParamikoWrapper:
                            def __init__(self, client):
                                self.client = client
                                self.device_type = device_type_collect
                            
                            def send_command(self, command, expect_string=None, delay_factor=1, max_loops=500, strip_prompt=True, strip_command=True):
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Executing with Paramiko: {command}")
                                stdin, stdout, stderr = self.client.exec_command(command)
                                output = stdout.read().decode('utf-8')
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Command completed with Paramiko")
                                return output
                            
                            def enable(self):
                                # For Paramiko, we'd need to implement enable mode differently
                                # This is a simplified version
                                if secret_collect:
                                    self.send_command("enable")
                                    self.send_command(secret_collect)
                            
                            def disconnect(self):
                                self.client.close()
                        
                        net_connect = ParamikoWrapper(router_client)
                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Paramiko connection successful")
                    except Exception as e_paramiko:
                        # If both fail, re-raise the original Netmiko exception
                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Paramiko connection also failed: {type(e_paramiko).__name__} - {e_paramiko}")
                        raise e_netmiko
                if secret_collect: net_connect.enable()
                log_to_ui_and_console(f"  {mark_audit(True)} Successfully connected to {r_name_collect}."); last_run_summary_data["per_router_status"][r_name_collect] = "Collection in Progress"; current_audit_progress["current_device_hostname"] = f"{r_name_collect} (Connected)"
                if r_name_collect in detailed_reports_manifest: detailed_reports_manifest[r_name_collect]["status"] = "Collection in Progress"
                # Verbose logging for router commands
                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Executing: terminal length 0")
                try:
                    output = net_connect.send_command("terminal length 0", expect_string=r"#")
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Set terminal length 0 successful")
                    if output:
                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Command output:\n{output[:200]}{'...' if len(output) > 200 else ''}")
                except Exception as e:
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Error setting terminal length: {e}")
                
                # Hostname is now derived from connection info
                hostname = r_name_collect
                
                # Initialize command output variables
                show_line_cmd_output = ""; show_run_line_0_output = ""
                
                
                
                # Get show line output with verbose logging
                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Executing: show line")
                try:
                    show_line_cmd_output = net_connect.send_command("show line", read_timeout=60)
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Successfully retrieved 'show line' output")
                    if show_line_cmd_output:
                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Command output:\n{show_line_cmd_output[:500]}{'...' if len(show_line_cmd_output) > 500 else ''}")
                except Exception as e_cmd_show_line:
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.YELLOW}Warning:{Style.RESET_ALL} Could not retrieve 'show line' output: {str(e_cmd_show_line)}")
                    # Get show run | section 'line 0/' output with verbose logging
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Executing: show run | section 'line 0/'")
                    show_run_line_0_output = ''
                    try:
                        show_run_line_0_output = net_connect.send_command("show run | section 'line 0/'", read_timeout=120)
                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Successfully retrieved 'show run | section \'line 0/\' output")
                        if show_run_line_0_output:
                            log_to_ui_and_console(f"[ROUTER:{r_name_collect}] Command output:\n{show_run_line_0_output[:500]}{'...' if len(show_run_line_0_output) > 500 else ''}")
                    except Exception as e_cmd_run_line_0:
                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.YELLOW}Warning:{Style.RESET_ALL} Could not retrieve 'show run | section \'line 0/\' output: {str(e_cmd_run_line_0)}")
                violations_count_for_this_router = 0; physical_line_telnet_violations_details = []; router_has_physical_line_violation = False
                current_physical_line = None; physical_line_buffer = []
                target_line_pattern = re.compile(r"^\s*line\s+(?!(?:con|aux|vty)\b)(\d+(?:/\d+)*(?:\s+\d+)?)\s*$")
                
                # If show_run_line_0_output is empty or only contains whitespace, it's a PASS (no physical lines configured)
                if not show_run_line_0_output or show_run_line_0_output.strip() == "":
                    log_to_ui_and_console(f"[ROUTER:{r_name_collect}] No physical line configurations found - PASS")
                
                # Process the output lines to check for telnet security issues
                config_lines_to_parse = (show_run_line_0_output or "").splitlines()
                for line_cfg in config_lines_to_parse:
                    stripped_line_cfg = line_cfg.strip(); match = target_line_pattern.match(stripped_line_cfg)
                    if match:
                        if current_physical_line and physical_line_buffer:
                            line_cfg_block_text = "\n".join(physical_line_buffer); line_is_telnet_open = False; line_failure_reason = "N/A"
                            transport_input_lines_for_block = [l.strip().lower() for l in physical_line_buffer if "transport input" in l.lower()]
                            if not transport_input_lines_for_block: 
                                # For physical lines, no transport input might be secure by default in modern IOS
                                # But we'll still log it as a potential issue to investigate
                                line_is_telnet_open = True
                                line_failure_reason = "default_telnet_no_transport_input"
                                last_run_summary_data["failure_category_default_telnet"] +=1
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.YELLOW}Warning:{Style.RESET_ALL} Line {current_physical_line} has no 'transport input' command - potential default telnet")
                            else:
                                for ti_line in transport_input_lines_for_block:
                                    if "telnet" in ti_line: 
                                        line_is_telnet_open = True
                                        line_failure_reason = "explicit_telnet"
                                        last_run_summary_data["failure_category_explicit_telnet"] +=1
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} explicitly allows telnet")
                                        break 
                                    if "all" in ti_line: 
                                        line_is_telnet_open = True
                                        line_failure_reason = "transport_all_keyword_present"
                                        last_run_summary_data["failure_category_transport_all"] +=1
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} allows all transport protocols (including telnet)")
                                        break
                                    if ("ssh" in ti_line and "telnet" not in ti_line and "all" not in ti_line) or "none" in ti_line: 
                                        line_is_telnet_open = False
                                        line_failure_reason = "PASS_SSH_OR_NONE"
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.GREEN}PASS:{Style.RESET_ALL} Line {current_physical_line} only allows SSH or has transport input none")
                                        break 
                            if line_is_telnet_open and line_failure_reason not in ["PASS_SSH_OR_NONE"]: router_has_physical_line_violation = True; physical_line_telnet_violations_details.append({"line_id": current_physical_line, "reason": line_failure_reason, "config_snippet": line_cfg_block_text})
                        current_physical_line = match.group(1); physical_line_buffer = [stripped_line_cfg]
                    elif current_physical_line and (line_cfg.startswith(" ") or line_cfg.startswith("\t")): physical_line_buffer.append(stripped_line_cfg)
                
                # Process the last line block if we've reached the end of the config
                if current_physical_line and physical_line_buffer and line_cfg == config_lines_to_parse[-1]:
                    line_cfg_block_text = "\n".join(physical_line_buffer)
                    line_is_telnet_open = False
                    line_failure_reason = "N/A"
                    
                    transport_input_lines_for_block = [l.strip().lower() for l in physical_line_buffer if "transport input" in l.lower()]
                    
                    if not transport_input_lines_for_block:
                        line_is_telnet_open = True
                        line_failure_reason = "default_telnet_no_transport_input"
                        last_run_summary_data["failure_category_default_telnet"] +=1
                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.YELLOW}Warning:{Style.RESET_ALL} Line {current_physical_line} has no 'transport input' command - potential default telnet")
                    else:
                        for ti_line in transport_input_lines_for_block:
                            if "telnet" in ti_line:
                                line_is_telnet_open = True
                                line_failure_reason = "explicit_telnet"
                                last_run_summary_data["failure_category_explicit_telnet"] +=1
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} explicitly allows telnet")
                                break
                            if "all" in ti_line:
                                line_is_telnet_open = True
                                line_failure_reason = "transport_all_keyword_present"
                                last_run_summary_data["failure_category_transport_all"] +=1
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} allows all transport protocols (including telnet)")
                                break
                            if ("ssh" in ti_line and "telnet" not in ti_line and "all" not in ti_line) or "none" in ti_line:
                                line_is_telnet_open = False
                                line_failure_reason = "PASS_SSH_OR_NONE"
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.GREEN}PASS:{Style.RESET_ALL} Line {current_physical_line} only allows SSH or has transport input none")
                                break
                    
                    if line_is_telnet_open and line_failure_reason not in ["PASS_SSH_OR_NONE"]:
                        router_has_physical_line_violation = True
                        physical_line_telnet_violations_details.append({
                            "line_id": current_physical_line,
                            "reason": line_failure_reason,
                            "config_snippet": line_cfg_block_text
                        })
                    elif current_physical_line and not stripped_line_cfg.startswith("line "): physical_line_buffer.append(stripped_line_cfg)
                    elif stripped_line_cfg.lower().startswith(("line con", "line aux", "line vty")):
                        if current_physical_line and physical_line_buffer:
                            line_cfg_block_text = "\n".join(physical_line_buffer); line_is_telnet_open = False; line_failure_reason = "N/A"
                            transport_input_lines_for_block = [l.strip().lower() for l in physical_line_buffer if "transport input" in l.lower()]
                            if not transport_input_lines_for_block: 
                                # For physical lines, no transport input might be secure by default in modern IOS
                                # But we'll still log it as a potential issue to investigate
                                line_is_telnet_open = True
                                line_failure_reason = "default_telnet_no_transport_input"
                                last_run_summary_data["failure_category_default_telnet"] +=1
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.YELLOW}Warning:{Style.RESET_ALL} Line {current_physical_line} has no 'transport input' command - potential default telnet")
                            else:
                                for ti_line in transport_input_lines_for_block:
                                    if "telnet" in ti_line: 
                                        line_is_telnet_open = True
                                        line_failure_reason = "explicit_telnet"
                                        last_run_summary_data["failure_category_explicit_telnet"] +=1
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} explicitly allows telnet")
                                        break
                                    if "all" in ti_line: 
                                        line_is_telnet_open = True
                                        line_failure_reason = "transport_all_keyword_present"
                                        last_run_summary_data["failure_category_transport_all"] +=1
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} allows all transport protocols (including telnet)")
                                        break
                                    if ("ssh" in ti_line and "telnet" not in ti_line and "all" not in ti_line) or "none" in ti_line: 
                                        line_is_telnet_open = False
                                        line_failure_reason = "PASS_SSH_OR_NONE"
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.GREEN}PASS:{Style.RESET_ALL} Line {current_physical_line} only allows SSH or has transport input none")
                                        break
                            if line_is_telnet_open and line_failure_reason not in ["PASS_SSH_OR_NONE"]: router_has_physical_line_violation = True; physical_line_telnet_violations_details.append({"line_id": current_physical_line, "reason": line_failure_reason, "config_snippet": line_cfg_block_text})
                        current_physical_line = None; physical_line_buffer = []
                if current_physical_line and physical_line_buffer:
                    line_cfg_block_text = "\n".join(physical_line_buffer); line_is_telnet_open = False; line_failure_reason = "N/A"
                    transport_input_lines_for_block = [l.strip().lower() for l in physical_line_buffer if "transport input" in l.lower()]
                    if not transport_input_lines_for_block: 
                                # For physical lines, no transport input might be secure by default in modern IOS
                                # But we'll still log it as a potential issue to investigate
                                line_is_telnet_open = True
                                line_failure_reason = "default_telnet_no_transport_input"
                                last_run_summary_data["failure_category_default_telnet"] +=1
                                log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.YELLOW}Warning:{Style.RESET_ALL} Line {current_physical_line} has no 'transport input' command - potential default telnet")
                    else:
                        for ti_line in transport_input_lines_for_block:
                            if "telnet" in ti_line: 
                                        line_is_telnet_open = True
                                        line_failure_reason = "explicit_telnet"
                                        last_run_summary_data["failure_category_explicit_telnet"] +=1
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} explicitly allows telnet")
                                        break 
                            if "all" in ti_line: 
                                        line_is_telnet_open = True
                                        line_failure_reason = "transport_all_keyword_present"
                                        last_run_summary_data["failure_category_transport_all"] +=1
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.RED}FAIL:{Style.RESET_ALL} Line {current_physical_line} allows all transport protocols (including telnet)")
                                        break
                            if ("ssh" in ti_line and "telnet" not in ti_line and "all" not in ti_line) or "none" in ti_line: 
                                        line_is_telnet_open = False
                                        line_failure_reason = "PASS_SSH_OR_NONE"
                                        log_to_ui_and_console(f"[ROUTER:{r_name_collect}] {Fore.GREEN}PASS:{Style.RESET_ALL} Line {current_physical_line} only allows SSH or has transport input none")
                                        break
                    if line_is_telnet_open and line_failure_reason not in ["PASS_SSH_OR_NONE"]: router_has_physical_line_violation = True; physical_line_telnet_violations_details.append({"line_id": current_physical_line, "reason": line_failure_reason, "config_snippet": line_cfg_block_text})
                violations_count_for_this_router = len(physical_line_telnet_violations_details)
                if router_has_physical_line_violation: last_run_summary_data["with_violations"] += 1; last_run_summary_data["total_physical_line_issues"] += violations_count_for_this_router
                report_b_data = {"hostname": hostname, "ip_address": ip_addr_collect, "total_physical_line_violations": violations_count_for_this_router, "violation_details": physical_line_telnet_violations_details, "full_show_run_section_line": show_run_section_line_cmd_output, "full_show_line_output": show_line_cmd_output}
                report_b_path = os.path.join(b_dir, f"{hostname.lower().replace(' ', '_')}_audit_details_b.json"); detailed_reports_manifest.setdefault(r_name_collect, {})["b"] = os.path.basename(report_b_path)
                with open(report_b_path, "w") as f_report_b: json.dump(report_b_data, f_report_b, indent=2)
                opt_a_data = {"hostname": hostname, "ip_address": ip_addr_collect, "show_run_section_line_config": show_run_section_line_cmd_output, "show_line_output": show_line_cmd_output}
                report_a_path = os.path.join(a_dir, f"{hostname.lower().replace(' ', '_')}_raw_data_a.json"); detailed_reports_manifest[r_name_collect]["a"] = os.path.basename(report_a_path)
                with open(report_a_path, "w") as f_report_a: json.dump(opt_a_data, f_report_a, indent=2)
                opt_c_data = opt_c_collect(net_connect, hostname, device_type_collect, show_line_cmd_output); opt_c_data["hostname"] = hostname; opt_c_data["ip_address"] = ip_addr_collect
                report_c_path = os.path.join(c_dir, f"{hostname.lower().replace(' ', '_')}_parsed_line_c.json"); detailed_reports_manifest[r_name_collect]["c"] = os.path.basename(report_c_path)
                with open(report_c_path, "w") as f_report_c: json.dump(opt_c_data, f_report_c, indent=2)
                last_run_summary_data["collected"] += 1; log_to_ui_and_console(f"  {mark_audit(violations_count_for_this_router == 0)} Collection for {hostname} successful. Physical Line Telnet Violations: {violations_count_for_this_router}")
                # Determine status for AUDIT_PROGRESS based on violations
                device_final_status = "success" if violations_count_for_this_router == 0 else "warning"
                update_audit_progress(device=r_name_collect, status=device_final_status, completed=True) # Real-time update
                last_run_summary_data["per_router_status"][r_name_collect] = f"Collected (as {hostname}), Physical Line Violations: {violations_count_for_this_router}"; current_run_failures[r_name_collect] = None
            except NetmikoTimeoutException as e_nm_timeout: 
                err_msg = f"Netmiko Timeout for {r_name_collect}: {e_nm_timeout}"
                log_to_ui_and_console(f"{Fore.RED}  {mark_audit(False)} {sanitize_log_message(err_msg)}{Style.RESET_ALL}")
                update_audit_progress(device=r_name_collect, status="failure", completed=True) # Real-time update
                current_run_failures[r_name_collect] = f"COLLECTION_FAIL_TIMEOUT: {str(e_nm_timeout)[:150]}"
                last_run_summary_data["per_router_status"][r_name_collect] = "Collection Failed (Timeout)"
                last_run_summary_data["failed_collection"] += 1
            except NetmikoAuthenticationException as e_nm_auth: 
                err_msg = f"Netmiko Auth Error for {r_name_collect}: {e_nm_auth}"
                log_to_ui_and_console(f"{Fore.RED}  {mark_audit(False)} {sanitize_log_message(err_msg)}{Style.RESET_ALL}")
                update_audit_progress(device=r_name_collect, status="failure", completed=True) # Real-time update
                current_run_failures[r_name_collect] = f"COLLECTION_FAIL_AUTH: {str(e_nm_auth)[:150]}"
                last_run_summary_data["per_router_status"][r_name_collect] = "Collection Failed (Auth Error)"
                last_run_summary_data["failed_collection"] += 1
            except Exception as e_collect: 
                err_msg = f"Collection for {r_name_collect} failed: {type(e_collect).__name__} - {e_collect}"
                log_to_ui_and_console(f"{Fore.RED}  {mark_audit(False)} {sanitize_log_message(err_msg)}{Style.RESET_ALL}")
                update_audit_progress(device=r_name_collect, status="failure", completed=True) # Real-time update
                current_run_failures[r_name_collect] = f"COLLECTION_FAIL_GENERIC: {str(e_collect)[:150]}"
                last_run_summary_data["per_router_status"][r_name_collect] = f"Collection Failed ({type(e_collect).__name__})"
                last_run_summary_data["failed_collection"] += 1
            finally:
                if net_connect:
                    try: net_connect.disconnect(); log_to_ui_and_console(f"  Netmiko connection to {r_name_collect} closed.", console_only=True)
                    except Exception as e_disconnect: log_to_ui_and_console(f"  {Fore.YELLOW}Warning: Error disconnecting Netmiko from {r_name_collect}: {e_disconnect}{Style.RESET_ALL}", console_only=True)
            if num_ready_routers > 0: log_to_ui_and_console(bar_audit(collection_done_count * 100.0 / num_ready_routers), end="\r")
        if sys.stdout.isatty() and num_ready_routers > 0: 
            print()
            current_audit_progress["percentage_complete"] = 100
            current_audit_progress["status_message"] = "Collection & Audit Phase Complete"
            AUDIT_PROGRESS["status_message"] = "Collection & Audit Phase Complete"
            AUDIT_PROGRESS["completed_devices"] = AUDIT_PROGRESS["total_devices"]
            
            # Emit progress update
            try:
                socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
            except Exception as e:
                print(f"Error emitting progress update: {e}")
        banner_to_log_audit("Summary of Audit Results") # ... (rest of summary logging)
        # ... (Summary logging and report generation as before)
        with open(summary_txt_file_path, "w") as fh_summary:
            fh_summary.write(f"==== Telnet-Audit Summary Report ====\nTimestamp: {datetime.now().isoformat()}\nActive Inventory File: {APP_CONFIG['ACTIVE_INVENTORY_FILE']} (Format: {APP_CONFIG['ACTIVE_INVENTORY_FORMAT'].upper()})\n\n")
            # ... (rest of summary text file content as before, adding format)
            fh_summary.write(f"Total routers configured: {last_run_summary_data['total_routers']}\nICMP reachable: {last_run_summary_data['icmp_reachable']} (Failed: {last_run_summary_data['failed_icmp']})\nSSH authenticated: {last_run_summary_data['ssh_auth_ok']} (Failed: {last_run_summary_data['failed_ssh_auth']})\nData collected for: {last_run_summary_data['collected']} (Failed: {last_run_summary_data['failed_collection']})\nRouters with Physical Line Telnet violations: {last_run_summary_data['with_violations']}\n")
            fh_summary.write(f"  - Total physical lines with Telnet enabled: {last_run_summary_data['total_physical_line_issues']}\n    - Due to 'transport input telnet': {last_run_summary_data['failure_category_explicit_telnet']}\n    - Due to 'transport input all': {last_run_summary_data['failure_category_transport_all']}\n    - Due to default Telnet (no transport input): {last_run_summary_data['failure_category_default_telnet']}\n\n==== Per-Router Detailed Status ====\n")
            for r_name_iter_summary in _INV_ROUTERS: fh_summary.write(f"{r_name_iter_summary:<20}: {last_run_summary_data['per_router_status'].get(r_name_iter_summary, 'Unknown') + (f' (Reason: {str(current_run_failures.get(r_name_iter_summary))[:200]})' if current_run_failures.get(r_name_iter_summary) else '')}\n")
            fh_summary.write("\n==== Report File Index ====\n" + f"{'Router (Inventory Name)':<25} | {'Report A (Raw Data)':<40} | {'Report B (Audit Detail)':<40} | {'Report C (Parsed Line)':<40}\n" + f"{'-'*25} | {'-'*40} | {'-'*40} | {'-'*40}\n")
            for r_name_iter_manifest, files_dict_manifest in detailed_reports_manifest.items():
                if r_name_iter_manifest != 'summary_file' and r_name_iter_manifest != 'pdf_summary_file':
                    fh_summary.write(f"{r_name_iter_manifest:<25} | {files_dict_manifest.get('a', 'N/A'):<40} | {files_dict_manifest.get('b', 'N/A'):<40} | {files_dict_manifest.get('c', 'N/A'):<40}\n")
        detailed_reports_manifest["summary_file"] = {"folder": "", "filename": os.path.basename(summary_txt_file_path)}
        try:
            current_audit_progress["status_message"] = "Generating PDF Summary Report..."
            AUDIT_PROGRESS["status_message"] = "Generating PDF Summary Report..."
            
            # Emit progress update
            try:
                socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
            except Exception as e:
                print(f"Error emitting progress update: {e}")
                
            log_to_ui_and_console("Generating PDF summary report...")
            pdf_path = generate_pdf_summary_report(last_run_summary_data, detailed_reports_manifest, current_run_failures, _INV_ROUTERS, base_report_path)
            detailed_reports_manifest["pdf_summary_file"] = {"folder": "", "filename": os.path.basename(pdf_path)}
            log_to_ui_and_console(f"PDF summary report saved to {pdf_path}")
            
            # Generate Excel summary report
            log_to_ui_and_console("Generating Excel summary report...")
            excel_path = generate_excel_summary_report(last_run_summary_data, detailed_reports_manifest, current_run_failures, _INV_ROUTERS, base_report_path)
            detailed_reports_manifest["excel_summary_file"] = {"folder": "", "filename": os.path.basename(excel_path)}
            log_to_ui_and_console(f"Excel summary report saved to {excel_path}")
        except Exception as pdf_e: log_to_ui_and_console(f"{Fore.RED}Failed to generate PDF summary: {pdf_e}{Style.RESET_ALL}")
        log_to_ui_and_console(f"\nAudit process finished. Reports in {base_report_path}/")
        audit_status = "Completed"
        current_audit_progress["status_message"] = "Audit Completed"
        current_audit_progress["percentage_complete"] = 100
        
        AUDIT_PROGRESS["status"] = "completed"
        AUDIT_PROGRESS["status_message"] = "Audit Completed"
        AUDIT_PROGRESS["completed_devices"] = AUDIT_PROGRESS["total_devices"]
        AUDIT_PROGRESS["end_time"] = datetime.now()
        
        # Ensure the final progress state is emitted to update the UI
        try:
            socketio.emit('progress_update', prepare_progress_for_json(AUDIT_PROGRESS), namespace='/')
            sync_progress_with_ui() # Make sure the UI is fully synchronized
        except Exception as e:
            print(f"Error emitting final progress update: {e}")
        
        # Emit final progress update
        try:
            socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
    except AuditError as ae: 
        log_to_ui_and_console(f"{Fore.RED}AUDIT HALTED: {ae}{Style.RESET_ALL}")
        current_audit_progress["status_message"] = audit_status
        AUDIT_PROGRESS["status"] = "failed"
        AUDIT_PROGRESS["status_message"] = f"AUDIT HALTED: {ae}"
        
        # Emit error progress update
        try:
            socketio.emit('progress_update', AUDIT_PROGRESS, namespace='/')
        except Exception as e:
            print(f"Error emitting progress update: {e}")
    except Exception as e_audit_main:
        error_msg = f"UNEXPECTED CRITICAL ERROR: {type(e_audit_main).__name__} - {e_audit_main}"; log_to_ui_and_console(f"{Fore.RED}{sanitize_log_message(error_msg)}{Style.RESET_ALL}")
        import traceback; tb_str = traceback.format_exc(); log_to_ui_and_console(sanitize_log_message(tb_str)); ui_logs.append(strip_ansi(sanitize_log_message(tb_str)))
        audit_status = f"Failed Critically: {type(e_audit_main).__name__}"; current_audit_progress["status_message"] = audit_status
    finally:
        current_audit_progress["current_phase"] = "Finalizing"
        if jump_client:
            try: jump_client.close()
            except Exception as e_jump_close: log_to_ui_and_console(f"{Fore.YELLOW}Warn: Error closing jump client: {e_jump_close}{Style.RESET_ALL}", console_only=True)
        if ssh_cfg_file_obj:
            try: os.unlink(ssh_cfg_file_obj.name)
            except OSError as e_unlink: log_to_ui_and_console(f"{Fore.YELLOW}Warn: Temp SSH cfg delete fail: {e_unlink}{Style.RESET_ALL}")
            ssh_cfg_file_obj.close()
        ui_logs = [log_entry for log_entry in ui_logs if " कार्य प्रगति पर है " not in log_entry]
        if audit_status == "Running": audit_status = "Failed: Interrupted"; current_audit_progress["status_message"] = audit_status
        elif audit_status == "Completed": last_successful_audit_completion_time = datetime.now(); current_audit_progress["status_message"] = "Audit Completed"; current_audit_progress["percentage_complete"] = 100;
        if not last_successful_audit_completion_time and audit_status == "Completed": last_successful_audit_completion_time = datetime.now()
        audit_paused = False; audit_pause_event.set()

def generate_audit_charts_for_pdf(summary_data, base_report_dir_for_charts):
    chart_paths = {}; plt.style.use('seaborn-v0_8-whitegrid')
    labels_overall = ['Collected (Clean)', 'Collected (Violations)', 'Failed Collection', 'Failed SSH Auth', 'Failed ICMP']
    succeeded_cleanly = summary_data.get('collected', 0) - summary_data.get('with_violations', 0)
    sizes_overall = [max(0, succeeded_cleanly), max(0, summary_data.get('with_violations', 0)), max(0, summary_data.get('failed_collection', 0)), max(0, summary_data.get('failed_ssh_auth', 0)), max(0, summary_data.get('failed_icmp', 0))]
    filtered_labels_overall = [label_item for i, label_item in enumerate(labels_overall) if sizes_overall[i] > 0]
    filtered_sizes_overall = [size_item for size_item in sizes_overall if size_item > 0]
    if sum(filtered_sizes_overall) > 0:
        fig1, ax1 = plt.subplots(figsize=(8, 5)); ax1.pie(filtered_sizes_overall, labels=filtered_labels_overall, autopct='%1.1f%%', startangle=140, colors=plt.cm.Paired.colors); ax1.axis('equal'); plt.title('Overall Audit Status Distribution', fontsize=14); chart_path1 = os.path.join(base_report_dir_for_charts, "overall_status_chart.png"); plt.savefig(chart_path1, bbox_inches='tight'); plt.close(fig1); chart_paths['overall'] = chart_path1
    else: chart_paths['overall'] = None
    labels_success_fail = ['Total Configured', 'Data Collected', 'Any Failure Stage']; total_configured = summary_data.get('total_routers', 0); data_collected = summary_data.get('collected', 0); any_failure = total_configured - data_collected
    if total_configured > 0:
        sizes_success_fail = [total_configured, data_collected, max(0, any_failure)]; fig2, ax2 = plt.subplots(figsize=(7, 5)); bars = ax2.bar(labels_success_fail, sizes_success_fail, color=['skyblue', 'lightgreen', 'salmon']); ax2.set_ylabel('Number of Routers'); ax2.set_title('Audit Stage Completion Counts', fontsize=14)
        for bar_item in bars: yval = bar_item.get_height(); ax2.text(bar_item.get_x() + bar_item.get_width()/2.0, yval + 0.5, int(yval), ha='center', va='bottom', fontweight='bold')
        plt.xticks(rotation=15, ha="right"); chart_path2 = os.path.join(base_report_dir_for_charts, "success_failure_chart.png"); plt.savefig(chart_path2, bbox_inches='tight'); plt.close(fig2); chart_paths['success_fail'] = chart_path2
    else: chart_paths['success_fail'] = None
    return chart_paths

def generate_pdf_summary_report(summary_data, manifest_data, failure_details, inv_routers_map, base_report_dir):
    pdf_filepath = os.path.join(base_report_dir, PDF_SUMMARY_FILENAME); doc = SimpleDocTemplate(pdf_filepath); styles = getSampleStyleSheet(); story = []
    story.append(Paragraph("Router Audit Summary Report", styles['h1'])); story.append(Paragraph(f"<b>Date of Audit:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])); story.append(Paragraph(f"<b>Active Inventory File:</b> {APP_CONFIG.get('ACTIVE_INVENTORY_FILE', 'N/A')} (Format: {APP_CONFIG.get('ACTIVE_INVENTORY_FORMAT','yaml').upper()})", styles['Normal'])); story.append(Spacer(1, 0.2 * inch)) # Added format
    chart_image_paths = generate_audit_charts_for_pdf(summary_data, base_report_dir)
    if chart_image_paths.get('overall') and os.path.exists(chart_image_paths['overall']): story.append(ReportLabImage(chart_image_paths['overall'], width=6 * inch, height=3.75 * inch)); story.append(Spacer(1, 0.2 * inch))
    if chart_image_paths.get('success_fail') and os.path.exists(chart_image_paths['success_fail']): story.append(ReportLabImage(chart_image_paths['success_fail'], width=5.5 * inch, height=3.9 * inch)); story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph("Overall Audit Metrics", styles['h2']))
    summary_table_content = [ [Paragraph("<b>Metric</b>", styles['Normal']), Paragraph("<b>Value</b>", styles['Normal'])], ["Total Routers in Inventory", summary_data.get('total_routers', 0)], ["ICMP Reachable / Failed", f"{summary_data.get('icmp_reachable', 0)} / {summary_data.get('failed_icmp', 0)}"], ["SSH Authenticated / Failed", f"{summary_data.get('ssh_auth_ok', 0)} / {summary_data.get('failed_ssh_auth', 0)}"], ["Data Collected / Failed", f"{summary_data.get('collected', 0)} / {summary_data.get('failed_collection', 0)}"], ["Routers with Physical Line Telnet Violations", summary_data.get('with_violations', 0)], ["Total Physical Lines with Telnet Enabled", summary_data.get('total_physical_line_issues', 0)]]
    summary_table = Table(summary_table_content, colWidths=[3 * inch, 3 * inch]); summary_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.darkblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke), ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('BOTTOMPADDING', (0, 0), (-1, 0), 12), ('BACKGROUND', (0, 1), (-1, -1), colors.beige), ('GRID', (0, 0), (-1, -1), 1, colors.black)])); story.append(summary_table); story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph("Per-Router Detailed Status", styles['h2']))
    router_status_data = [[Paragraph("<b>Router (Inventory Name)</b>", styles['Normal']), Paragraph("<b>Status & Details</b>", styles['Normal'])]]; router_keys_for_pdf = inv_routers_map.keys() if isinstance(inv_routers_map, dict) else []
    for r_name_pdf in router_keys_for_pdf:
        status_pdf = summary_data.get("per_router_status", {}).get(r_name_pdf, "Status Unknown"); fail_reason_pdf = failure_details.get(r_name_pdf); details_pdf_str = status_pdf
        if fail_reason_pdf: details_pdf_str += f" (Reason: {str(fail_reason_pdf)[:200]})"
        router_status_data.append([Paragraph(r_name_pdf, styles['Normal']), Paragraph(details_pdf_str, styles['Normal'])])
    status_table = Table(router_status_data, colWidths=[1.5 * inch, 4.5 * inch]); status_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightblue), ('TEXTCOLOR', (0, 0), (-1, 0), colors.black), ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('GRID', (0, 0), (-1, -1), 1, colors.black), ('VALIGN', (0, 0), (-1, -1), 'TOP')])); story.append(status_table); story.append(Spacer(1, 0.2 * inch))
    if any(k not in ['summary_file', 'pdf_summary_file'] for k in manifest_data):
        story.append(Paragraph("Individual Report File Index", styles['h2']))
        file_index_data = [[Paragraph(s_header, styles['Normal']) for s_header in ["<b>Router</b>", "<b>Report A (Raw Data)</b>", "<b>Report B (Audit Detail)</b>", "<b>Report C (Parsed Line)</b>"]]]
        for r_name_pdf_manifest, files_dict_pdf_manifest in manifest_data.items():
            if r_name_pdf_manifest not in ['summary_file', 'pdf_summary_file']:
                file_index_data.append([r_name_pdf_manifest, files_dict_pdf_manifest.get('a', 'N/A'), files_dict_pdf_manifest.get('b', 'N/A'), files_dict_pdf_manifest.get('c', 'N/A')])
        index_table = Table(file_index_data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch]); index_table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen), ('TEXTCOLOR', (0, 0), (-1, 0), colors.black), ('ALIGN', (0, 0), (-1, -1), 'LEFT'), ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), ('GRID', (0, 0), (-1, -1), 1, colors.black)])); story.append(index_table)
    doc.build(story); print(f"PDF report generated at {pdf_filepath}")
    return pdf_filepath

def generate_excel_summary_report(summary_data, manifest_data, failure_details, inv_routers_map, base_report_dir):
    """Generate an Excel summary report of the audit results.
    
    Args:
        summary_data: Dictionary containing summary metrics (last_run_summary_data)
        manifest_data: Dictionary containing paths to reports (detailed_reports_manifest)
        failure_details: Dictionary containing failure reasons (current_run_failures)
        inv_routers_map: Dictionary of routers to audit (_INV_ROUTERS)
        base_report_dir: Base directory to save the report
        
    Returns:
        Path to the generated Excel file
    """
    excel_filepath = os.path.join(base_report_dir, EXCEL_SUMMARY_FILENAME)
    
    # Create a new workbook and get the active sheet
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Audit Summary"
    
    # Define styles
    title_font = Font(name='Calibri', size=16, bold=True)
    header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
    normal_font = Font(name='Calibri', size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alt_row_fill = PatternFill(start_color="E6F0FD", end_color="E6F0FD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Sheet title and metadata
    sheet.merge_cells('A1:C1')
    sheet['A1'] = "Router Audit Summary Report"
    sheet['A1'].font = title_font
    sheet['A1'].alignment = center_align
    sheet.row_dimensions[1].height = 30
    
    # Metadata
    sheet['A3'] = "Date of Audit:"
    sheet['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sheet['A4'] = "Active Inventory:"
    sheet['B4'] = f"{APP_CONFIG.get('ACTIVE_INVENTORY_FILE', 'N/A')} (Format: {APP_CONFIG.get('ACTIVE_INVENTORY_FORMAT','csv').upper()})"
    
    # Overall Audit Metrics section
    sheet['A6'] = "Overall Audit Metrics"
    sheet['A6'].font = Font(name='Calibri', size=14, bold=True)
    sheet.merge_cells('A6:C6')
    
    # Metrics header row
    sheet['A8'] = "Metric"
    sheet['B8'] = "Value"
    sheet['A8'].font = header_font
    sheet['B8'].font = header_font
    sheet['A8'].fill = header_fill
    sheet['B8'].fill = header_fill
    sheet['A8'].alignment = left_align
    sheet['B8'].alignment = left_align
    
    # Metrics data
    metrics = [
        ("Total Routers in Inventory", summary_data.get('total_routers', 0)),
        ("ICMP Reachable", summary_data.get('icmp_reachable', 0)),
        ("ICMP Failed", summary_data.get('failed_icmp', 0)),
        ("SSH Authenticated", summary_data.get('ssh_auth_ok', 0)),
        ("SSH Authentication Failed", summary_data.get('failed_ssh_auth', 0)),
        ("Data Collected", summary_data.get('collected', 0)),
        ("Collection Failed", summary_data.get('failed_collection', 0)),
        ("Routers with Physical Line Telnet Violations", summary_data.get('with_violations', 0)),
        ("Total Physical Lines with Telnet Enabled", summary_data.get('total_physical_line_issues', 0)),
        ("Default Telnet (no transport input)", summary_data.get('failure_category_default_telnet', 0)),
        ("Explicit 'transport input telnet'", summary_data.get('failure_category_explicit_telnet', 0)),
        ("'transport input all'", summary_data.get('failure_category_transport_all', 0))
    ]
    
    row_num = 9
    for i, (desc, val) in enumerate(metrics):
        sheet[f'A{row_num}'] = desc
        sheet[f'B{row_num}'] = val
        if i % 2 == 1:  # Apply alternate row coloring
            sheet[f'A{row_num}'].fill = alt_row_fill
            sheet[f'B{row_num}'].fill = alt_row_fill
        row_num += 1
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 40
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 50
    
    # Per-Router Detailed Status section
    row_num += 2  # Add some space
    sheet[f'A{row_num}'] = "Per-Router Detailed Status"
    sheet[f'A{row_num}'].font = Font(name='Calibri', size=14, bold=True)
    sheet.merge_cells(f'A{row_num}:D{row_num}')
    row_num += 2
    
    # Router status header row
    sheet[f'A{row_num}'] = "Router (Inventory Name)"
    sheet[f'B{row_num}'] = "Status"
    sheet[f'C{row_num}'] = "Details/Failure Reason"
    for col in ['A', 'B', 'C']:
        sheet[f'{col}{row_num}'].font = header_font
        sheet[f'{col}{row_num}'].fill = header_fill
        sheet[f'{col}{row_num}'].alignment = left_align
    row_num += 1
    
    # Router status data
    router_keys = inv_routers_map.keys() if isinstance(inv_routers_map, dict) else []
    for i, r_name in enumerate(router_keys):
        status = summary_data.get("per_router_status", {}).get(r_name, "Status Unknown")
        fail_reason = failure_details.get(r_name)
        details_str = str(fail_reason) if fail_reason else ""
        
        sheet[f'A{row_num}'] = r_name
        sheet[f'B{row_num}'] = status
        sheet[f'C{row_num}'] = details_str
        
        if i % 2 == 1:  # Apply alternate row coloring
            for col in ['A', 'B', 'C']:
                sheet[f'{col}{row_num}'].fill = alt_row_fill
        row_num += 1
    
    # Report File Index section
    row_num += 2  # Add some space
    sheet[f'A{row_num}'] = "Report File Index"
    sheet[f'A{row_num}'].font = Font(name='Calibri', size=14, bold=True)
    sheet.merge_cells(f'A{row_num}:D{row_num}')
    row_num += 2
    
    # Report index header row
    sheet[f'A{row_num}'] = "Router"
    sheet[f'B{row_num}'] = "Report A (Raw Data)"
    sheet[f'C{row_num}'] = "Report B (Audit Detail)"
    sheet[f'D{row_num}'] = "Report C (Parsed Line)"
    for col in ['A', 'B', 'C', 'D']:
        sheet[f'{col}{row_num}'].font = header_font
        sheet[f'{col}{row_num}'].fill = header_fill
        sheet[f'{col}{row_num}'].alignment = left_align
    row_num += 1
    
    # Report index data
    for i, (r_name, files_dict) in enumerate(manifest_data.items()):
        if r_name not in ['summary_file', 'pdf_summary_file', 'excel_summary_file']:
            sheet[f'A{row_num}'] = r_name
            sheet[f'B{row_num}'] = files_dict.get('a', 'N/A')
            sheet[f'C{row_num}'] = files_dict.get('b', 'N/A')
            sheet[f'D{row_num}'] = files_dict.get('c', 'N/A')
            
            if i % 2 == 1:  # Apply alternate row coloring
                for col in ['A', 'B', 'C', 'D']:
                    sheet[f'{col}{row_num}'].fill = alt_row_fill
            row_num += 1
    
    # Save the workbook
    wb.save(excel_filepath)
    print(f"Excel report generated at {excel_filepath}")
    return excel_filepath

HTML_BASE_LAYOUT = r"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{% block title %}Router Audit & Terminal Pro{% endblock %}</title>
    <link rel="stylesheet" href="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/css/bootstrap.min.css">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/5.15.1/css/all.min.css">
    <style>
        body { padding-top: 20px; }
        .footer { margin-top: 30px; padding: 10px; background-color: #f8f9fa; }
        pre { white-space: pre-wrap; }
        .log-container { max-height: 400px; overflow-y: auto; background-color: #f8f9fa; }
        .terminal-container { background-color: black; color: #33ff33; font-family: monospace; }
    </style>
    {% block head_extra %}{% endblock %}

    <style>
        /* Refresh indicator styles */
        #refresh-indicator {
            position: fixed;
            top: 10px;
            right: 10px;
            background: rgba(0,0,0,0.7);
            color: white;
            padding: 5px 10px;
            border-radius: 4px;
            display: flex;
            align-items: center;
            z-index: 9999;
            opacity: 0;
            transition: opacity 0.3s ease;
            pointer-events: none;
        }
        
        #refresh-indicator.refreshing {
            opacity: 1;
        }
        
        .refresh-spinner {
            width: 16px;
            height: 16px;
            border: 2px solid rgba(255,255,255,0.3);
            border-radius: 50%;
            border-top-color: white;
            animation: spin 1s linear infinite;
            margin-right: 8px;
        }
        
        @keyframes spin {
            to { transform: rotate(360deg); }
        }
        
        /* Highlight new log entries */
        .new-log-entry {
            animation: highlightNew 2s ease;
        }
        
        @keyframes highlightNew {
            0% { background-color: rgba(255, 255, 0, 0.3); }
            100% { background-color: transparent; }
        }
        
        /* Notifications container */
        #notifications-container {
            position: fixed;
            top: 20px;
            right: 20px;
            z-index: 9999;
            max-width: 350px;
        }
        
        .notification-toast {
            margin-bottom: 10px;
            box-shadow: 0 4px 8px rgba(0,0,0,0.1);
        }
    </style>
    </head>
<body>
    
    <!-- Refresh indicator -->
    <div id="refresh-indicator">
        <div class="refresh-spinner"></div>
        <span>Refreshing data...</span>
    </div>
    
    <!-- Notifications container -->
    <div id="notifications-container"></div>
    
    <div class="container">
        <nav class="navbar navbar-expand-lg navbar-light bg-light mb-4">
            <a class="navbar-brand" href="/"><i class="fas fa-network-wired"></i> Router Audit & Terminal Pro</a>
            <button class="navbar-toggler" type="button" data-toggle="collapse" data-target="#navbarNav">
                <span class="navbar-toggler-icon"></span>
            </button>
            <div class="collapse navbar-collapse" id="navbarNav">
                <ul class="navbar-nav mr-auto">
                    <li class="nav-item">
                        <a class="nav-link" href="/"><i class="fas fa-home"></i> Home</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/settings"><i class="fas fa-cogs"></i> Settings</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/manage_inventories"><i class="fas fa-tasks"></i> Manage Inventories</a>
                    </li>
                    <li class="nav-item">
                        <a class="nav-link" href="/captured_configs"><i class="fas fa-file-code"></i> Captured Configs</a>
                    </li>
                </ul>
                <ul class="navbar-nav ml-auto mr-3">
                    <li class="nav-item">
                        <a class="nav-link btn btn-danger text-white" href="#" id="stop-reset-audit"><i class="fas fa-stop-circle"></i> Stop/Reset Audit</a>
                    </li>
                </ul>
                <span class="navbar-text">
                    <span class="badge badge-info">Port: {{ APP_PORT }}</span>
                </span>
            </div>
        </nav>
        
        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category if category != 'message' else 'info' }} alert-dismissible fade show">
                        {{ message }}
                        <button type="button" class="close" data-dismiss="alert">&times;</button>
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        
        {% block content %}{% endblock %}
        
        <footer class="footer text-center">
            <p>&copy; {{ BROWSER_TIMESTAMP.year }} Router Audit & Terminal Pro</p>
        </footer>
    </div>
    
    <!-- Core JavaScript Libraries -->
    <script src="https://code.jquery.com/jquery-3.5.1.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/popper.js@1.16.1/dist/umd/popper.min.js"></script>
    <script src="https://stackpath.bootstrapcdn.com/bootstrap/4.5.2/js/bootstrap.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/socket.io/4.0.1/socket.io.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    
    <!-- Error Boundary System -->
    <script>
    // ErrorBoundary: Comprehensive error handling system for Frontend
    class ErrorBoundary {
        constructor(options = {}) {
            this.options = Object.assign({
                retryLimit: 3,              // Maximum retry attempts
                logErrorsToServer: true,     // Whether to send errors to server
                autoRecover: true,           // Try to auto-recover from errors
                trackState: true,            // Track state before errors
                notifyUser: true             // Show user notifications on errors
            }, options);
            
            this.errors = [];                // Error history
            this.componentStates = new Map(); // Component states before errors
            this.retryAttempts = new Map();   // Track retry attempts per component
            
            // Initialize error listener for uncaught errors
            this.initGlobalErrorListener();
        }
        
        // Start monitoring a UI component
        monitorComponent(componentId, componentName, initialState = null) {
            console.log(`ErrorBoundary: Started monitoring ${componentName} (${componentId})`);
            
            if (this.options.trackState && initialState) {
                this.componentStates.set(componentId, initialState);
            }
            
            this.retryAttempts.set(componentId, 0);
            
            return {
                updateState: (newState) => this.updateComponentState(componentId, newState),
                handleError: (error, context) => this.handleComponentError(componentId, componentName, error, context)
            };
        }
        
        // Update component state (for recovery)
        updateComponentState(componentId, newState) {
            if (this.options.trackState) {
                this.componentStates.set(componentId, newState);
            }
        }
        
        // Handle component-specific error
        handleComponentError(componentId, componentName, error, context = {}) {
            const timestamp = new Date();
            const errorInfo = {
                id: `error_${Date.now()}_${Math.random().toString(36).substr(2, 5)}`,
                componentId,
                componentName,
                timestamp,
                error: {
                    name: error.name,
                    message: error.message,
                    stack: error.stack
                },
                context,
                recoverable: true // Assume recoverable until determined otherwise
            };
            
            console.error(`ErrorBoundary: Error in ${componentName}:`, error);
            this.errors.push(errorInfo);
            
            // Log to server if enabled
            if (this.options.logErrorsToServer) {
                this.logErrorToServer(errorInfo);
            }
            
            // Show notification if enabled
            if (this.options.notifyUser) {
                this.showErrorNotification(errorInfo);
            }
            
            // Try to recover if enabled
            if (this.options.autoRecover) {
                return this.attemptRecovery(errorInfo);
            }
            
            return false;
        }
        
        // Attempt to recover from an error
        attemptRecovery(errorInfo) {
            const { componentId, componentName } = errorInfo;
            const attempts = this.retryAttempts.get(componentId) || 0;
            
            if (attempts < this.options.retryLimit) {
                this.retryAttempts.set(componentId, attempts + 1);
                
                console.log(`ErrorBoundary: Attempting recovery for ${componentName} (attempt ${attempts + 1}/${this.options.retryLimit})`);
                
                // Get previous state if available
                const previousState = this.componentStates.get(componentId);
                
                // Create error recovery element
                const recoveryEl = this.createRecoveryUI(errorInfo, attempts + 1);
                
                // If we have a recovery element, insert it in the DOM where the error occurred
                if (recoveryEl && errorInfo.componentId) {
                    const targetEl = document.getElementById(componentId);
                    if (targetEl) {
                        targetEl.innerHTML = '';
                        targetEl.appendChild(recoveryEl);
                    }
                }
                
                return { recovered: true, previousState, attempts: attempts + 1 };
            }
            
            console.error(`ErrorBoundary: Recovery limit reached for ${componentName} after ${attempts} attempts`);
            return { recovered: false, attempts };
        }
        
        // Create recovery UI element
        createRecoveryUI(errorInfo, attempts) {
            const container = document.createElement('div');
            container.className = 'error-boundary-container';
            container.style.padding = '15px';
            container.style.margin = '10px 0';
            container.style.backgroundColor = '#f8d7da';
            container.style.borderRadius = '4px';
            container.style.border = '1px solid #f5c6cb';
            container.style.color = '#721c24';
            
            const header = document.createElement('h6');
            header.textContent = `Component Error: ${errorInfo.componentName}`;
            header.style.fontWeight = 'bold';
            container.appendChild(header);
            
            const message = document.createElement('p');
            message.textContent = `An error occurred while updating this component. Retry attempt ${attempts}/${this.options.retryLimit}.`;
            container.appendChild(message);
            
            const errorDetails = document.createElement('details');
            errorDetails.style.marginTop = '10px';
            errorDetails.style.fontSize = '12px';
            
            const summary = document.createElement('summary');
            summary.textContent = 'Technical Details';
            errorDetails.appendChild(summary);
            
            const detailsContent = document.createElement('pre');
            detailsContent.textContent = `${errorInfo.error.name}: ${errorInfo.error.message}`;
            detailsContent.style.whiteSpace = 'pre-wrap';
            detailsContent.style.marginTop = '5px';
            errorDetails.appendChild(detailsContent);
            
            container.appendChild(errorDetails);
            
            const retryButton = document.createElement('button');
            retryButton.className = 'btn btn-sm btn-danger mt-2';
            retryButton.textContent = 'Retry Now';
            retryButton.addEventListener('click', () => this.manualRetry(errorInfo.componentId));
            container.appendChild(retryButton);
            
            return container;
        }
        
        // Manual retry initiated by the user
        manualRetry(componentId) {
            const component = document.getElementById(componentId);
            if (!component) return false;
            
            console.log(`ErrorBoundary: Manual retry for component ${componentId}`);
            
            // Dispatch custom event that components can listen for
            const retryEvent = new CustomEvent('error-boundary-retry', {
                detail: { componentId, timestamp: Date.now() }
            });
            component.dispatchEvent(retryEvent);
            window.dispatchEvent(retryEvent);
            
            return true;
        }
        
        // Show error notification to the user
        showErrorNotification(errorInfo) {
            // Create or get notification container
            let notificationContainer = document.getElementById('error-boundary-notifications');
            if (!notificationContainer) {
                notificationContainer = document.createElement('div');
                notificationContainer.id = 'error-boundary-notifications';
                notificationContainer.style.position = 'fixed';
                notificationContainer.style.bottom = '20px';
                notificationContainer.style.right = '20px';
                notificationContainer.style.zIndex = '9999';
                document.body.appendChild(notificationContainer);
            }
            
            // Create notification
            const notification = document.createElement('div');
            notification.className = 'error-notification';
            notification.style.backgroundColor = '#f8d7da';
            notification.style.color = '#721c24';
            notification.style.padding = '10px 15px';
            notification.style.marginTop = '10px';
            notification.style.borderRadius = '4px';
            notification.style.boxShadow = '0 2px 5px rgba(0,0,0,0.2)';
            notification.style.width = '300px';
            notification.style.position = 'relative';
            
            // Create notification content
            const title = document.createElement('div');
            title.style.fontWeight = 'bold';
            title.style.marginBottom = '5px';
            title.textContent = `Error in ${errorInfo.componentName}`;
            
            const message = document.createElement('div');
            message.textContent = errorInfo.error.message || 'An error occurred in this component';
            
            const closeBtn = document.createElement('button');
            closeBtn.textContent = '×';
            closeBtn.style.position = 'absolute';
            closeBtn.style.top = '5px';
            closeBtn.style.right = '10px';
            closeBtn.style.background = 'none';
            closeBtn.style.border = 'none';
            closeBtn.style.cursor = 'pointer';
            closeBtn.style.fontSize = '20px';
            closeBtn.style.fontWeight = 'bold';
            closeBtn.style.color = '#721c24';
            closeBtn.addEventListener('click', () => notification.remove());
            
            // Add everything to the notification
            notification.appendChild(closeBtn);
            notification.appendChild(title);
            notification.appendChild(message);
            
            // Add auto-dismiss after 8 seconds
            notificationContainer.appendChild(notification);
            setTimeout(() => {
                notification.style.opacity = '0';
                notification.style.transition = 'opacity 0.5s';
                setTimeout(() => notification.remove(), 500);
            }, 8000);
        }
        
        // Log error to server for monitoring
        logErrorToServer(errorInfo) {
            // Strip sensitive data
            const safeErrorInfo = {
                ...errorInfo,
                context: this.sanitizeContext(errorInfo.context)
            };
            
            // Send to server via fetch API
            fetch('/log_client_error', {
                method: 'POST',
                headers: { 'Content-Type': 'application/json' },
                body: JSON.stringify(safeErrorInfo)
            }).catch(err => {
                console.error('Failed to send error to server:', err);
            });
        }
        
        // Remove sensitive data from error context
        sanitizeContext(context) {
            if (!context) return {};
            
            const sanitized = {...context};
            const sensitiveKeys = ['password', 'token', 'secret', 'key', 'auth', 'credentials'];
            
            // Recursively sanitize nested objects
            const sanitizeObj = (obj) => {
                if (!obj || typeof obj !== 'object') return obj;
                
                Object.keys(obj).forEach(key => {
                    // Check if this key contains sensitive information
                    const isKeyMatch = sensitiveKeys.some(sensitiveKey => 
                        key.toLowerCase().includes(sensitiveKey)
                    );
                    
                    if (isKeyMatch) {
                        obj[key] = '[REDACTED]';
                    } else if (typeof obj[key] === 'object' && obj[key] !== null) {
                        obj[key] = sanitizeObj(obj[key]);
                    }
                });
                
                return obj;
            };
            
            return sanitizeObj(sanitized);
        }
        
        // Initialize global error listener
        initGlobalErrorListener() {
            // Listen for unhandled errors
            window.addEventListener('error', (event) => {
                const error = event.error || new Error(event.message);
                console.error('Global error caught by ErrorBoundary:', error);
                
                // Try to determine which component caused the error
                let componentId = 'unknown';
                let componentName = 'Global Scope';
                
                // Create error info
                const errorInfo = {
                    id: `global_${Date.now()}_${Math.random().toString(36).substr(2, 5)}`,
                    componentId,
                    componentName,
                    timestamp: new Date(),
                    error: {
                        name: error.name,
                        message: error.message,
                        stack: error.stack
                    },
                    context: { url: window.location.href, source: event.filename, line: event.lineno, col: event.colno },
                    recoverable: false
                };
                
                this.errors.push(errorInfo);
                
                if (this.options.logErrorsToServer) {
                    this.logErrorToServer(errorInfo);
                }
                
                if (this.options.notifyUser) {
                    this.showErrorNotification(errorInfo);
                }
                
                // Prevent the browser's default error handling
                event.preventDefault();
                return true;
            });
            
            // Listen for unhandled promise rejections
            window.addEventListener('unhandledrejection', (event) => {
                const error = event.reason instanceof Error ? event.reason 
                              : new Error(String(event.reason || 'Promise rejection'));
                              
                console.error('Unhandled Promise rejection caught by ErrorBoundary:', error);
                
                // Create error info
                const errorInfo = {
                    id: `promise_${Date.now()}_${Math.random().toString(36).substr(2, 5)}`,
                    componentId: 'promise',
                    componentName: 'Asynchronous Operation',
                    timestamp: new Date(),
                    error: {
                        name: error.name,
                        message: error.message,
                        stack: error.stack
                    },
                    context: { url: window.location.href },
                    recoverable: false
                };
                
                this.errors.push(errorInfo);
                
                if (this.options.logErrorsToServer) {
                    this.logErrorToServer(errorInfo);
                }
                
                if (this.options.notifyUser) {
                    this.showErrorNotification(errorInfo);
                }
                
                // Prevent the browser's default rejection handling
                event.preventDefault();
                return true;
            });
        }
    }
    
    // Create global ErrorBoundary instance
    window.errorBoundary = new ErrorBoundary();
    console.log('Error Boundary System initialized successfully');
    
    // Safe function wrapper to catch and handle errors
    function safeCall(fn, componentId, componentName, context = {}) {
        try {
            return fn();
        } catch (error) {
            window.errorBoundary.handleComponentError(componentId, componentName, error, context);
            return null;
        }
    }
    
    // Wrap setTimeout to catch errors in callbacks
    const originalSetTimeout = window.setTimeout;
    window.setTimeout = function(callback, delay, ...args) {
        const wrappedCallback = function() {
            try {
                callback.apply(this, args);
            } catch (error) {
                window.errorBoundary.handleComponentError(
                    'async-timeout', 
                    'Async Timeout Callback', 
                    error, 
                    { delay }
                );
            }
        };
        return originalSetTimeout(wrappedCallback, delay);
    };
    
    // Patch Socket.IO event handlers to catch errors
    if (typeof io !== 'undefined' && io && io.Socket && io.Socket.prototype) {
        const originalOn = io.Socket.prototype.on;
        io.Socket.prototype.on = function(event, callback) {
            const wrappedCallback = function(...args) {
                try {
                    return callback.apply(this, args);
                } catch (error) {
                    window.errorBoundary.handleComponentError(
                        'socket-' + event, 
                        'Socket.IO Event: ' + event, 
                        error, 
                        { eventArgs: args }
                    );
                }
            };
            return originalOn.call(this, event, wrappedCallback);
        };
    } else {
        console.warn('Socket.IO not available yet, skipping event handler patching');
    }
        return originalOn.call(this, event, wrappedCallback);
    };
    
    // Helper to create monitored components
    function createMonitoredComponent(id, name, options = {}) {
        const el = document.getElementById(id);
        if (!el) {
            console.warn(`Cannot monitor component ${name} - element with ID ${id} not found`);
            return null;
        }
        
        // Create initial state from options
        const initialState = options.initialState || {};
        
        // Register with error boundary
        const monitor = window.errorBoundary.monitorComponent(id, name, initialState);
        
        // Return the component helper
        return {
            el,
            monitor,
            update: function(updateFn) {
                return safeCall(
                    () => updateFn(el), 
                    id, 
                    name, 
                    { action: 'update' }
                );
            },
            setContent: function(content) {
                return safeCall(
                    () => { el.innerHTML = content; }, 
                    id, 
                    name, 
                    { action: 'setContent', contentLength: String(content).length }
                );
            },
            updateState: function(newState) {
                this.monitor.updateState(newState);
            }
        };
    }
    </script>

    <!-- Template block for page-specific scripts -->
    {% block custom_scripts %}{% endblock %}
    
    <!-- Core application script -->
    <script>
    // Function to update the enhanced progress tracking UI
    function updateEnhancedProgressUI(data) {
        console.log('Updating Enhanced Progress UI with data:', data);
        if (!data.enhanced_progress) {
            console.warn('No enhanced_progress data available');
            return;
        }
        
        const progress = data.enhanced_progress;
        const statusCounts = progress.status_counts || {};
        
        // Helper function to safely update DOM elements
        function safeUpdateElement(id, updateFunc) {
            const element = document.getElementById(id);
            if (element) {
                updateFunc(element);
            } else {
                console.warn(`Element with id '${id}' not found`);
            }
        }
        
        // Update progress bar
        const percentComplete = progress.percentage || 0;
        safeUpdateElement('enhanced-progress-bar', (el) => {
            el.style.width = percentComplete + '%';
            el.setAttribute('aria-valuenow', percentComplete);
            el.textContent = percentComplete.toFixed(1) + '%';
        });
        
        // Update status information
        safeUpdateElement('enhanced-audit-status', (el) => {
            el.textContent = progress.status || 'Idle';
        });
        
        safeUpdateElement('enhanced-elapsed-time', (el) => {
            el.textContent = progress.elapsed_time || '00:00:00';
        });
        
        safeUpdateElement('enhanced-completed-count', (el) => {
            el.textContent = progress.completed_devices || 0;
        });
        
        safeUpdateElement('enhanced-total-count', (el) => {
            el.textContent = progress.total_devices || 0;
        });
        
        // Update ETA
        safeUpdateElement('enhanced-eta', (el) => {
            if (progress.estimated_completion_time && progress.status === 'running') {
                try {
                    const etaDate = new Date(progress.estimated_completion_time);
                    const hours = etaDate.getHours().toString().padStart(2, '0');
                    const minutes = etaDate.getMinutes().toString().padStart(2, '0');
                    const seconds = etaDate.getSeconds().toString().padStart(2, '0');
                    el.textContent = `${hours}:${minutes}:${seconds}`;
                } catch (e) {
                    console.error('Error formatting ETA:', e);
                    el.textContent = '--:--:--';
                }
            } else {
                el.textContent = '--:--:--';
            }
        });
        
        // Update current device
        safeUpdateElement('current-device-name', (el) => {
            el.textContent = progress.current_device || 'None';
        });
        
        // Update status counts
        safeUpdateElement('success-count', (el) => {
            el.textContent = statusCounts.success || 0;
        });
        
        safeUpdateElement('warning-count', (el) => {
            el.textContent = statusCounts.warning || 0;
        });
        
        safeUpdateElement('failure-count', (el) => {
            el.textContent = statusCounts.failure || 0;
        });
        
        console.log('Enhanced Progress UI updated successfully');
    }
    
    // Function to fetch progress data and update UI
    function fetchProgressData() {
        // Add visual indicator that data is being refreshed
        const refreshIndicator = document.getElementById('refresh-indicator');
        if (refreshIndicator) {
            refreshIndicator.classList.add('refreshing');
        }
        
        // Show a small loading message in the UI
        const statusIndicator = document.createElement('div');
        statusIndicator.id = 'temp-refresh-indicator';
        statusIndicator.innerHTML = '<span class="badge badge-info">Refreshing data...</span>';
        statusIndicator.style.position = 'fixed';
        statusIndicator.style.top = '10px';
        statusIndicator.style.right = '10px';
        statusIndicator.style.zIndex = '9999';
        
        // Remove any existing indicator
        const existingIndicator = document.getElementById('temp-refresh-indicator');
        if (existingIndicator) {
            existingIndicator.remove();
        }
        
        // Add the new indicator
        document.body.appendChild(statusIndicator);
        
        // Use the more reliable /get_audit_progress endpoint instead of /audit_progress_data
        fetch('/get_audit_progress')
            .then(response => {
                if (!response.ok) {
                    throw new Error(`HTTP error ${response.status}`);
                }
                return response.json();
            })
            .then(apiData => {
                // Also fetch UI logs and other dashboard data
                return fetch('/audit_progress_data')
                    .then(response => {
                        if (!response.ok) {
                            throw new Error(`HTTP error ${response.status}`);
                        }
                        return response.json();
                    })
                    .then(uiData => {
                        // Combine the data from both endpoints
                        const data = { ...uiData, ...apiData };
                        return data;
                    });
            })
            .then(data => {
                // Update standard progress elements
                const progressBar = document.querySelector('.progress-bar');
                if (progressBar) {
                    const percent = data.progress?.percentage_complete || 0;
                    progressBar.style.width = percent + '%';
                    progressBar.setAttribute('aria-valuenow', percent);
                    progressBar.textContent = percent + '%';
                }
                
                // Update overall status display
                const statusElement = document.getElementById('audit-status');
                if (statusElement && data.overall_audit_status) {
                    statusElement.textContent = data.overall_audit_status;
                }
                
                // Update current device display
                const deviceElement = document.getElementById('current-device');
                if (deviceElement && data.progress && data.progress.current_device) {
                    deviceElement.textContent = data.progress.current_device || 'N/A';
                }
                
                // Update completed/total devices
                const devicesElement = document.getElementById('devices-processed');
                if (devicesElement && data.progress) {
                    const completed = data.progress.completed_devices || 0;
                    const total = data.progress.total_devices || 0;
                    devicesElement.textContent = `${completed} / ${total}`;
                }
                
                // Update enhanced progress tracking UI
                updateEnhancedProgressUI(data);
                
                // Update logs with smart diffing (only append new logs)
                const logsContainer = document.getElementById('logs-container');
                if (logsContainer && data.ui_logs) {
                    // Store current log count to determine if new logs were added
                    const currentLogCount = logsContainer.childElementCount;
                    const newLogsCount = data.ui_logs.length;
                    
                    // Only clear and rebuild if the structure changed significantly
                    if (Math.abs(currentLogCount - newLogsCount) > 5 || newLogsCount < currentLogCount) {
                        logsContainer.innerHTML = '';
                        data.ui_logs.forEach(log => {
                            const logLine = document.createElement('div');
                            logLine.innerHTML = log;
                            logsContainer.appendChild(logLine);
                        });
                    } else if (newLogsCount > currentLogCount) {
                        // Append only new logs for efficiency
                        for (let i = currentLogCount; i < newLogsCount; i++) {
                            const logLine = document.createElement('div');
                            logLine.innerHTML = data.ui_logs[i];
                            logLine.classList.add('new-log-entry');
                            logsContainer.appendChild(logLine);
                            
                            // Remove highlight after animation
                            setTimeout(() => {
                                logLine.classList.remove('new-log-entry');
                            }, 2000);
                        }
                    }
                    
                    // Always scroll to bottom for new logs
                    logsContainer.scrollTop = logsContainer.scrollHeight;
                }
                
                // Force update of Summary Reports section
                if (data.reports_manifest) {
                    // Check if there are summary reports available
                    const summaryReportsContainer = document.querySelector('.card-body');
                    if (summaryReportsContainer) {
                        // Check for PDF summary report
                        if (data.reports_manifest.pdf_summary_file) {
                            const pdfButton = document.querySelector('a.btn-outline-danger');
                            if (!pdfButton) {
                                const pdfLink = document.createElement('p');
                                pdfLink.innerHTML = `<a href="/reports/${data.reports_manifest.pdf_summary_file.folder}/${data.reports_manifest.pdf_summary_file.filename}" class="btn btn-outline-danger">
                                    <i class="fas fa-file-pdf"></i> Download PDF Summary Report
                                </a>`;
                                summaryReportsContainer.insertBefore(pdfLink, summaryReportsContainer.firstChild);
                            }
                        }
                        
                        // Check for Excel summary report
                        if (data.reports_manifest.excel_summary_file) {
                            const excelButton = document.querySelector('a.btn-outline-success');
                            if (!excelButton) {
                                const excelLink = document.createElement('p');
                                excelLink.innerHTML = `<a href="/reports/${data.reports_manifest.excel_summary_file.folder}/${data.reports_manifest.excel_summary_file.filename}" class="btn btn-outline-success">
                                    <i class="fas fa-file-excel"></i> Download Excel Summary Report
                                </a>`;
                                summaryReportsContainer.insertBefore(excelLink, summaryReportsContainer.firstChild);
                            }
                        }
                        
                        // Check for JSON summary report
                        if (data.reports_manifest.summary_file) {
                            const jsonButton = document.querySelector('a.btn-outline-primary');
                            if (!jsonButton) {
                                const jsonLink = document.createElement('p');
                                jsonLink.innerHTML = `<a href="/view_json_report?folder=${data.reports_manifest.summary_file.folder}&filename=${data.reports_manifest.summary_file.filename}" class="btn btn-outline-primary">
                                    <i class="fas fa-file-code"></i> View JSON Summary Report
                                </a>`;
                                summaryReportsContainer.insertBefore(jsonLink, summaryReportsContainer.firstChild);
                            }
                        }
                    }
                }
                
                // Update page title with status for better user awareness
                document.title = `NetAuditPro - ${data.overall_audit_status || 'Dashboard'}`;
                
                // Remove refresh indicator
                if (refreshIndicator) {
                    refreshIndicator.classList.remove('refreshing');
                }
                
                // Remove the temporary status indicator
                const tempIndicator = document.getElementById('temp-refresh-indicator');
                if (tempIndicator) {
                    tempIndicator.remove();
                }
                
                // Adaptive polling rate based on activity
                let refreshRate = 5000; // Default 5 seconds
                
                if (data.overall_audit_status === 'running' || data.overall_audit_status === 'Running') {
                    refreshRate = 1000; // 1 second for active audits
                } else if (data.audit_paused) {
                    refreshRate = 2000; // 2 seconds for paused audits
                } else if (data.progress && data.progress.status === 'completed') {
                    refreshRate = 10000; // 10 seconds for completed audits
                }
                
                // Schedule next update with adaptive rate
                if (window.nextFetchTimeout) {
                    clearTimeout(window.nextFetchTimeout);
                }
                window.nextFetchTimeout = setTimeout(fetchProgressData, refreshRate);
                
                // Log refresh event
                console.log(`Data refreshed. Next refresh in ${refreshRate/1000}s. Status: ${data.overall_audit_status}`);
            })
            .catch(error => {
                console.error('Error fetching progress data:', error);
                
                // Remove refresh indicator on error
                if (refreshIndicator) {
                    refreshIndicator.classList.remove('refreshing');
                }
                
                // Update the temporary status indicator to show error
                const tempIndicator = document.getElementById('temp-refresh-indicator');
                if (tempIndicator) {
                    tempIndicator.innerHTML = '<span class="badge badge-danger">Error refreshing data. Retrying...</span>';
                    
                    // Remove after 3 seconds
                    setTimeout(() => {
                        if (tempIndicator && tempIndicator.parentNode) {
                            tempIndicator.remove();
                        }
                    }, 3000);
                }
                
                // Exponential backoff for retries (up to 30 seconds)
                const retryDelay = Math.min(30000, (window.lastRetryDelay || 5000) * 1.5);
                window.lastRetryDelay = retryDelay;
                
                console.log(`Retrying in ${retryDelay/1000} seconds...`);
                
                // Clear any existing timeout
                if (window.nextFetchTimeout) {
                    clearTimeout(window.nextFetchTimeout);
                }
                
                window.nextFetchTimeout = setTimeout(fetchProgressData, retryDelay);
            });
    }
    
    // Function to cancel current fetch timeout and start a new one
    function restartDataFetch() {
        if (window.nextFetchTimeout) {
            clearTimeout(window.nextFetchTimeout);
        }
        window.lastRetryDelay = 5000; // Reset retry delay
        fetchProgressData();
    }
    

    // Enhanced Stop/Reset functionality
    function stopAuditWithConfirmation() {
        if (confirm("Are you sure you want to stop and reset the current audit? This will terminate all connections.")) {
            // Show loading notification
            showNotification("Stopping audit...", "info");
            
            // Call the correct stop/reset audit endpoint
            fetch('/stop_reset_audit', { 
                method: 'POST',
                headers: {
                    'Content-Type': 'application/json'
                }
            })
                .then(response => {
                    if (!response.ok) {
                        throw new Error(`HTTP error ${response.status}`);
                    }
                    return response.json();
                })
                .then(data => {
                    console.log('Audit stopped:', data);
                    
                    // Show success notification
                    showNotification('Audit stopped and reset successfully', 'success');
                    
                    // Force immediate data refresh
                    restartDataFetch();
                    
                    // Reload the page after a short delay to ensure fresh state
                    setTimeout(() => {
                        location.reload();
                    }, 2000);
                })
                .catch(error => {
                    console.error('Error stopping audit:', error);
                    // Show error notification
                    showNotification('Error stopping audit: ' + error, 'danger');
                    
                    // Still try to refresh the data
                    restartDataFetch();
                });
        }
    }
    
    // Notification helper
    function showNotification(message, type = 'info') {
        const notificationId = 'notification-' + Date.now();
        const notificationHTML = `
            <div id="${notificationId}" class="alert alert-${type} alert-dismissible fade show notification-toast" role="alert">
                ${message}
                <button type="button" class="close" data-dismiss="alert" aria-label="Close">
                    <span aria-hidden="true">&times;</span>
                </button>
            </div>
        `;
        
        // Add notification to container
        const notificationsContainer = document.getElementById('notifications-container');
        if (notificationsContainer) {
            notificationsContainer.innerHTML += notificationHTML;
            
            // Auto-remove after 5 seconds
            setTimeout(() => {
                const notification = document.getElementById(notificationId);
                if (notification) {
                    notification.classList.remove('show');
                    setTimeout(() => notification.remove(), 500);
                }
    }
}

// Document ready handler
document.addEventListener('DOMContentLoaded', function() {
    fetchProgressData();
    loadInventoryDataFromAPI();
    
    // Initialize charts if they exist
    const resultsChartCanvas = document.getElementById('resultsChart');
    if (resultsChartCanvas) {
        // Use chart data if available
        try {
            // Check if chart_data is defined in the template context
            let chartData;
            // Use chart data if available
            try {
                // Check if chart_data is defined in the template context
                let chartData;
                try {
                    // Use a safer approach with JSON.parse
                    const chartDataStr = '{{ chart_data|default({})|tojson|safe }}';
                    chartData = JSON.parse(chartDataStr);
                } catch (e) {
                    console.error('Error parsing chart data:', e);
                    // If chart_data is not defined or invalid, use default data
                    chartData = {"labels":["No Data"],"values":[1],"colors":["#D3D3D3"]};
                }
                new Chart(resultsChartCanvas, {
                    type: 'pie',
                    data: {
                        labels: chartData.labels,
                        datasets: [{
                            data: chartData.values,
                            backgroundColor: chartData.colors
                        }]
                    },
                    options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        legend: {
                            position: 'right'
                        }
                    }
                });
            } catch (e) {
                console.error('Error initializing chart:', e);
            }
        }
    });
</script>
{% block extra_scripts %}{% endblock %}
</body>
</html>"""

HTML_INDEX_PAGE = """{% extends "base_layout.html" %}
{% block title %}Home - Router Audit & Terminal Pro{% endblock %}

{% block head_extra %}
<script src="https://cdn.jsdelivr.net/npm/chart.js@2.9.4/dist/Chart.min.js"></script>
<style>
    .progress { height: 25px; }
    .audit-controls { margin-bottom: 20px; }
    .status-card { margin-bottom: 20px; }
    .chart-container { position: relative; height: 300px; width: 100%; }
</style>
{% endblock %}

{% block content %}
<h1><i class="fas fa-network-wired"></i> Router Audit Dashboard</h1>
<p class="lead">Monitor and control network device audits. Current active inventory: <strong>{{ active_inventory_file }}</strong></p>

<!-- Audit Control Buttons -->
<div class="card status-card">
    <div class="card-header bg-primary text-white">
        <h4><i class="fas fa-play-circle"></i> Audit Controls</h4>
    </div>
    <div class="card-body">
        <div class="row">
            <div class="col-md-4">
                <form method="POST" action="{{ url_for('start_audit_route') }}">
                    <button type="submit" class="btn btn-success btn-lg btn-block" {% if audit_status == "Running" %}disabled{% endif %}>
                        <i class="fas fa-play"></i> Run Audit
                    </button>
                </form>
            </div>
            <div class="col-md-4">
                <form method="POST" action="{{ url_for('pause_audit_route') }}">
                    <button type="submit" class="btn btn-warning btn-lg btn-block" {% if audit_status != "Running" or audit_paused %}disabled{% endif %}>
                        <i class="fas fa-pause"></i> Pause Audit
                    </button>
                </form>
            </div>
            <div class="col-md-4">
                <form method="POST" action="{{ url_for('resume_audit_route') }}">
                    <button type="submit" class="btn btn-info btn-lg btn-block" {% if not audit_paused %}disabled{% endif %}>
                        <i class="fas fa-play"></i> Resume Audit
                    </button>
                </form>
            </div>
        </div>
    </div>
</div>

<!-- Audit Status -->
<div class="card status-card">
    <div class="card-header bg-info text-white">
        <h4><i class="fas fa-tasks"></i> Current Audit Status: {{ audit_status }}</h4>
    </div>
    <div class="card-body">
        {% if audit_status == "Running" or audit_status == "Completed" %}
            <h5>Progress: <span id="progress-percentage">{{ current_audit_progress.percentage_complete }}</span>% Complete</h5>
            <div class="progress mb-3">
                <div class="progress-bar progress-bar-striped progress-bar-animated" role="progressbar" 
                     style="width: {{ current_audit_progress.percentage_complete }}%" 
                     aria-valuenow="{{ current_audit_progress.percentage_complete }}" aria-valuemin="0" aria-valuemax="100">
                    {{ current_audit_progress.percentage_complete }}%
                </div>
            </div>
            <p><strong>Status:</strong> <span id="status-message">{{ current_audit_progress.status_message }}</span></p>
            <p><strong>Current Device:</strong> <span id="current-device-display">{{ current_audit_progress.current_device_hostname }}</span></p>
            <p><strong>Devices Processed:</strong> <span id="devices-processed">{{ current_audit_progress.devices_processed_count }} / {{ current_audit_progress.total_devices_to_process }}</span></p>
            
            <!-- Enhanced Progress Tracking UI -->
            <div class="card mt-4 mb-3">
                <div class="card-header bg-primary text-white">
                    <h5><i class="fas fa-tachometer-alt"></i> Enhanced Progress Tracking</h5>
                </div>
                <div class="card-body" id="enhanced-progress-container">
                    <div class="row">
                        <div class="col-md-6">
                            <h6 class="text-muted">Overall Progress</h6>
                            <div class="progress mb-3" style="height: 25px;">
                                <div id="enhanced-progress-bar" class="progress-bar progress-bar-striped progress-bar-animated" 
                                     role="progressbar" style="width: 0%;" aria-valuenow="0" aria-valuemin="0" aria-valuemax="100">
                                    0%
                                </div>
                            </div>
                            
                            <div class="row">
                                <div class="col-6">
                                    <p><strong>Status:</strong> <span id="progress-status">Idle</span></p>
                                    <p><strong>Elapsed:</strong> <span id="progress-elapsed">00:00:00</span></p>
                                </div>
                                <div class="col-6">
                                    <p><strong>Completed:</strong> <span id="progress-completed">0/0</span></p>
                                    <p><strong>ETA:</strong> <span id="progress-eta">--:--:--</span></p>
                                </div>
                            </div>
                        </div>
                        
                        <div class="col-md-6">
                            <h6 class="text-muted">Current Device</h6>
                            <div class="alert alert-info">
                                <h5 id="current-device">None</h5>
                                <div class="progress mt-2" style="height: 20px;">
                                    <div id="device-progress-bar" class="progress-bar progress-bar-striped progress-bar-animated bg-info" 
                                         role="progressbar" style="width: 100%;" aria-valuenow="100" aria-valuemin="0" aria-valuemax="100">
                                        In Progress
                                    </div>
                                </div>
                            </div>
                            
                            <div class="mt-3">
                                <h6 class="text-muted">Device Status</h6>
                                <div class="row">
                                    <div class="col-4 text-center">
                                        <div class="alert alert-success mb-1">
                                            <span id="success-count">0</span>
                                        </div>
                                        <small>Success</small>
                                    </div>
                                    <div class="col-4 text-center">
                                        <div class="alert alert-warning mb-1">
                                            <span id="warning-count">0</span>
                                        </div>
                                        <small>Warning</small>
                                    </div>
                                    <div class="col-4 text-center">
                                        <div class="alert alert-danger mb-1">
                                            <span id="failure-count">0</span>
                                        </div>
                                        <small>Failure</small>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        {% else %}
            <p>No audit is currently running. Click "Run Audit" to start.</p>
        {% endif %}
    </div>
</div>

<!-- Results Summary -->
{% if last_run_summary %}
<div class="row">
    <div class="col-md-6">
        <div class="card status-card">
            <div class="card-header bg-success text-white">
                <h4><i class="fas fa-chart-pie"></i> Audit Results Summary</h4>
            </div>
            <div class="card-body">
                <div class="chart-container">
                    <canvas id="resultsChart"></canvas>
                </div>
            </div>
        </div>
    </div>
    <div class="col-md-6">
        <div class="card status-card">
            <div class="card-header bg-secondary text-white">
                <h4><i class="fas fa-file-alt"></i> Summary Reports</h4>
            </div>
            <div class="card-body">
                {% if reports_manifest and reports_manifest.summary_file %}
                    <p><a href="{{ url_for('view_json_report', folder=reports_manifest.summary_file.folder, filename=reports_manifest.summary_file.filename) }}" class="btn btn-outline-primary">
                        <i class="fas fa-file-code"></i> View JSON Summary Report
                    </a></p>
                {% endif %}
                
                {% if reports_manifest and reports_manifest.pdf_summary_file %}
                    <p><a href="{{ url_for('serve_report_file', folder_relative_path=reports_manifest.pdf_summary_file.folder + '/' + reports_manifest.pdf_summary_file.filename) }}" class="btn btn-outline-danger">
                        <i class="fas fa-file-pdf"></i> Download PDF Summary Report
                    </a></p>
                {% endif %}
                
                {% if reports_manifest and reports_manifest.excel_summary_file %}
                    <p><a href="{{ url_for('serve_report_file', folder_relative_path=reports_manifest.excel_summary_file.folder + '/' + reports_manifest.excel_summary_file.filename) }}" class="btn btn-outline-success">
                        <i class="fas fa-file-excel"></i> Download Excel Summary Report
                    </a></p>
                {% endif %}
                
                <hr>
                <h5>Statistics:</h5>
                <ul>
                    <li><strong>Total Routers:</strong> {{ last_run_summary.total_routers }}</li>
                    <li><strong>Successfully Collected:</strong> {{ last_run_summary.collected }}</li>
                    <li><strong>With Violations:</strong> {{ last_run_summary.with_violations }}</li>
                    <li><strong>Failed Collection:</strong> {{ last_run_summary.failed_collection }}</li>
                    <li><strong>Failed SSH Auth:</strong> {{ last_run_summary.failed_ssh_auth }}</li>
                    <li><strong>Failed ICMP:</strong> {{ last_run_summary.failed_icmp }}</li>
                </ul>
            </div>
        </div>
    </div>
</div>
{% endif %}

<!-- Detailed Reports -->
{% if reports_manifest and reports_manifest|length > 2 %}
<div class="card status-card">
    <div class="card-header bg-info text-white">
        <h4><i class="fas fa-file-alt"></i> Detailed Router Reports</h4>
    </div>
    <div class="card-body">
        <div class="table-responsive">
            <table class="table table-striped table-hover">
                <thead class="thead-dark">
                    <tr>
                        <th>Router</th>
                        <th>Report A (Raw)</th>
                        <th>Report B (Audit)</th>
                        <th>Report C (Parsed)</th>
                                            </tr>
                </thead>
                <tbody>
                    {% for router_name, files in reports_manifest.items() %}
                        {% if router_name != 'summary_file' and router_name != 'pdf_summary_file' and router_name != 'excel_summary_file' %}
                        <tr>
                            <td><strong>{{ router_name }}</strong></td>
                            <td>
                                {% if files.a %}
                                <a href="{{ url_for('view_json_report', folder=files.a.folder, filename=files.a.filename) }}" class="btn btn-sm btn-outline-secondary">
                                    <i class="fas fa-file-alt"></i> View
                                </a>
                                {% else %}
                                <span class="badge badge-danger">Not Available</span>
                                {% endif %}
                            </td>
                            <td>
                                {% if files.b %}
                                <a href="{{ url_for('view_json_report', folder=files.b.folder, filename=files.b.filename) }}" class="btn btn-sm btn-outline-secondary">
                                    <i class="fas fa-file-alt"></i> View
                                </a>
                                {% else %}
                                <span class="badge badge-danger">Not Available</span>
                                {% endif %}
                            </td>
                            <td>
                                {% if files.c %}
                                <a href="{{ url_for('view_json_report', folder=files.c.folder, filename=files.c.filename) }}" class="btn btn-sm btn-outline-secondary">
                                    <i class="fas fa-file-alt"></i> View
                                </a>
                                {% else %}
                                <span class="badge badge-danger">Not Available</span>
                                {% endif %}
                            </td>
                                            </tr>
                        {% endif %}
                    {% endfor %}
                </tbody>
            </table>
        </div>
    </div>
</div>
{% endif %}

<!-- Live Logs -->
<div class="card">
    <div class="card-header bg-dark text-white">
        <h4><i class="fas fa-terminal"></i> Live Audit Logs</h4>
        <div class="btn-group btn-group-sm" role="group">
            <button class="btn btn-outline-secondary btn-sm" id="clearLogsBtn">Clear Logs</button>
            <button class="btn btn-outline-secondary btn-sm" id="toggleAutoScrollBtn">Auto-scroll: ON</button>
        </div>
    </div>
    <div class="card-body">
        <div class="log-container p-3" id="logContainer" style="height: 400px; overflow-y: auto; font-family: monospace; font-size: 0.85rem; background-color: #1e1e1e; color: #f0f0f0;">
            {% for log in logs %}
                <div class="log-entry">
                    {% if '[ROUTER:' in log %}
                        <span style="color: #4FC1FF;">{{ log }}</span>
                    {% elif '[JUMP' in log %}
                        <span style="color: #C586C0;">{{ log }}</span>
                    {% elif 'Error' in log or 'Failed' in log or 'FAILED' in log %}
                        <span style="color: #F14C4C;">{{ log }}</span>
                    {% elif 'Warning' in log or 'WARN' in log %}
                        <span style="color: #FFCC00;">{{ log }}</span>
                    {% elif 'Success' in log or 'OK' in log %}
                        <span style="color: #6A9955;">{{ log }}</span>
                    {% else %}
                        <span>{{ log }}</span>
                    {% endif %}
                </div>
            {% endfor %}
        </div>
    </div>
</div>
{% endblock %}

{% block custom_scripts %}
<script>
    // Log handling and auto-scroll functionality
    document.addEventListener('DOMContentLoaded', function() {
        const logContainer = document.getElementById('logContainer');
        const clearLogsBtn = document.getElementById('clearLogsBtn');
        const toggleAutoScrollBtn = document.getElementById('toggleAutoScrollBtn');
        let autoScrollEnabled = true;
        
        // Function to scroll logs to bottom
        function scrollLogsToBottom() {
            if (autoScrollEnabled && logContainer) {
                logContainer.scrollTop = logContainer.scrollHeight;
            }
        }
        
        // Initial scroll to bottom
        scrollLogsToBottom();
        
        // WebSocket handling for real-time log updates with enhanced reliability
        // Configure socket with better reconnection settings
        let socket;
        try {
            // Check if io is defined before using it
            if (typeof io !== 'undefined') {
                socket = io('/', {
                    reconnection: true,            // Enable reconnection
                    reconnectionAttempts: Infinity, // Keep trying to reconnect indefinitely
                    reconnectionDelay: 1000,       // Start with 1s delay
                    reconnectionDelayMax: 5000,    // Maximum 5s delay
                    timeout: 20000,                // Longer connection timeout
                    autoConnect: true              // Connect automatically
                });
            } else {
                console.error('Socket.IO (io) is not defined. Make sure socket.io.js is loaded.');
                // Create a dummy socket object to prevent errors
                socket = {
                    on: function() {},
                    emit: function() {}
                };
            }
        } catch (e) {
            console.error('Error initializing Socket.IO:', e);
            // Create a dummy socket object to prevent errors
            socket = {
                on: function() {},
                emit: function() {}
            };
        }
        
        // Track connection state for UI feedback
        let isConnected = false;
        let reconnectAttempts = 0;
        const connectionStatusEl = document.createElement('div');
        connectionStatusEl.className = 'socket-connection-status';
        connectionStatusEl.style.position = 'fixed';
        connectionStatusEl.style.bottom = '10px';
        connectionStatusEl.style.right = '10px';
        connectionStatusEl.style.padding = '5px 10px';
        connectionStatusEl.style.borderRadius = '3px';
        connectionStatusEl.style.fontSize = '12px';
        connectionStatusEl.style.zIndex = '1000';
        connectionStatusEl.style.display = 'none'; // Hidden by default
        document.body.appendChild(connectionStatusEl);
        
        // Debug and manage socket connection events
        socket.on('connect', function() {
            console.log('Socket.IO connected successfully');
            isConnected = true;
            reconnectAttempts = 0;
            
            // Update connection status indicator
            connectionStatusEl.textContent = 'Connected';
            connectionStatusEl.style.backgroundColor = '#28a745';
            connectionStatusEl.style.color = 'white';
            
            // Hide the indicator after 3 seconds
            setTimeout(() => {
                connectionStatusEl.style.display = 'none';
            }, 3000);
            
            // Request latest state on reconnection
            socket.emit('request_full_state', {}, function(response) {
                console.log('Received full state after reconnection:', response);
                // Update UI with latest state
                if (response && response.logs) {
                    updateLogDisplay(response.logs);
                }
                if (response && response.progress) {
                    updateProgressDisplay(response.progress);
                }
            });
        });
        
        socket.on('disconnect', function(reason) {
            console.warn('Socket.IO disconnected:', reason);
            isConnected = false;
            
            // Show connection status
            connectionStatusEl.textContent = 'Disconnected - Attempting to reconnect...';
            connectionStatusEl.style.backgroundColor = '#dc3545';
            connectionStatusEl.style.color = 'white';
            connectionStatusEl.style.display = 'block';
        });
        
        socket.on('reconnecting', function(attemptNumber) {
            reconnectAttempts = attemptNumber;
            console.log(`Socket.IO reconnection attempt ${attemptNumber}`);
            
            // Update reconnection status
            connectionStatusEl.textContent = `Reconnecting (Attempt ${attemptNumber})...`;
            connectionStatusEl.style.backgroundColor = '#ffc107';
            connectionStatusEl.style.color = 'black';
            connectionStatusEl.style.display = 'block';
        });
        
        socket.on('reconnect_failed', function() {
            console.error('Socket.IO reconnection failed after multiple attempts');
            
            // Update connection status to show permanent failure
            connectionStatusEl.textContent = 'Connection failed - Please refresh page';
            connectionStatusEl.style.backgroundColor = '#dc3545';
            connectionStatusEl.style.color = 'white';
            connectionStatusEl.style.display = 'block';
            
            // Alert user about refresh needed
            if (confirm('Lost connection to the server. Would you like to refresh the page?')) {
                window.location.reload();
            }
        });
        
        socket.on('connect_error', function(error) {
            console.error('Socket.IO connection error:', error);
            
            // Only show the error if we've tried multiple times
            if (reconnectAttempts > 3) {
                connectionStatusEl.textContent = 'Connection error - Retrying...';
                connectionStatusEl.style.backgroundColor = '#dc3545';
                connectionStatusEl.style.color = 'white';
                connectionStatusEl.style.display = 'block';
            }
        });
        
        // Helper function to update log display
        function updateLogDisplay(logs) {
            if (!logContainer || !logs) return;
            
            // Clear existing logs
            while (logContainer.firstChild) {
                logContainer.removeChild(logContainer.firstChild);
            }
            
            // Add new logs with proper formatting
            logs.forEach(function(log) {
                const logEntry = document.createElement('div');
                logEntry.className = 'log-entry';
                
                const logSpan = document.createElement('span');
                if (log.includes('[ROUTER:')) {
                    logSpan.style.color = '#4FC1FF';
                } else if (log.includes('[JUMP')) {
                    logSpan.style.color = '#C586C0';
                } else if (log.includes('Error') || log.includes('Failed') || log.includes('FAILED')) {
                    logSpan.style.color = '#F14C4C';
                } else if (log.includes('Warning') || log.includes('WARN')) {
                    logSpan.style.color = '#FFCC00';
                } else if (log.includes('Success') || log.includes('OK')) {
                    logSpan.style.color = '#6A9955';
                }
                
                logSpan.textContent = log;
                logEntry.appendChild(logSpan);
                logContainer.appendChild(logEntry);
            });
            
            // Scroll to bottom if auto-scroll is enabled
            scrollLogsToBottom();
        }
        
        // Helper function to update progress display
        function updateProgressDisplay(progressData) {
            // Update all progress elements based on the data
            console.log('Updating progress display with:', progressData);
            // Implementation depends on your UI structure
        }
        
        socket.on('log_update', function(data) {
            if (logContainer && data.logs) {
                // Clear existing logs
                while (logContainer.firstChild) {
                    logContainer.removeChild(logContainer.firstChild);
                }
                
                // Add new logs with proper formatting
                data.logs.forEach(function(log) {
                    const logEntry = document.createElement('div');
                    logEntry.className = 'log-entry';
                    
                    const logSpan = document.createElement('span');
                    if (log.includes('[ROUTER:')) {
                        logSpan.style.color = '#4FC1FF';
                    } else if (log.includes('[JUMP')) {
                        logSpan.style.color = '#C586C0';
                    } else if (log.includes('Error') || log.includes('Failed') || log.includes('FAILED')) {
                        logSpan.style.color = '#F14C4C';
                    } else if (log.includes('Warning') || log.includes('WARN')) {
                        logSpan.style.color = '#FFCC00';
                    } else if (log.includes('Success') || log.includes('OK')) {
                        logSpan.style.color = '#6A9955';
                    }
                    
                    logSpan.textContent = log;
                    logEntry.appendChild(logSpan);
                    logContainer.appendChild(logEntry);
                });
                
                // Scroll to bottom if auto-scroll is enabled
                scrollLogsToBottom();
            }
        });
        
        // Handle progress updates from the server
        socket.on('progress_update', function(progressData) {
            console.log('Progress update received via socket:', progressData);
            
            // Debug comprehensive progress data
            console.group('Enhanced Progress Tracking Update');
            console.log('Status:', progressData.status);
            console.log('Current Device:', progressData.current_device);
            console.log('Progress:', `${progressData.completed_devices || 0}/${progressData.total_devices || 0}`);
            console.log('Success Count:', progressData.overall_success_count || 0);
            console.log('Warning Count:', progressData.overall_warning_count || 0);
            console.log('Failure Count:', progressData.overall_failure_count || 0);
            console.groupEnd();
            
            if (!progressData) {
                console.warn('Received empty progressData via socket.');
                return;
            }
            
            try {
                // --- Update "Enhanced Progress Tracking" section ---
                // Update progress bar with appropriate styling based on status
                const enhancedProgressBar = document.getElementById('enhanced-progress-bar');
                if (enhancedProgressBar) {
                    const percent = progressData.total_devices > 0 ? (progressData.completed_devices / progressData.total_devices * 100) : 0;
                    enhancedProgressBar.style.width = percent.toFixed(1) + '%';
                    enhancedProgressBar.setAttribute('aria-valuenow', percent.toFixed(1));
                    enhancedProgressBar.textContent = percent.toFixed(1) + '%';
                    
                    // Update progress bar styling based on status
                    enhancedProgressBar.classList.remove('bg-success', 'bg-warning', 'bg-info', 'bg-secondary');
                    
                    if (progressData.status === 'completed') {
                        enhancedProgressBar.classList.add('bg-success');
                        enhancedProgressBar.classList.remove('progress-bar-animated');
                    } else if (progressData.status === 'paused') {
                        enhancedProgressBar.classList.add('bg-warning');
                        enhancedProgressBar.classList.remove('progress-bar-animated');
                    } else if (progressData.status === 'running') {
                        enhancedProgressBar.classList.add('bg-info');
                        enhancedProgressBar.classList.add('progress-bar-animated');
                    } else {
                        enhancedProgressBar.classList.add('bg-secondary');
                        enhancedProgressBar.classList.remove('progress-bar-animated');
                    }
                    
                    console.log(`Updated progress bar: ${percent.toFixed(1)}% (${progressData.status})`);
                } else {
                    console.warn('Could not find enhanced-progress-bar element');
                }
                
                // Status with appropriate styling
                const progressStatusEl = document.getElementById('progress-status');
                if (progressStatusEl) {
                    // Get status from status_message for better user-friendly display
                    const statusText = progressData.status_message || progressData.status || 'Idle';
                    progressStatusEl.textContent = statusText;
                    
                    // Apply appropriate styling based on status
                    progressStatusEl.classList.remove('text-success', 'text-warning', 'text-info', 'text-secondary', 'text-danger');
                    
                    if (progressData.status === 'completed') {
                        progressStatusEl.classList.add('text-success');
                    } else if (progressData.status === 'paused') {
                        progressStatusEl.classList.add('text-warning');
                    } else if (progressData.status === 'running') {
                        progressStatusEl.classList.add('text-info');
                    } else if (progressData.status === 'error') {
                        progressStatusEl.classList.add('text-danger');
                    } else {
                        progressStatusEl.classList.add('text-secondary');
                    }
                    
                    console.log(`Updated status: ${statusText}`);
                } else {
                    console.warn('Could not find progress-status element');
                }
                
                // Elapsed time
                const progressElapsedEl = document.getElementById('progress-elapsed');
                if (progressElapsedEl) {
                    let elapsedDisplay = '00:00:00';
                    
                    if (progressData.start_time) {
                        try {
                            const startTime = new Date(progressData.start_time);
                            const now = (progressData.status === 'completed' && progressData.end_time) ? 
                                        new Date(progressData.end_time) : new Date();
                            const elapsedMs = now - startTime;
                            
                            if (elapsedMs >= 0) {
                                const hours = Math.floor(elapsedMs / 3600000).toString().padStart(2, '0');
                                const minutes = Math.floor((elapsedMs % 3600000) / 60000).toString().padStart(2, '0');
                                const seconds = Math.floor((elapsedMs % 60000) / 1000).toString().padStart(2, '0');
                                elapsedDisplay = `${hours}:${minutes}:${seconds}`;
                            }
                        } catch (timeErr) {
                            console.error('Error calculating elapsed time:', timeErr);
                        }
                    }
                    
                    progressElapsedEl.textContent = elapsedDisplay;
                    console.log(`Updated elapsed time: ${elapsedDisplay}`);
                } else {
                    console.warn('Could not find progress-elapsed element');
                }
                
                // Completion count
                const progressCompletedEl = document.getElementById('progress-completed');
                if (progressCompletedEl) {
                    const completedText = `${progressData.completed_devices || 0}/${progressData.total_devices || 0}`;
                    progressCompletedEl.textContent = completedText;
                    console.log(`Updated completion count: ${completedText}`);
                } else {
                    console.warn('Could not find progress-completed element');
                }
                
                // ETA
                const progressEtaEl = document.getElementById('progress-eta');
                if (progressEtaEl) {
                    let etaDisplay = '--:--:--';
                    
                    if (progressData.estimated_completion_time && progressData.status === 'running') {
                        try {
                            const eta = new Date(progressData.estimated_completion_time);
                            const hours = eta.getHours().toString().padStart(2, '0');
                            const minutes = eta.getMinutes().toString().padStart(2, '0');
                            const seconds = eta.getSeconds().toString().padStart(2, '0');
                            etaDisplay = `${hours}:${minutes}:${seconds}`;
                        } catch (etaErr) {
                            console.error('Error parsing ETA:', etaErr);
                        }
                    }
                    
                    progressEtaEl.textContent = etaDisplay;
                    console.log(`Updated ETA: ${etaDisplay}`);
                } else {
                    console.warn('Could not find progress-eta element');
                }
                
                // Current device with appropriate styling based on audit state
                const currentDeviceEl = document.getElementById('current-device');
                if (currentDeviceEl) {
                    const deviceName = progressData.current_device || 'None';
                    currentDeviceEl.textContent = deviceName;
                    
                    // Update the device container styling based on audit status
                    const deviceContainer = currentDeviceEl.closest('.alert');
                    if (deviceContainer) {
                        deviceContainer.classList.remove('alert-info', 'alert-success', 'alert-warning', 'alert-secondary', 'alert-danger');
                        
                        if (progressData.status === 'running') {
                            deviceContainer.classList.add('alert-info');
                        } else if (progressData.status === 'completed') {
                            deviceContainer.classList.add('alert-success');
                        } else if (progressData.status === 'paused') {
                            deviceContainer.classList.add('alert-warning');
                        } else if (progressData.status === 'error') {
                            deviceContainer.classList.add('alert-danger');
                        } else {
                            deviceContainer.classList.add('alert-secondary');
                        }
                    }
                    
                    console.log(`Updated current device: ${deviceName}`);
                    
                    // Also update the device progress bar with appropriate styling
                    const deviceProgressBar = document.getElementById('device-progress-bar');
                    if (deviceProgressBar) {
                        deviceProgressBar.classList.remove('bg-info', 'bg-success', 'bg-warning', 'bg-secondary', 'bg-danger');
                        
                        if (progressData.status === 'running') {
                            deviceProgressBar.style.width = '100%';
                            deviceProgressBar.classList.add('bg-info', 'progress-bar-animated');
                            deviceProgressBar.textContent = 'In Progress';
                        } else if (progressData.status === 'completed') {
                            deviceProgressBar.style.width = '100%';
                            deviceProgressBar.classList.add('bg-success');
                            deviceProgressBar.classList.remove('progress-bar-animated');
                            deviceProgressBar.textContent = 'Completed';
                        } else if (progressData.status === 'paused') {
                            deviceProgressBar.style.width = '100%';
                            deviceProgressBar.classList.add('bg-warning');
                            deviceProgressBar.classList.remove('progress-bar-animated');
                            deviceProgressBar.textContent = 'Paused';
                        } else if (progressData.status === 'error') {
                            deviceProgressBar.style.width = '100%';
                            deviceProgressBar.classList.add('bg-danger');
                            deviceProgressBar.classList.remove('progress-bar-animated');
                            deviceProgressBar.textContent = 'Error';
                        } else {
                            deviceProgressBar.style.width = '0%';
                            deviceProgressBar.classList.add('bg-secondary');
                            deviceProgressBar.classList.remove('progress-bar-animated');
                            deviceProgressBar.textContent = 'Not Started';
                        }
                    }
                } else {
                    console.warn('Could not find current-device element');
                }
                
                // Status counts - using the correct field names from the backend
                const successCountEl = document.getElementById('success-count');
                if (successCountEl) {
                    const successCount = progressData.overall_success_count || 0;
                    successCountEl.textContent = successCount;
                    console.log(`Updated success count: ${successCount}`);
                } else {
                    console.warn('Could not find success-count element');
                }
                
                const warningCountEl = document.getElementById('warning-count');
                if (warningCountEl) {
                    const warningCount = progressData.overall_warning_count || 0;
                    warningCountEl.textContent = warningCount;
                    console.log(`Updated warning count: ${warningCount}`);
                } else {
                    console.warn('Could not find warning-count element');
                }
                
                const failureCountEl = document.getElementById('failure-count');
                if (failureCountEl) {
                    const failureCount = progressData.overall_failure_count || 0;
                    failureCountEl.textContent = failureCount;
                    console.log(`Updated failure count: ${failureCount}`);
                } else {
                    console.warn('Could not find failure-count element');
                }
            } catch (e) {
                console.error("Error updating Enhanced Progress Tracking (socket):", e, e.stack);
            }
            
            try {
                // --- Update "Current Audit Status" (older/simpler section on top) ---
                const oldProgressBar = document.querySelector('.progress-bar:not(#enhanced-progress-bar):not(#device-progress-bar)');
                if (oldProgressBar) {
                    const percentage = progressData.total_devices > 0 ? (progressData.completed_devices / progressData.total_devices * 100) : 0;
                    oldProgressBar.style.width = `${percentage.toFixed(1)}%`;
                    oldProgressBar.setAttribute('aria-valuenow', percentage.toFixed(1));
                    oldProgressBar.textContent = `${percentage.toFixed(1)}%`;
                }
                
                const progressPercentageDisplay = document.getElementById('progress-percentage');
                if (progressPercentageDisplay) {
                    const percentage = progressData.total_devices > 0 ? (progressData.completed_devices / progressData.total_devices * 100) : 0;
                    progressPercentageDisplay.textContent = `${Math.round(percentage)}`;
                }
                
                const currentDeviceDisplayOld = document.getElementById('current-device-display');
                if (currentDeviceDisplayOld) currentDeviceDisplayOld.textContent = progressData.current_device || 'N/A';
                
                const statusMessageOld = document.getElementById('status-message');
                if (statusMessageOld) statusMessageOld.textContent = progressData.status_message || 'Idle';
                
                const devicesProcessedOld = document.getElementById('devices-processed');
                if (devicesProcessedOld) devicesProcessedOld.textContent = `${progressData.completed_devices || 0} / ${progressData.total_devices || 0}`;
            } catch (e) {
                console.error("Error updating Current Audit Status (socket):", e);
            }

                
                // Update main progress bar
                const progressBar = document.querySelector('.progress-bar');
                if (progressBar) {
                    const percentage = progress.completed_devices / progress.total_devices * 100 || 0;
                    progressBar.style.width = `${percentage}%`;
                    progressBar.setAttribute('aria-valuenow', percentage);
                    progressBar.textContent = `${Math.round(percentage)}%`;
                }
                
                // Update main progress display
                const progressPercentage = document.getElementById('progress-percentage');
                if (progressPercentage) {
                    const percentage = progress.completed_devices / progress.total_devices * 100 || 0;
                    progressPercentage.textContent = `${Math.round(percentage)}%`;
                }
                
                // Update current device and status message
                const currentDeviceDisplay = document.getElementById('current-device-display');
                if (currentDeviceDisplay) {
                    currentDeviceDisplay.textContent = progress.current_device || 'N/A';
                }
                
                const statusMessage = document.getElementById('status-message');
                if (statusMessage) {
                    statusMessage.textContent = progress.status_message || 'Idle';
                }
                
                const devicesProcessed = document.getElementById('devices-processed');
                if (devicesProcessed) {
                    devicesProcessed.textContent = `${progress.completed_devices || 0} / ${progress.total_devices || 0}`;
                }
            }
        });
        
        // Stop/Reset Audit button handler
        const stopResetAuditBtn = document.getElementById('stop-reset-audit');
        if (stopResetAuditBtn) {
            stopResetAuditBtn.addEventListener('click', function(e) {
                e.preventDefault();
                
                if (confirm('Are you sure you want to stop and reset the current audit process?')) {
                    // Show loading indicator or disable button
                    stopResetAuditBtn.disabled = true;
                    stopResetAuditBtn.innerHTML = '<i class="fas fa-spinner fa-spin"></i> Processing...';
                    
                    fetch('/stop_reset_audit', {
                        method: 'POST',
                        headers: {
                            'Content-Type': 'application/json'
                        }
                    })
                    .then(response => response.json())
                    .then(data => {
                        if (data.status === 'success') {
                            // Provide specific feedback based on the message
                            if (data.message.includes('No audit was running')) {
                                alert('No audit was running. The dashboard will be reset.');
                            } else if (data.message.includes('Completed audit')) {
                                alert('The completed audit has been reset. The dashboard will refresh.');
                            } else {
                                alert(data.message);
                            }
                            
                            console.log('Stop/Reset Audit response:', data);
                            // Refresh the page to update UI
                            window.location.reload();
                        } else {
                            alert('Error: ' + data.message);
                            // Reset button state
                            stopResetAuditBtn.disabled = false;
                            stopResetAuditBtn.innerHTML = 'Stop/Reset Audit';
                        }
                    })
                    .catch(error => {
                        console.error('Error stopping audit:', error);
                        alert('An error occurred while trying to stop the audit. Please check the console for details.');
                        // Reset button state
                        stopResetAuditBtn.disabled = false;
                        stopResetAuditBtn.innerHTML = 'Stop/Reset Audit';
                    });
                }
            });
        }
        
        // Clear logs button
        if (clearLogsBtn) {
            clearLogsBtn.addEventListener('click', function() {
                if (logContainer) {
                    while (logContainer.firstChild) {
                        logContainer.removeChild(logContainer.firstChild);
                    }
                    // Send request to clear logs on server
                    fetch('/clear_logs', { method: 'POST' });
                }
            });
        }
        
        // Toggle auto-scroll button
        if (toggleAutoScrollBtn) {
            toggleAutoScrollBtn.addEventListener('click', function() {
                autoScrollEnabled = !autoScrollEnabled;
                toggleAutoScrollBtn.textContent = 'Auto-scroll: ' + (autoScrollEnabled ? 'ON' : 'OFF');
                if (autoScrollEnabled) {
                    scrollLogsToBottom();
                }
            });
        }
        
        // Create chart if data exists
        var ctx = document.getElementById('resultsChart');
        // Use JSON.parse with a string to avoid Jinja2 syntax issues
        var chartDataStr = '{{ chart_data|tojson|safe }}';
        var chartData = null;
        try {
            // Ensure the string is properly formatted before parsing
            if (chartDataStr && typeof chartDataStr === 'string') {
                chartData = JSON.parse(chartDataStr);
            } else if (typeof chartDataStr === 'object') {
                // In case it's already an object
                chartData = chartDataStr;
            }
        } catch (e) {
            console.error('Error parsing chart data:', e);
            // Provide fallback data
            chartData = {"labels":["No Data"], "values":[1], "colors":["#D3D3D3"]};
        }
        
        if (ctx && chartData) {
            new Chart(ctx, {
                type: 'pie',
                data: {
                    labels: chartData.labels,
                    datasets: [{
                        data: chartData.values,
                        backgroundColor: chartData.colors,
                        borderWidth: 1
                    }]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    legend: {
                        position: 'right'
                    }
                }
            });
        }
    });
</script>
{% endblock %}
"""
# Template for Captured Configs page
HTML_CAPTURED_CONFIGS_TEMPLATE = """{% extends "base_layout.html" %}
{% block content %}
<div class="container mt-4">
    <h2><i class="fas fa-file-code"></i> Captured Router Configurations</h2>
    
    <div class="card mb-4">
        <div class="card-header bg-primary text-white">
            <h5><i class="fas fa-info-circle"></i> About Router Configurations</h5>
        </div>
        <div class="card-body">
            <p>This page displays all captured router configurations from previous audit runs. You can view the configuration files for each router and compare changes over time.</p>
        </div>
    </div>
    
    {% if sorted_router_names|length > 0 %}
        <div class="accordion" id="routerConfigsAccordion">
            {% for router_name in sorted_router_names %}
                <div class="card">
                    <div class="card-header" id="heading{{ loop.index }}">
                        <h2 class="mb-0">
                            <button class="btn btn-link btn-block text-left" type="button" data-toggle="collapse" data-target="#collapse{{ loop.index }}" aria-expanded="{% if loop.first %}true{% else %}false{% endif %}" aria-controls="collapse{{ loop.index }}">
                                <i class="fas fa-router mr-2"></i> {{ router_name }}
                                <span class="badge badge-info ml-2">{{ router_configs[router_name]|length }} configs</span>
                            </button>
                        </h2>
                    </div>
                    
                    <div id="collapse{{ loop.index }}" class="collapse {% if loop.first %}show{% endif %}" aria-labelledby="heading{{ loop.index }}" data-parent="#routerConfigsAccordion">
                        <div class="card-body">
                            <div class="mb-3">
                                <a href="/download_all_configs/{{ router_name }}" class="btn btn-success">
                                    <i class="fas fa-download"></i> Download All {{ router_name }} Configs
                                </a>
                            </div>
                            <div class="mb-3">
                                <a href="/download_all_configs/{{ router_name }}" class="btn btn-success">
                                    <i class="fas fa-download"></i> Download All {{ router_name }} Configs
                                </a>
                            </div>
                            <div class="table-responsive">
                                <table class="table table-striped">
                                    <thead>
                                        <tr>
                                            <th>Date</th>
                                            <th>Filename</th>
                                            <th>Size</th>
                                            <th>Actions</th>
                                            </tr>
                                    </thead>
                                    <tbody>
                                        {% for config in router_configs[router_name] %}
                                            <tr>
                                                <td>{{ config.audit_date }}</td>
                                                <td>{{ config.filename }}</td>
                                                <td>{{ (config.filesize / 1024)|round(1) }} KB</td>
                                                <td>
                                                    <div class="btn-group">
                                                        <a href="/view_config/{{ config.audit_date }}/{{ config.filename }}" class="btn btn-sm btn-primary">
                                                            <i class="fas fa-eye"></i> View
                                                        </a>
                                                        <a href="/download_config/{{ config.audit_date }}/{{ config.filename }}" class="btn btn-sm btn-success">
                                                            <i class="fas fa-download"></i> Download
                                                        </a>
                                                    </div>
                                                </td>
                                            </tr>
                                        {% endfor %}
                                    </tbody>
                                </table>
                            </div>
                        </div>
                    </div>
                </div>
            {% endfor %}
        </div>
    {% else %}
        <div class="alert alert-info">
            <i class="fas fa-info-circle"></i> No router configurations have been captured yet. Run an audit to collect router configurations.
        </div>
    {% endif %}
</div>
{% endblock %}
"""

# Template for View Config page
HTML_VIEW_CONFIG_TEMPLATE = """
{% extends "base_layout.html" %}
{% block content %}
<div class="container mt-4">
    <div class="d-flex justify-content-between align-items-center mb-3">
        <h2><i class="fas fa-file-code"></i> Router Configuration</h2>
        <a href="/captured_configs" class="btn btn-secondary">
            <i class="fas fa-arrow-left"></i> Back to All Configs
        </a>
    </div>
    
    <div class="card mb-4">
        <div class="card-header bg-primary text-white">
            <div class="d-flex justify-content-between align-items-center">
                <h5><i class="fas fa-info-circle"></i> Configuration Details</h5>
            </div>
        </div>
        <div class="card-body">
            <div class="row">
                <div class="col-md-6">
                    <p><strong>Router:</strong> {{ router_name }}</p>
                    <p><strong>Filename:</strong> {{ filename }}</p>
                </div>
                <div class="col-md-6">
                    <p><strong>Last Modified:</strong> {{ last_modified.strftime('%Y-%m-%d %H:%M:%S') }}</p>
                    <p><strong>Size:</strong> {{ (filesize / 1024)|round(1) }} KB</p>
                </div>
            </div>
            <div class="mt-3">
                <a href="/download_config/{{ audit_date }}/{{ filename }}" class="btn btn-success">
                    <i class="fas fa-download"></i> Download Configuration
                </a>
            </div>
        </div>
    </div>
    
    <div class="card">
        <div class="card-header bg-dark text-white">
            <div class="d-flex justify-content-between align-items-center">
                <h5><i class="fas fa-code"></i> Configuration Content</h5>
                <div>
                    <button class="btn btn-sm btn-outline-light" id="btn-copy-config">
                        <i class="fas fa-copy"></i> Copy
                    </button>
                </div>
            </div>
        </div>
        <div class="card-body p-0">
            <pre id="config-content" class="p-3 mb-0 bg-light" style="max-height: 600px; overflow-y: auto;">{{ config_content }}</pre>
        </div>
    </div>
</div>

<script>
    document.addEventListener('DOMContentLoaded', function() {
        // Add copy button functionality
        document.getElementById('btn-copy-config').addEventListener('click', function() {
            const configContent = document.getElementById('config-content').textContent;
            navigator.clipboard.writeText(configContent).then(function() {
                alert('Configuration copied to clipboard!');
            }, function() {
                alert('Failed to copy configuration. Please select and copy manually.');
            });
        });
    });
</script>
{% endblock %}
"""

HTML_SETTINGS_TEMPLATE_CONTENT = """
{% extends "base_layout.html" %}
{% block title %}Settings - Router Audit & Terminal Pro{% endblock %}
{% block content %}
<h1><i class="fas fa-cogs"></i> Application Settings</h1>
<p class="lead">Configure jump host details, default device credentials, and global inventory settings.</p>

<div class="card mt-4">
    <div class="card-header"><h4><i class="fas fa-server"></i> Jump Host & Device Credentials</h4></div>
    <div class="card-body">
        <form method="POST" action="{{ url_for('settings_route') }}">
            <h5>Jump Host Configuration</h5>
            <div class="form-row">
                <div class="form-group col-md-6"><label for="jump_host">Jump Host IP/Hostname:</label><input type="text" class="form-control" id="jump_host" name="jump_host" value="{{ config.JUMP_HOST }}"></div>
                <div class="form-group col-md-6"><label for="jump_ping_path">Jump Host Ping Executable Path:</label><input type="text" class="form-control" id="jump_ping_path" name="jump_ping_path" value="{{ config.JUMP_PING_PATH }}" placeholder="e.g., /bin/ping"></div>
            </div>
            <div class="form-row">
                <div class="form-group col-md-6"><label for="jump_username">Jump Host Username:</label><input type="text" class="form-control" id="jump_username" name="jump_username" value="{{ config.JUMP_USERNAME }}"></div>
                <div class="form-group col-md-6"><label for="jump_password">Jump Host Password:</label><input type="password" class="form-control" id="jump_password" name="jump_password" placeholder="Leave blank to keep current"><small class="form-text text-muted">Passwords stored in .env file.</small></div>
            </div>
            <hr>
            <h5>Default Device Credentials (used if not in inventory)</h5>
            <div class="form-row">
                <div class="form-group col-md-4"><label for="device_username">Default Device Username:</label><input type="text" class="form-control" id="device_username" name="device_username" value="{{ config.DEVICE_USERNAME }}"></div>
                <div class="form-group col-md-4"><label for="device_password">Default Device Password:</label><input type="password" class="form-control" id="device_password" name="device_password" placeholder="Leave blank to keep current"></div>
                <div class="form-group col-md-4"><label for="device_enable_password">Default Device Enable Password:</label><input type="password" class="form-control" id="device_enable_password" name="device_enable_password" placeholder="Leave blank or empty to clear"></div>
            </div>
            <hr>
             <h5>Global Inventory Settings</h5>
             <div class="form-group">
                <label>Inventory Format:</label>
                <p class="form-control-static">CSV (.csv) <small class="text-muted">- The application now exclusively uses CSV format for inventories</small></p>
            </div>
            <button type="submit" class="btn btn-primary mt-3"><i class="fas fa-save"></i> Save All Settings</button>
        </form>
    </div>
</div>
{% endblock %}
"""
HTML_VIEW_JSON_TEMPLATE_CONTENT = """{% extends "base_layout.html" %} ...""" # (Same as before)
HTML_EDIT_INVENTORY_TEMPLATE_CONTENT = """
{% extends "base_layout.html" %}
{% block title %}Manage Inventories{% endblock %}
{% block content %}
<h1><i class="fas fa-tasks"></i> Manage Inventory Files</h1>
<p class="lead">Upload new inventory files (YAML or CSV), select an existing file to be active, edit content, or export the current active inventory to CSV.</p>
<style>
    .filter-card {
        margin-bottom: 20px;
        border-left: 4px solid #007bff;
    }
    .filter-header {
        display: flex;
        justify-content: space-between;
        align-items: center;
    }
    #filter-results-count {
        font-size: 0.9rem;
        font-weight: 500;
        color: #6c757d;
    }
    #no-filter-results {
        display: none;
        padding: 20px;
        text-align: center;
        background-color: #f8f9fa;
        border-radius: 4px;
        margin-top: 15px;
    }
    .hostname-cell, .ip-cell, .device-type-cell {
        font-weight: 500;
    }
</style>

<div class="card mt-4 filter-card">
    <div class="card-header filter-header">
        <h4><i class="fas fa-filter"></i> Filter Inventory</h4>
        <span id="filter-results-count"></span>
    </div>
    <div class="card-body">
        <form id="inventory-filter-form">
            <div class="form-row">
                <div class="form-group col-md-4">
                    <label for="hostname-filter">Hostname</label>
                    <input type="text" class="form-control" id="hostname-filter" placeholder="Filter by hostname...">
                </div>
                <div class="form-group col-md-4">
                    <label for="ip-filter">IP Address</label>
                    <input type="text" class="form-control" id="ip-filter" placeholder="Filter by IP address...">
                </div>
                <div class="form-group col-md-4">
                    <label for="device-type-filter">Device Type</label>
                    <select class="form-control" id="device-type-filter">
                        <option value="all">All Device Types</option>
                    </select>
                </div>
            </div>
            <div class="form-row">
                <div class="col-md-6">
                    <button type="submit" class="btn btn-primary" id="apply-filters"><i class="fas fa-search"></i> Apply Filters</button>
                    <button type="button" class="btn btn-secondary" id="reset-filters"><i class="fas fa-undo"></i> Reset</button>
                </div>
            </div>
        </form>
        <div id="no-filter-results" class="mt-3">
            <i class="fas fa-exclamation-circle text-warning"></i> No devices match the current filters. Try adjusting your criteria.
        </div>
    </div>
</div>

<div class="card mt-4">
    <div class="card-header"><h4><i class="fas fa-list-alt"></i> Active Inventory & Available Files</h4></div>
    <div class="card-body">
        <form method="POST" action="{{ url_for('manage_inventories_route') }}">
            <input type="hidden" name="action" value="set_active">
            <div class="form-row align-items-end">
                <div class="form-group col-md-8">
                    <label for="active_inventory_file_manage"><strong>Current Active Inventory:</strong> {{ active_inventory }} (Format: {{ active_inventory_format.upper() }})</label>
                    <br>
                    <label for="active_inventory_file_manage">Select an inventory file to make active:</label>
                    <select class="form-control" id="active_inventory_file_manage" name="active_inventory_file_manage">
                        {% for inv_file in inventories %}
                        <option value="{{ inv_file }}" {% if inv_file == active_inventory %}selected{% endif %}>{{ inv_file }}</option>
                        {% endfor %}
                    </select>
                </div>
                <div class="form-group col-md-4">
                     <button type="submit" class="btn btn-info btn-block"><i class="fas fa-check-circle"></i> Set as Active</button>
                </div>
            </div>
        </form>
        <hr>
        <a href="{{ url_for('export_inventory_csv_route') }}" class="btn btn-success mt-2"><i class="fas fa-file-csv"></i> Export Current Active Inventory to CSV</a>
    </div>
</div>

<div class="card mt-4">
    <div class="card-header"><h4><i class="fas fa-upload"></i> Upload New Inventory File</h4></div>
    <div class="card-body">
        <form method="POST" action="{{ url_for('manage_inventories_route') }}" enctype="multipart/form-data">
            <input type="hidden" name="action" value="upload">
            <div class="form-group">
                <label for="inventory_file_upload_manage">Upload CSV (.csv) inventory file:</label>
                <input type="file" class="form-control-file" id="inventory_file_upload_manage" name="inventory_file_upload_manage" accept=".csv">
                <small class="form-text text-muted">
                    File will be validated and versioned (e.g., inventory-list-v01.csv). 
                    If valid, it will be set as active.
                </small>
            </div>
            <button type="submit" class="btn btn-success"><i class="fas fa-cloud-upload-alt"></i> Upload and Set Active</button>
        </form>
    </div>
</div>

<div class="card mt-4">
    <div class="card-header">
        <h4><i class="fas fa-file-alt"></i> Edit Current Active Inventory ({{ active_inventory }} - CSV)</h4>
    </div>
    <div class="card-body">
        <ul class="nav nav-tabs" id="editorTabs" role="tablist">
            <li class="nav-item">
                <a class="nav-link active" id="csv-tab" data-toggle="tab" href="#csv-editor" role="tab" aria-controls="csv-editor" aria-selected="true">Table View</a>
            </li>
            <li class="nav-item">
                <a class="nav-link" id="raw-tab" data-toggle="tab" href="#raw-editor" role="tab" aria-controls="raw-editor" aria-selected="false">Raw CSV</a>
            </li>
        </ul>
        
        <div class="tab-content" id="editorTabsContent">
            <!-- CSV Table Editor Tab -->
            <div class="tab-pane fade show active" id="csv-editor" role="tabpanel" aria-labelledby="csv-tab">
                <form id="csvTableForm" method="POST" action="{{ url_for('edit_active_inventory_content_route') }}">
                    <input type="hidden" name="inventory_content_edit" id="csv-data-hidden">
                    
                    <div class="table-responsive mt-3">
                        <table class="table table-bordered table-hover" id="csvTable">
                            <thead id="csvTableHeader">
                                <!-- Headers will be populated by JavaScript -->
                            </thead>
                            <tbody id="csvTableBody">
                                <!-- Rows will be populated by JavaScript -->
                            </tbody>
                        </table>
                    </div>
                    
                    <div class="btn-toolbar mb-3">
                        <div class="btn-group mr-2">
                            <button type="button" class="btn btn-sm btn-success" id="addRowBtn"><i class="fas fa-plus-circle"></i> Add Row</button>
                            <button type="button" class="btn btn-sm btn-warning" id="addColumnBtn"><i class="fas fa-columns"></i> Add Column</button>
                        </div>
                    </div>
                    
                    <button type="submit" class="btn btn-primary mt-3"><i class="fas fa-save"></i> Save Changes</button>
                    <a href="{{ url_for('index') }}" class="btn btn-secondary mt-3">Cancel</a>
                </form>
            </div>
            
            <!-- Raw CSV Editor Tab -->
            <div class="tab-pane fade" id="raw-editor" role="tabpanel" aria-labelledby="raw-tab">
                <form method="POST" action="{{ url_for('edit_active_inventory_content_route') }}">
                    <div class="form-group mt-3">
                        <textarea class="form-control" id="raw_inventory_content_edit" name="inventory_content_edit" rows="20" placeholder="Loading CSV content...">{{ current_inventory_content_raw }}</textarea>
                        <small class="form-text text-muted">CSV format with header row. Changes will be saved as a new CSV version.</small>
                    </div>
                    <button type="submit" class="btn btn-primary"><i class="fas fa-save"></i> Save Changes</button>
                    <a href="{{ url_for('index') }}" class="btn btn-secondary">Cancel</a>
                </form>
            </div>
        </div>
    </div>
</div>

<script>
// Helper function to get raw inventory content
function getRawInventoryContent() {
    const textarea = document.getElementById('raw_inventory_content_edit');
    return textarea ? textarea.value : '';
}

// Parse CSV string into array of arrays
function parseCSV(csvStr) {
    if (!csvStr || !csvStr.trim()) {
        console.error("Empty CSV string");
        return [];
    }
    
    console.log(`Parsing CSV string of length ${csvStr.length}`);
    const lines = csvStr.trim().split('\n');
    console.log(`Found ${lines.length} lines in CSV`);
    
    const result = [];
    for (let i = 0; i < lines.length; i++) {
        if (!lines[i].trim()) continue;
        
        // Simple CSV parsing - split by commas
        // This could be enhanced for quoted fields if needed
        const row = lines[i].split(',').map(cell => cell.trim());
        result.push(row);
    }
    
    return result;
}

// Populate the CSV table with data
function populateCSVTable(data) {
    if (!data || !data.length) {
        console.error("No data to populate table");
        return false;
    }
    
    const tableHeader = document.getElementById('csvTableHeader');
    const tableBody = document.getElementById('csvTableBody');
    
    if (!tableHeader || !tableBody) {
        console.error("Table elements not found");
        return false;
    }
    
    // Clear existing content
    tableHeader.innerHTML = '';
    tableBody.innerHTML = '';
    
    // Create header row
    const headerRow = document.createElement('tr');
    for (const header of data[0]) {
        const th = document.createElement('th');
        th.textContent = header;
        th.contentEditable = "true";
        headerRow.appendChild(th);
    }
    tableHeader.appendChild(headerRow);
    
    // Add data rows
    let rowCount = 0;
    for (let i = 1; i < data.length; i++) {
        const tr = document.createElement('tr');
        
        for (let j = 0; j < data[i].length; j++) {
            const td = document.createElement('td');
            td.textContent = data[i][j] || '';
            td.contentEditable = "true";
            
            // Add classes for filtering
            if (data[0][j] === 'hostname') {
                td.classList.add('hostname-cell');
            } else if (data[0][j] === 'ip') {
                td.classList.add('ip-cell');
            } else if (data[0][j] === 'device_type') {
                td.classList.add('device-type-cell');
            }
            
            tr.appendChild(td);
        }
        
        // Add delete button
        const deleteCell = document.createElement('td');
        const deleteBtn = document.createElement('button');
        deleteBtn.className = "btn btn-sm btn-danger";
        deleteBtn.innerHTML = '<i class="fas fa-trash"></i>';
        deleteBtn.addEventListener('click', () => {
            tr.remove();
            updateFilterResultsCount();
        });
        deleteCell.appendChild(deleteBtn);
        tr.appendChild(deleteCell);
        
        tableBody.appendChild(tr);
        rowCount++;
    }
    
    console.log(`Added ${rowCount} rows to the table`);
    updateFilterResultsCount();
    populateDeviceTypeDropdown();
    return rowCount > 0;
}

// Load inventory data from API and populate the table
async function loadInventoryDataFromAPI() {
    try {
        console.log('Fetching inventory data from API...');
        const response = await fetch('/get_active_inventory_info');
        
        if (!response.ok) {
            throw new Error(`API returned ${response.status}: ${response.statusText}`);
        }
        
        const data = await response.json();
        
        if (data.status === 'success' && data.data && data.headers) {
            console.log(`Successfully fetched inventory data with ${data.data.length} rows`);
            
            // Convert data to array format for table display
            const tableData = [];
            
            // Add headers row
            tableData.push(data.headers);
            
            // Add data rows
            for (const row of data.data) {
                const rowArray = [];
                for (const header of data.headers) {
                    rowArray.push(row[header] || '');
                }
                tableData.push(rowArray);
            }
            
            // Populate the table
            populateCSVTable(tableData);
            console.log('Table populated successfully from API data');
            return true;
        } else {
            console.error('API returned error:', data.message);
            return false;
        }
    } catch (error) {
        console.error('Error fetching inventory data:', error);
        return false;
    }
}

// Load from raw content textarea
function loadFromRawContent() {
    const rawContent = getRawInventoryContent();
    if (!rawContent) {
        console.error("Could not get raw inventory content");
        return false;
    }
    
    const parsedData = parseCSV(rawContent);
    if (!parsedData.length) {
        console.error("Failed to parse CSV data");
        return false;
    }
    
    return populateCSVTable(parsedData);
}

// Populate device type dropdown
function populateDeviceTypeDropdown() {
    const dropdown = document.getElementById('device-type-filter');
    if (!dropdown) {
        console.error("Device type dropdown not found");
        return false;
    }
    
    // Get all device type cells
    const cells = document.querySelectorAll('.device-type-cell');
    if (!cells.length) {
        console.error("No device type cells found");
        return false;
    }
    
    // Extract unique device types
    const types = new Set();
    cells.forEach(cell => {
        const type = cell.textContent.trim().toLowerCase();
        if (type) types.add(type);
    });
    
    // Build options HTML
    let options = '<option value="all">All Device Types</option>';
    types.forEach(type => {
        const capitalized = type.charAt(0).toUpperCase() + type.slice(1);
        options += `<option value="${type}">${capitalized}</option>`;
    });
    
    // Set dropdown HTML
    dropdown.innerHTML = options;
    console.log(`Populated device type dropdown with ${types.size} options`);
    return true;
}

// Update filter results count
function updateFilterResultsCount() {
    const countElement = document.getElementById('filter-results-count');
    if (!countElement) return;
    
    const tbody = document.getElementById('csvTableBody');
    if (!tbody) return;
    
    const totalRows = tbody.querySelectorAll('tr').length;
    const visibleRows = Array.from(tbody.querySelectorAll('tr'))
        .filter(row => row.style.display !== 'none').length;
    
    countElement.textContent = `Showing ${visibleRows} of ${totalRows} devices`;
    
    // Show/hide "no results" message
    const noResults = document.getElementById('no-filter-results');
    if (noResults) {
        noResults.style.display = visibleRows === 0 ? 'block' : 'none';
    }
}

// Apply filters function
function applyFilters() {
    const hostnameFilter = document.getElementById('hostname-filter').value.toLowerCase();
    const ipFilter = document.getElementById('ip-filter').value.toLowerCase();
    const deviceTypeFilter = document.getElementById('device-type-filter').value.toLowerCase();
    
    const tbody = document.getElementById('csvTableBody');
    if (!tbody) return;
    
    const rows = tbody.querySelectorAll('tr');
    let visibleCount = 0;
    
    rows.forEach(row => {
        const hostname = row.querySelector('.hostname-cell')?.textContent.toLowerCase() || '';
        const ip = row.querySelector('.ip-cell')?.textContent.toLowerCase() || '';
        const deviceType = row.querySelector('.device-type-cell')?.textContent.toLowerCase() || '';
        
        const matchesHostname = !hostnameFilter || hostname.includes(hostnameFilter);
        const matchesIp = !ipFilter || ip.includes(ipFilter);
        const matchesDeviceType = !deviceTypeFilter || deviceTypeFilter === 'all' || 
                                deviceType === deviceTypeFilter;
        
        if (matchesHostname && matchesIp && matchesDeviceType) {
            row.style.display = '';
            visibleCount++;
        } else {
            row.style.display = 'none';
        }
    });
    
    updateFilterResultsCount();
    return visibleCount;
}

// Reset filters function
function resetFilters() {
    const filterForm = document.getElementById('inventory-filter-form');
    if (filterForm) filterForm.reset();
    
    const tbody = document.getElementById('csvTableBody');
    if (!tbody) return;
    
    const rows = tbody.querySelectorAll('tr');
    rows.forEach(row => {
        row.style.display = '';
    });
    
    updateFilterResultsCount();
}

// Add a new row to the table
function addNewRow() {
    const tableBody = document.getElementById('csvTableBody');
    const tableHeader = document.getElementById('csvTableHeader');
    
    if (!tableBody || !tableHeader) return;
    
    const headerRow = tableHeader.querySelector('tr');
    if (!headerRow) return;
    
    const columnCount = headerRow.querySelectorAll('th').length;
    
    const newRow = document.createElement('tr');
    
    for (let i = 0; i < columnCount; i++) {
        const td = document.createElement('td');
        td.contentEditable = "true";
        
        // Add special classes based on header
        const headerCell = headerRow.querySelectorAll('th')[i];
        if (headerCell) {
            const headerText = headerCell.textContent.toLowerCase();
            if (headerText === 'hostname') {
                td.classList.add('hostname-cell');
            } else if (headerText === 'ip') {
                td.classList.add('ip-cell');
            } else if (headerText === 'device_type') {
                td.classList.add('device-type-cell');
            }
        }
        
        newRow.appendChild(td);
    }
    
    // Add delete button
    const deleteCell = document.createElement('td');
    const deleteBtn = document.createElement('button');
    deleteBtn.className = "btn btn-sm btn-danger";
    deleteBtn.innerHTML = '<i class="fas fa-trash"></i>';
    deleteBtn.addEventListener('click', function() {
        newRow.remove();
        updateFilterResultsCount();
    });
    deleteCell.appendChild(deleteBtn);
    newRow.appendChild(deleteCell);
    
    tableBody.appendChild(newRow);
    updateFilterResultsCount();
}

// Initialize inventory filtering
function initializeInventoryFiltering() {
    // Set up filter form events
    const filterForm = document.getElementById('inventory-filter-form');
    if (filterForm) {
        filterForm.addEventListener('submit', function(e) {
            e.preventDefault();
            applyFilters();
        });
        
        const resetButton = document.getElementById('reset-filters');
        if (resetButton) {
            resetButton.addEventListener('click', function(e) {
                e.preventDefault();
                resetFilters();
            });
        }
    }
}

document.addEventListener('DOMContentLoaded', async function() {
    // Initialize tab switching
    const csvTab = document.getElementById('csv-tab');
    const rawTab = document.getElementById('raw-tab');
    
    if (csvTab && rawTab) {
        console.log('Setting up tab event listeners');
        csvTab.addEventListener('click', function() {
            document.getElementById('csv-editor').classList.add('show', 'active');
            document.getElementById('raw-editor').classList.remove('show', 'active');
            csvTab.classList.add('active');
            rawTab.classList.remove('active');
            
            // Refresh the table when switching to CSV tab
            loadFromRawContent();
        });
        
        rawTab.addEventListener('click', function() {
            document.getElementById('raw-editor').classList.add('show', 'active');
            document.getElementById('csv-editor').classList.remove('show', 'active');
            rawTab.classList.add('active');
            csvTab.classList.remove('active');
        });
    }
    
    // Try loading the table using different methods
    // First try API, then raw content, then template data
    const apiSuccess = await loadInventoryDataFromAPI();
    
    if (!apiSuccess) {
        console.log('API loading failed, trying raw content');
        if (!loadFromRawContent()) {
            console.log('Raw content loading failed, trying template data');
            const templateData = '{{ current_inventory_content_raw|replace("\n", "\\n")|replace("'", "\'") | safe }}';
            if (templateData) {
                const csvData = parseCSV(templateData);
                populateCSVTable(csvData);
            } else {
                console.error('No data sources available to load table');
            }
        }
    }
    
    // Set up add row button
    const addRowBtn = document.getElementById('addRowBtn');
    if (addRowBtn) {
        addRowBtn.addEventListener('click', addNewRow);
    }
    
    // Set up form submission for CSV data
    const csvTableForm = document.getElementById('csvTableForm');
    if (csvTableForm) {
        csvTableForm.addEventListener('submit', function(e) {
            // Prepare the CSV data before submission
            const tableHeader = document.getElementById('csvTableHeader');
            const tableBody = document.getElementById('csvTableBody');
            const hiddenInput = document.getElementById('csv-data-hidden');
            
            if (!tableHeader || !tableBody || !hiddenInput) return;
            
            // Get all headers
            const headers = [];
            const headerCells = tableHeader.querySelectorAll('th');
            headerCells.forEach(cell => headers.push(cell.textContent));
            
            // Build CSV string starting with headers
            let csvString = headers.join(',') + '\n';
            
            // Add each row
            const rows = tableBody.querySelectorAll('tr');
            rows.forEach(row => {
                const cells = row.querySelectorAll('td:not(:last-child)'); // Skip delete button cell
                const rowValues = [];
                cells.forEach(cell => rowValues.push(cell.textContent));
                csvString += rowValues.join(',') + '\n';
            });
            
            // Set the hidden input value
            hiddenInput.value = csvString;
        });
    }
    
    // Initialize filtering functionality
    initializeInventoryFiltering();
});



// Add a new row to the table
function addRow() {
    const tableBody = document.getElementById('csvTableBody');
    const columnCount = document.getElementById('csvTableHeader').firstChild.childElementCount;
    
    let newRow = document.createElement('tr');
    
    // Create empty cells for each column
    for (let i = 0; i < columnCount; i++) {
        let td = document.createElement('td');
        td.contentEditable = "true";
        td.textContent = "";
        newRow.appendChild(td);
    }
    
    // Add delete row button
    let deleteCell = document.createElement('td');
    let deleteBtn = document.createElement('button');
    deleteBtn.className = "btn btn-sm btn-danger";
    deleteBtn.innerHTML = '<i class="fas fa-trash"></i>';
    deleteBtn.addEventListener('click', function() {
        newRow.remove();
    });
    deleteCell.appendChild(deleteBtn);
    newRow.appendChild(deleteCell);
    
    tableBody.appendChild(newRow);
}

// Add a new column to the table
function addColumn() {
    const headerRow = document.getElementById('csvTableHeader').firstChild;
    const tableRows = document.getElementById('csvTableBody').children;
    
    // Add new header cell
    let newHeaderCell = document.createElement('th');
    newHeaderCell.contentEditable = "true";
    newHeaderCell.textContent = "New Column";
    headerRow.insertBefore(newHeaderCell, headerRow.lastChild);
    
    // Add new cell to each data row
    for (let i = 0; i < tableRows.length; i++) {
        let newCell = document.createElement('td');
        newCell.contentEditable = "true";
        newCell.textContent = "";
        tableRows[i].insertBefore(newCell, tableRows[i].lastChild);
    }
}

// Prepare the CSV data from the table for submission
function prepareCSVData() {
    const headerRow = document.getElementById('csvTableHeader').firstChild;
    const tableRows = document.getElementById('csvTableBody').children;
    let csvData = [];
    
    // Get header row values
    let headers = [];
    for (let i = 0; i < headerRow.children.length; i++) {
        headers.push(headerRow.children[i].textContent);
    }
    csvData.push(headers);
    
    // Get data row values
    for (let i = 0; i < tableRows.length; i++) {
        let rowData = [];
        for (let j = 0; j < headers.length; j++) {
            rowData.push(tableRows[i].children[j].textContent);
        }
        csvData.push(rowData);
    }
    
    // Convert to CSV string
    let csvString = '';
    for (let i = 0; i < csvData.length; i++) {
        for (let j = 0; j < csvData[i].length; j++) {
            // Handle values with commas, quotes, or newlines
            let value = csvData[i][j];
            if (value.includes(',') || value.includes('"') || value.includes('\n')) {
                value = '"' + value.replace(/"/g, '""') + '"';
            }
            csvString += value;
            if (j < csvData[i].length - 1) {
                csvString += ',';
            }
        }
        if (i < csvData.length - 1) {
            csvString += '\n';
        }
    }
    
    // Set the hidden input value
    document.getElementById('csv-data-hidden').value = csvString;
}

// Initialize inventory filtering functionality
function initializeInventoryFiltering() {
    const filterForm = document.getElementById('inventory-filter-form');
    if (!filterForm) return;

    // Apply filters when form is submitted
    filterForm.addEventListener('submit', function(e) {
        e.preventDefault();
        applyFilters();
    });

    // Reset filters when reset button is clicked
    const resetButton = document.getElementById('reset-filters');
    if (resetButton) {
        resetButton.addEventListener('click', function(e) {
            e.preventDefault();
            resetFilters();
        });
    }

    // Initialize device type dropdown with unique values
    populateDeviceTypeDropdown();
    
    // Initialize results count
    updateFilterResultsCount();
}

// Populate device type dropdown with unique values from the table
function populateDeviceTypeDropdown() {
    console.log('Populating device type dropdown');
    const dropdown = document.getElementById('device-type-filter');
    if (!dropdown) {
        console.error('Device type dropdown not found');
        return;
    }
    
    const table = document.getElementById('csvTable');
    if (!table) {
        console.error('CSV table not found');
        return;
    }

    const tbody = document.getElementById('csvTableBody');
    if (!tbody) {
        console.error('CSV table body not found');
        return;
    }
    
    // Get all device type cells
    const deviceTypeCells = tbody.querySelectorAll('.device-type-cell');
    console.log(`Found ${deviceTypeCells.length} device type cells`);
    
    // Extract unique device types
    const deviceTypes = new Set();
    deviceTypeCells.forEach(cell => {
        const deviceType = cell.textContent.trim().toLowerCase();
        if (deviceType) {
            deviceTypes.add(deviceType);
            console.log(`Added device type: ${deviceType}`);
        }
    });
    
    // Add an "All" option
    let options = '<option value="all">All Device Types</option>';
    
    // Add options for each unique device type
    deviceTypes.forEach(deviceType => {
        options += `<option value="${deviceType}">${deviceType.charAt(0).toUpperCase() + deviceType.slice(1)}</option>`;
    });
    
    // Set the dropdown options
    dropdown.innerHTML = options;
    console.log(`Device type dropdown populated with ${deviceTypes.size} options plus 'All'`);
}

// Apply filters to the inventory table
function applyFilters() {
    const hostnameFilter = document.getElementById('hostname-filter').value.toLowerCase();
    const ipFilter = document.getElementById('ip-filter').value.toLowerCase();
    const deviceTypeFilter = document.getElementById('device-type-filter').value.toLowerCase();
    
    // Get all rows in the CSV table
    const table = document.getElementById('csvTable');
    if (!table) return;

    const tbody = table.querySelector('tbody');
    if (!tbody) return;

    const rows = tbody.querySelectorAll('tr');
    
    // Filter the rows
    let visibleCount = 0;
    rows.forEach(row => {
        const hostname = row.querySelector('.hostname-cell')?.textContent.toLowerCase() || '';
        const ip = row.querySelector('.ip-cell')?.textContent.toLowerCase() || '';
        const deviceType = row.querySelector('.device-type-cell')?.textContent.toLowerCase() || '';
        
        // Check if row matches all active filters
        const matchesHostname = !hostnameFilter || hostname.includes(hostnameFilter);
        const matchesIp = !ipFilter || ip.includes(ipFilter);
        const matchesDeviceType = !deviceTypeFilter || deviceTypeFilter === 'all' || deviceType === deviceTypeFilter;
        
        // Show or hide the row based on filter matches
        if (matchesHostname && matchesIp && matchesDeviceType) {
            row.style.display = '';
            visibleCount++;
        } else {
            row.style.display = 'none';
        }
    });
    
    // Update filter results count
    updateFilterResultsCount(visibleCount, rows.length);
    
    // Show/hide the "no results" message
    const noResultsMsg = document.getElementById('no-filter-results');
    if (noResultsMsg) {
        noResultsMsg.style.display = visibleCount === 0 ? 'block' : 'none';
    }
}

// Reset all filters to their default state
function resetFilters() {
    const filterForm = document.getElementById('inventory-filter-form');
    if (!filterForm) return;
    
    // Reset all input fields
    filterForm.reset();
    
    // Show all rows
    const table = document.getElementById('csvTable');
    if (!table) return;

    const tbody = table.querySelector('tbody');
    if (!tbody) return;

    const rows = tbody.querySelectorAll('tr');
    rows.forEach(row => {
        row.style.display = '';
    });
    
    // Reset filter results count
    updateFilterResultsCount(rows.length, rows.length);
    
    // Hide the "no results" message
    const noResultsMsg = document.getElementById('no-filter-results');
    if (noResultsMsg) {
        noResultsMsg.style.display = 'none';
    }
}

// Update the filter results count display
function updateFilterResultsCount(visibleCount, totalCount) {
    const countElement = document.getElementById('filter-results-count');
    if (countElement) {
        if (!visibleCount && !totalCount) {
            // Initialize with current table state
            const table = document.getElementById('csvTable');
            if (table) {
                const tbody = table.querySelector('tbody');
                if (tbody) {
                    const allRows = tbody.querySelectorAll('tr');
                    const visibleRows = Array.from(allRows).filter(row => row.style.display !== 'none');
                    countElement.textContent = `Showing ${visibleRows.length} of ${allRows.length} devices`;
                    return;
                }
            }
        }
        countElement.textContent = `Showing ${visibleCount} of ${totalCount} devices`;
    }
}
</script>
{% endblock %}
"""

app.jinja_loader = DictLoader({
    TEMPLATE_BASE_NAME: HTML_BASE_LAYOUT, TEMPLATE_INDEX_NAME: HTML_INDEX_PAGE,
    TEMPLATE_SETTINGS_NAME: HTML_SETTINGS_TEMPLATE_CONTENT, TEMPLATE_VIEW_JSON_NAME: HTML_VIEW_JSON_TEMPLATE_CONTENT,
    TEMPLATE_EDIT_INVENTORY_NAME: HTML_EDIT_INVENTORY_TEMPLATE_CONTENT
})

def allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def read_csv_data_from_str(csv_string):
    """Parses a CSV string into a list of dictionaries (data) and a list of headers.
       Returns (None, None) on critical error or if CSV is fundamentally unparsable.
       Returns ([], headers) for empty CSV with only headers.
       Returns ([], []) for completely empty CSV string.
    """
    if not csv_string.strip():
        return [], [] # Empty string, empty data, empty headers
    try:
        # Use io.StringIO to treat the string as a file
        f = io.StringIO(csv_string)
        # Sniff to find the dialect (handles various CSV styles)
        try:
            dialect = csv.Sniffer().sniff(f.read(1024)) # Read a sample
            f.seek(0) # Rewind to the beginning of the string buffer
            reader = csv.reader(f, dialect)
        except csv.Error: # If sniffer fails (e.g. too simple, or not CSV)
            f.seek(0)
            # Basic check: if it has commas, assume standard CSV, else treat as single column
            if ',' in f.readline():
                f.seek(0)
                reader = csv.reader(f) # Default dialect
            else: # No commas, might be single column or not CSV at all
                f.seek(0)
                # If it's just one line without commas, it could be a header or a single data point.
                # Or if multiple lines without commas, treat each as a single-column row.
                # This basic reader will try its best.
                reader = csv.reader(f)

        headers = next(reader, None) # Get the header row
        if headers is None:
             # This case means the CSV string was empty or had some issue where headers couldn't be read
            app.logger.warning("CSV string seems to be empty or headers could not be read.")
            return [], [] # No headers, no data

        # Clean headers (strip whitespace)
        cleaned_headers = [header.strip() for header in headers]
        
        # Read data rows into a list of dictionaries
        data = []
        for row_values in reader:
            if not any(field.strip() for field in row_values): # Skip completely blank lines
                continue
            # Create a dict, ensuring row has a value for each header (even if empty)
            # If row is shorter than headers, missing values are empty strings implicitly via get
            # If row is longer, extra values are ignored by zip if headers is shorter.
            # For robustness, we ensure all headers are present in each dict.
            row_dict = {}
            for i, header in enumerate(cleaned_headers):
                row_dict[header] = row_values[i].strip() if i < len(row_values) else ''
            data.append(row_dict)
        
        app.logger.debug(f"Parsed CSV. Headers: {cleaned_headers}, Data rows: {len(data)}")
        return data, cleaned_headers
    except Exception as e:
        app.logger.error(f"Error parsing CSV string: {e}\nString was: '{csv_string[:200]}...'", exc_info=True)
        return None, None # Indicate critical parsing failure

# CSV-related functions have been moved to before the load_active_inventory function

def get_parsed_inventory_with_details(filepath):
    """Reads a CSV inventory file and returns its parsed content, format, and headers.
    Returns: (data_structure, format_string, headers)
             e.g., (list_of_dicts, 'csv', ['h1', 'h2'])
             Returns (None, 'csv', None) on failure to parse or file not found.
    """
    if not os.path.exists(filepath):
        app.logger.warning(f"File not found: {filepath}")
        return None, 'csv', None

    _, ext = os.path.splitext(filepath)
    if ext.lower() != '.csv':
        app.logger.error(f"Unsupported file extension '{ext}' for parsing from {filepath}. Only .csv is supported.")
        return None, 'csv', None

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content_str = f.read()
            if not content_str.strip():
                app.logger.debug(f"CSV file {filepath} is empty. Returning empty list and empty headers.")
                return [], 'csv', [] # Empty CSV file, return empty list and headers
                
            data, headers = read_csv_data_from_str(content_str)
            if data is None: # Indicates parsing error in read_csv_data_from_str
                raise ValueError("read_csv_data_from_str failed to parse CSV content.")
            return data, 'csv', headers
    except Exception as e:
        app.logger.error(f"Error parsing CSV inventory file {filepath}: {e}")
        return None, 'csv', None 

# Functions to read current inventory as specific string formats for UI
def read_current_inventory_as_csv_str() -> str:
    """Reads the active inventory and returns its content as a CSV formatted string.
       Since we only support CSV inventory, this simply reads the file and returns its content.
    """
    active_inventory_filepath = get_inventory_path()
    if not active_inventory_filepath or not os.path.exists(active_inventory_filepath):
        app.logger.warning("read_current_inventory_as_csv_str: Active inventory file not found or path is invalid.")
        return "# Active inventory file not found or path is invalid."

    try:
        # Direct file read approach for CSV
        with open(active_inventory_filepath, 'r', newline='', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        app.logger.error(f"Error reading CSV inventory file {active_inventory_filepath}: {e}")
        return f"# Error: Could not read CSV inventory file: {e}"


@app.route('/settings', methods=['GET', 'POST'])
def settings_route():
    global APP_CONFIG, ACTIVE_INVENTORY_DATA
    if request.method == 'POST':
        set_key(DOTENV_PATH, "JUMP_HOST", request.form.get("jump_host", APP_CONFIG.get("JUMP_HOST","")))
        set_key(DOTENV_PATH, "JUMP_USERNAME", request.form.get("jump_username", APP_CONFIG.get("JUMP_USERNAME","")))
        if request.form.get("jump_password"): set_key(DOTENV_PATH, "JUMP_PASSWORD", request.form.get("jump_password"))
        jump_ping_path = request.form.get("jump_ping_path", APP_CONFIG.get("JUMP_PING_PATH", "/bin/ping"))
        set_key(DOTENV_PATH, "JUMP_PING_PATH", jump_ping_path if jump_ping_path else "/bin/ping")
        set_key(DOTENV_PATH, "DEVICE_USERNAME", request.form.get("device_username", APP_CONFIG.get("DEVICE_USERNAME","")))
        if request.form.get("device_password"): set_key(DOTENV_PATH, "DEVICE_PASSWORD", request.form.get("device_password"))
        if "device_enable_password" in request.form: set_key(DOTENV_PATH, "DEVICE_ENABLE", request.form.get("device_enable_password",""))
        
        # Always set inventory format to CSV
        set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FORMAT", "csv")

        load_app_config(); load_active_inventory()
        flash("Settings updated successfully.", "success")
        return redirect(url_for('settings_route'))

    current_config_display = APP_CONFIG.copy()
    for pwd_field in ["JUMP_PASSWORD", "DEVICE_PASSWORD", "DEVICE_ENABLE"]:
        if current_config_display.get(pwd_field): current_config_display[pwd_field] = "********" 
        else: current_config_display[pwd_field] = ""
    return render_template(TEMPLATE_SETTINGS_NAME, config=current_config_display)

@app.route('/manage_inventories', methods=['GET', 'POST'])
def manage_inventories_route():
    global APP_CONFIG, ACTIVE_INVENTORY_DATA
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'upload':
            if 'inventory_file_upload_manage' not in request.files:
                flash("No file part in the request.", "warning"); return redirect(url_for('manage_inventories_route'))
            file_upload = request.files['inventory_file_upload_manage']
            if file_upload.filename == '':
                flash("No file selected for uploading.", "warning"); return redirect(url_for('manage_inventories_route'))

            if file_upload and allowed_file(file_upload.filename):
                filename_secure = secure_filename(file_upload.filename)
                file_ext = filename_secure.rsplit('.', 1)[1].lower()
                content_str = file_upload.read().decode('utf-8')
                
                yaml_equivalent_data = None
                is_valid = False
                msg = "Unknown error during validation."

                if file_ext in ['yml', 'yaml']:
                    try:
                        yaml_equivalent_data = yaml.safe_load(content_str)
                        is_valid, msg = validate_inventory_data_dict(yaml_equivalent_data)
                    except yaml.YAMLError as e:
                        is_valid, msg = False, f"Invalid YAML format: {e}"
                elif file_ext == 'csv':
                    # Parse and validate CSV content directly
                    parsed_csv_data, parsed_csv_headers = read_csv_data_from_str(content_str)
                    if parsed_csv_data is None:
                        is_valid, msg = False, "CSV content is malformed or unparsable"
                    else:
                        # Validate CSV data
                        is_valid, msg = validate_csv_data_list(parsed_csv_data, parsed_csv_headers)
                        
                        if is_valid:
                            # For backward compatibility, also create a YAML equivalent
                            yaml_equivalent_data = {"routers": {}}
                            
                            # Process each row in the CSV data
                            for row in parsed_csv_data:
                                hostname = row.get('hostname')
                                if not hostname:
                                    continue
                                    
                                # Create router entry
                                router_data = {}
                                for key, value in row.items():
                                    if key != 'hostname' and value:  # Skip empty values
                                        router_data[key] = value
                                        
                                # Add to routers dictionary
                                yaml_equivalent_data["routers"][hostname] = router_data
                
                if not is_valid:
                    flash(f"Uploaded inventory '{filename_secure}' is invalid: {msg}", "danger")
                else:
                    version = get_next_inventory_version()
                    
                    # Save the file in its original format
                    if file_ext == 'csv':
                        # Save as CSV
                        versioned_filename = f"inventory-list-v{version:02d}.csv"
                        save_path = get_inventory_path(versioned_filename, "csv")
                        try:
                            # For CSV, save the original content
                            with open(save_path, 'w', newline='') as f_csv_out:
                                f_csv_out.write(content_str)
                                
                            # Set as active inventory
                            set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FILE", versioned_filename)
                            set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FORMAT", "csv")
                            load_app_config()
                            load_active_inventory()
                            flash(f"CSV Inventory '{filename_secure}' uploaded, validated, saved as '{versioned_filename}', and set active.", "success")
                        except Exception as e_save:
                            flash(f"Error saving CSV inventory '{versioned_filename}': {e_save}", "danger")
                    else:
                        # Save as YAML
                        versioned_filename = f"inventory-list-v{version:02d}.yml"
                        save_path = get_inventory_path(versioned_filename, "yaml")
                        try:
                            with open(save_path, 'w') as f_save:
                                yaml.dump(yaml_equivalent_data, f_save, sort_keys=False, indent=2)
                            
                            set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FILE", versioned_filename)
                            set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FORMAT", "yaml")
                            load_app_config()
                            load_active_inventory()
                            flash(f"YAML Inventory '{filename_secure}' uploaded, validated, saved as '{versioned_filename}', and set active.", "success")
                        except Exception as e_save:
                            flash(f"Error saving YAML inventory '{versioned_filename}': {e_save}", "danger")
            elif file_upload.filename:
                flash(f"Invalid file type for '{file_upload.filename}'. Allowed: {', '.join(app.config['ALLOWED_EXTENSIONS'])}.", "warning")
        
        elif action == 'set_active':
            selected_inv_filename = request.form.get("active_inventory_file_manage")
            available_inventories = list_inventory_files()
            if selected_inv_filename and selected_inv_filename in available_inventories:
                selected_ext = selected_inv_filename.rsplit('.',1)[1].lower()
                if selected_ext not in ['yaml', 'yml', 'csv']:
                    flash(f"Invalid file type for selected inventory: {selected_inv_filename}", "danger")
                else:
                    if selected_inv_filename != APP_CONFIG.get("ACTIVE_INVENTORY_FILE") or \
                       selected_ext != APP_CONFIG.get("ACTIVE_INVENTORY_FORMAT"):
                        set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FILE", selected_inv_filename)
                        set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FORMAT", "csv" if selected_ext == "csv" else "yaml")
                        load_app_config(); load_active_inventory()
                        flash(f"Active inventory changed to {selected_inv_filename} (Format: {selected_ext.upper()}).", "success")
                    else:
                        flash(f"{selected_inv_filename} is already the active inventory.", "info")
            else:
                flash("Invalid inventory file selected or file does not exist.", "danger")
        return redirect(url_for('manage_inventories_route'))

    available_inventories = list_inventory_files()
    if not available_inventories: create_default_inventory_if_missing(); available_inventories = list_inventory_files()

    # Get CSV content for editing
    csv_content = read_current_inventory_as_csv_str()
    
    # Create a default chart_data to avoid JSON serialization errors
    default_chart_data = {"labels": ["No Data"], "values": [1], "colors": ["#D3D3D3"]}
    
    return render_template(TEMPLATE_EDIT_INVENTORY_NAME, 
                           inventories=available_inventories, 
                           active_inventory=APP_CONFIG.get("ACTIVE_INVENTORY_FILE"), 
                           active_inventory_format="csv",
                           current_inventory_content_raw=csv_content,
                           chart_data=default_chart_data)

@app.route('/export_inventory_csv', methods=['GET'])
def export_inventory_csv_route():
    # Get the active inventory filepath
    active_inventory_filepath = get_inventory_path()
    if not active_inventory_filepath:
        flash("No active inventory to export.", "warning")
        return redirect(url_for('manage_inventories_route'))
    
    # Check if the file exists
    if not os.path.exists(active_inventory_filepath):
        flash(f"Active inventory file '{active_inventory_filepath}' not found.", "danger")
        return redirect(url_for('manage_inventories_route'))
    
    # Read the file content
    try:
        with open(active_inventory_filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            
        # Send as a downloadable file
        return Response(
            content,
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment;filename={os.path.basename(active_inventory_filepath)}"}
        )
    except Exception as e:
        flash(f"Error exporting inventory: {e}", "danger")
        return redirect(url_for('manage_inventories_route'))

@app.route('/get_active_inventory_info', methods=['GET'])
def get_active_inventory_info():
    """API endpoint to get information about the active inventory for the frontend"""
    try:
        # Get active inventory info
        active_inventory_filepath = get_inventory_path()
        
        if not active_inventory_filepath or not os.path.exists(active_inventory_filepath):
            return jsonify({
                'status': 'error',
                'message': 'No active inventory or file not found',
                'data': None,
                'headers': None
            }), 404
        
        # Read and parse the active inventory
        if ACTIVE_INVENTORY_DATA and 'data' in ACTIVE_INVENTORY_DATA and 'headers' in ACTIVE_INVENTORY_DATA:
            # Use the cached data
            inventory_data = ACTIVE_INVENTORY_DATA['data']
            headers = ACTIVE_INVENTORY_DATA['headers']
        else:
            # Parse the file directly
            with open(active_inventory_filepath, 'r', encoding='utf-8') as f:
                content = f.read()
                inventory_data, headers = read_csv_data_from_str(content)
        
        # Check if data was parsed successfully
        if inventory_data is None:
            return jsonify({
                'status': 'error',
                'message': 'Failed to parse inventory data',
                'data': None,
                'headers': None
            }), 500
        
        # Return the data as JSON
        return jsonify({
            'status': 'success',
            'message': 'Active inventory retrieved successfully',
            'data': inventory_data,
            'headers': headers,
            'filename': os.path.basename(active_inventory_filepath)
        })
    except Exception as e:
        app.logger.error(f"Error getting active inventory info: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': f'Error: {str(e)}',
            'data': None,
            'headers': None
        }), 500

@app.route('/edit_active_inventory_content', methods=['POST'])
def edit_active_inventory_content_route():
    content_str = request.form.get('inventory_content_edit')
    
    if not content_str:
        flash("Inventory content (CSV) cannot be empty.", "danger")
        return redirect(url_for('manage_inventories_route'))
    
    version = get_next_inventory_version()
    
    # Parse and validate CSV content
    parsed_csv_data, parsed_csv_headers = read_csv_data_from_str(content_str)

    if parsed_csv_data is None:
        flash("Edited CSV content is malformed or unparsable. Fix and retry.", "danger")
        return render_template(TEMPLATE_EDIT_INVENTORY_NAME, 
                              inventories=list_inventory_files(), 
                              active_inventory=APP_CONFIG.get("ACTIVE_INVENTORY_FILE"),
                              active_inventory_format="csv",
                              current_inventory_content_raw=content_str)
    
    # Validate CSV data directly
    is_valid, msg = validate_csv_data_list(parsed_csv_data, parsed_csv_headers)
    if not is_valid and (parsed_csv_data or parsed_csv_headers): # Only fail validation if there was actual content to validate
        flash(f"Edited CSV inventory content is invalid: {msg}. Fix and retry.", "danger")
        return render_template(TEMPLATE_EDIT_INVENTORY_NAME, 
                              inventories=list_inventory_files(), 
                              active_inventory=APP_CONFIG.get("ACTIVE_INVENTORY_FILE"),
                              active_inventory_format="csv",
                              current_inventory_content_raw=content_str)
    
    # Save CSV content
    new_versioned_csv_filename = f"inventory-list-v{version:02d}.csv"
    csv_save_path = get_inventory_path(new_versioned_csv_filename, "csv")
    try:
        # Use write_csv_data_to_str for consistent saving from parsed data
        csv_string_to_save = write_csv_data_to_str(parsed_csv_data, parsed_csv_headers)
        with open(csv_save_path, 'w', newline='') as f_csv_out:
            f_csv_out.write(csv_string_to_save)
        app.logger.info(f"Saved edited content to CSV: {csv_save_path}")

        set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FILE", new_versioned_csv_filename)
        set_key(DOTENV_PATH, "ACTIVE_INVENTORY_FORMAT", "csv")
        load_app_config(); load_active_inventory()
        flash(f"Inventory changes saved as '{new_versioned_csv_filename}' (CSV) and set active.", "success")
    except Exception as e_save_csv:
        flash(f"Error saving edited CSV inventory '{new_versioned_csv_filename}': {e_save_csv}", "danger")
        app.logger.error(f"Error saving CSV: {e_save_csv}", exc_info=True)
    
    return redirect(url_for('manage_inventories_route'))


@app.route('/')
def index(): # (No changes needed in this route itself for inventory format, it uses ACTIVE_INVENTORY_DATA)
    chart_data = {}
    if last_run_summary_data and (last_run_summary_data.get('total_routers', 0) > 0 or audit_status != "Not Run"):
        succeeded_cleanly = last_run_summary_data.get('collected', 0) - last_run_summary_data.get('with_violations', 0)
        chart_data_values = [max(0, succeeded_cleanly), max(0, last_run_summary_data.get('with_violations', 0)), max(0, last_run_summary_data.get('failed_collection', 0)), max(0, last_run_summary_data.get('failed_ssh_auth', 0)), max(0, last_run_summary_data.get('failed_icmp', 0))]
        chart_data_labels = ['Collected (No Violations)', 'Collected (Violations)', 'Failed Collection', 'Failed SSH Auth', 'Failed ICMP']; chart_data_colors = ['#4CAF50', '#FFC107', '#FF9800', '#F44336', '#9E9E9E'] 
        filtered_labels = [label_item for i, label_item in enumerate(chart_data_labels) if chart_data_values[i] > 0]; filtered_values = [value_item for value_item in chart_data_values if value_item > 0]; filtered_colors = [color_item for i, color_item in enumerate(chart_data_colors) if chart_data_values[i] > 0]
        if filtered_values: chart_data = {"labels": filtered_labels, "values": filtered_values, "colors": filtered_colors}
        else: chart_data = {"labels": ["No Data"], "values": [1], "colors": ["#D3D3D3"]}
    else: chart_data = {"labels": ["Not Run Yet"], "values": [1], "colors": ["#D3D3D3"]}
    return render_template(TEMPLATE_INDEX_NAME, audit_status=audit_status, logs=ui_logs[-100:], reports_manifest=detailed_reports_manifest, last_run_summary=last_run_summary_data, current_run_failures=current_run_failures, active_inventory_file=APP_CONFIG.get('ACTIVE_INVENTORY_FILE', 'N/A'), chart_data=chart_data, current_audit_progress=current_audit_progress, audit_paused=audit_paused)

@app.route('/stop_reset_audit', methods=['POST'])
def stop_reset_audit():
    """Stop and reset the current audit process, closing all connections and resetting all state.
    
    This function now uses the centralized state management system for more robust operation.
    """
    global audit_status, audit_thread, audit_pause_event, AUDIT_SHOULD_PAUSE, AUDIT_SHOULD_STOP
    
    print("[DEBUG] /stop_reset_audit endpoint called")
    print(f"[DEBUG] Request method: {request.method}, headers: {dict(request.headers)}")
    try:
        print(f"[DEBUG] Current audit_status: {audit_status}")
        print(f"[DEBUG] audit_thread is {'alive' if audit_thread and audit_thread.is_alive() else 'not running'}")
        
        # Always log the stop/reset request
        log_to_ui_and_console("Stop/Reset Audit requested by user", log_level="INFO")
        
        # Determine message based on current state
        message = "Audit reset successfully"
        
        # Step 1: Signal the audit thread to stop if it exists
        if audit_status in ['running', 'paused']:
            # Force immediate stop of the audit
            with state_lock:
                AUDIT_SHOULD_PAUSE = False
                AUDIT_SHOULD_STOP = True  # Force the audit to stop completely
            
            if audit_thread and audit_thread.is_alive():
                print("[DEBUG] Stopping audit thread...")
                audit_pause_event.set()  # Signal to pause/stop
                audit_thread.join(timeout=5.0)  # Wait up to 5 seconds for the thread to complete
                
                # If thread is still alive after timeout, try to terminate it more forcefully
                if audit_thread.is_alive():
                    print("[WARNING] Audit thread did not stop gracefully, forcing termination...")
                    # We can't actually terminate threads in Python, but we can make sure
                    # the main code knows to exit ASAP
                    audit_thread = None  # Discard the reference to let it be garbage collected
                
                message = "Audit stopped and reset successfully"
                print("[DEBUG] Audit thread stopped.")
            else:
                print("[DEBUG] No running audit thread to stop.")
        else:
            print("[DEBUG] No audit was running, but reset requested.")
            if audit_status == "completed":
                message = "Completed audit has been reset"
            else:
                message = "No audit was running, dashboard reset"
        
        # Step 2: Close any open connections
        try:
            # Force close any connection that might be open using a more thorough approach
            # This includes searching all module-level variables and object attributes
            connections_closed = close_all_ssh_connections()
            
            # Log connection closure information
            if connections_closed > 0:
                log_to_ui_and_console(f"Reset audit: Closed {connections_closed} SSH connection(s)", log_level="INFO")
            else:
                log_to_ui_and_console("Reset audit: No active connections found to close", log_level="INFO")
                
        except Exception as conn_err:
            error_msg = f"Error while closing connections: {conn_err}"
            print(f"[WARNING] {error_msg}")
            log_to_ui_and_console(error_msg, log_level="WARNING")
        
        # Step 3: Use our improved stop_audit_progress function to reset all state
        with state_lock:
            # Reset flags that control audit execution
            AUDIT_SHOULD_PAUSE = False
            AUDIT_SHOULD_STOP = False
            
            # Update audit status before the full reset
            audit_status = 'idle'
        
        # Call our enhanced stop_audit_progress function to properly reset all state
        # This function has been improved to fully reset all global state and close connections
        stop_audit_progress()
        
        # Extra safety: explicitly set audit_status to 'idle' here as well
        with state_lock:
            audit_status = 'idle'
            
        # Explicitly reset AUDIT_PROGRESS global to a fresh state
        global AUDIT_PROGRESS
        AUDIT_PROGRESS = {
            'status': 'idle',
            'total_devices': 0,
            'completed_devices': 0,
            'current_device': None,
            'device_statuses': {},
            'overall_success_count': 0,
            'overall_warning_count': 0,
            'overall_failure_count': 0,
            'start_time': None,
            'end_time': None
        }
        
        # For extra safety, also call any legacy reset function
        if hasattr(app_state, 'reset_state'):
            app_state.reset_state()
        
        # Final log entries
        log_to_ui_and_console("Audit state has been reset completely", log_level="INFO")
        log_to_ui_and_console("System is ready for a new audit", log_level="SUCCESS")
        log_to_ui_and_console(f"Audit dashboard reset: {message}", log_level="SUCCESS")
        
        return jsonify({'status': 'success', 'message': message})
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[ERROR] Exception in /stop_reset_audit: {e}\n{tb}")
        log_to_ui_and_console(f"Error stopping audit: {str(e)}", log_level="ERROR")
        return jsonify({'status': 'error', 'message': f'Error stopping audit: {str(e)}'}), 500

@app.route('/captured_configs')
def captured_configs():
    """Display captured configurations for each router"""
    global APP_CONFIG, current_audit_progress, audit_status, last_run_summary_data
    
    report_base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), BASE_DIR_NAME)
    config_files = []
    
    # Find all configuration files from past audits
    try:
        # Check if the reports directory exists
        if os.path.exists(report_base_dir):
            # Get a list of subdirectories (audit timestamps)
            for audit_dir in sorted(os.listdir(report_base_dir), reverse=True):
                audit_path = os.path.join(report_base_dir, audit_dir)
                if os.path.isdir(audit_path):
                    # Scan for router config files in this audit
                    for router_file in os.listdir(audit_path):
                        if router_file.endswith('.conf') or router_file.endswith('.cfg') or 'running-config' in router_file:
                            router_name = os.path.splitext(router_file)[0]
                            router_name = router_name.replace('-running-config', '')
                            
                            # Get file stats
                            file_path = os.path.join(audit_path, router_file)
                            file_size = os.path.getsize(file_path)
                            last_modified = datetime.fromtimestamp(os.path.getmtime(file_path))
                            
                            # Add to the list of config files
                            config_files.append({
                                'router_name': router_name,
                                'audit_date': audit_dir,
                                'filename': router_file,
                                'filepath': file_path,
                                'filesize': file_size,
                                'last_modified': last_modified,
                                # Convert datetime to string to avoid JSON serialization issues
                                'last_modified_str': last_modified.strftime('%Y-%m-%d %H:%M:%S')
                            })
    except Exception as e:
        print(f"Error scanning for config files: {e}")
        
    # Group config files by router name
    router_configs = {}
    for config in config_files:
        router_name = config['router_name']
        if router_name not in router_configs:
            router_configs[router_name] = []
        router_configs[router_name].append(config)
    
    # Sort each router's configs by date
    for router_name in router_configs:
        router_configs[router_name] = sorted(router_configs[router_name], 
                                           key=lambda x: x['last_modified_str'], 
                                           reverse=True)
    
    # Sort routers alphabetically
    sorted_router_names = sorted(router_configs.keys())
    
    # Initialize chart_data with default values to avoid JSON serialization issues
    chart_data = {"labels": ["No Router Configs"], "values": [1], "colors": ["#D3D3D3"]}
    
    return render_template_string(HTML_CAPTURED_CONFIGS_TEMPLATE, 
                          router_configs=router_configs,
                          sorted_router_names=sorted_router_names,
                          current_audit_progress=current_audit_progress,
                          audit_status=audit_status,
                          last_run_summary=last_run_summary_data,
                          last_run_summarychart_data=chart_data,
                          APP_PORT=APP_CONFIG.get('PORT', 5007))

@app.route('/view_config/<audit_date>/<filename>')
def view_config(audit_date, filename):
    """View a specific router configuration file"""
    try:
        # Security check to ensure the path is within our report directory
        report_base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), BASE_DIR_NAME)
        full_path = os.path.join(report_base_dir, audit_date, filename)
        
        # Prevent path traversal attacks
        if not os.path.abspath(full_path).startswith(os.path.abspath(report_base_dir)):
            return jsonify({'status': 'error', 'message': 'Invalid path'}), 400
        
        if not os.path.exists(full_path):
            return jsonify({'status': 'error', 'message': 'File not found'}), 404
        
        # Read the config file
        with open(full_path, 'r') as f:
            config_content = f.read()
        
        # Get file metadata
        filename = os.path.basename(full_path)
        router_name = os.path.splitext(filename)[0]
        router_name = router_name.replace('-running-config', '')
        last_modified = datetime.fromtimestamp(os.path.getmtime(full_path))
        filesize = os.path.getsize(full_path)
        
        return render_template_string(HTML_VIEW_CONFIG_TEMPLATE,
                              config_content=config_content,
                              filename=filename,
                              router_name=router_name,
                              last_modified=last_modified,
                              filesize=filesize,
                              audit_status=audit_status,
                              current_audit_progress=current_audit_progress,
                              APP_PORT=APP_CONFIG.get('PORT', 5007))
    except Exception as e:
        app.logger.error(f"Error in view_config for {audit_date}/{filename}: {str(e)}")
        return f"Error viewing configuration file: {filename}. Details: {str(e)}", 500
@app.route('/download_config/<audit_date>/<filename>')
def download_config(audit_date, filename):
    """Download a specific router configuration file"""
    try:
        # Security check to ensure the path is within our report directory
        report_base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), BASE_DIR_NAME)
        full_path = os.path.join(report_base_dir, audit_date, filename)
        
        # Prevent path traversal attacks
        if not os.path.normpath(full_path).startswith(os.path.normpath(report_base_dir)):
            return "Invalid file path", 400
        
        if not os.path.exists(full_path):
            return "File not found", 404
        
        # Read the file content
        with open(full_path, 'r') as f:
            content = f.read()
        
        # Create a response with the file content
        response = make_response(content)
        response.headers['Content-Type'] = 'text/plain'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    except Exception as e:
        return f"Error downloading file: {str(e)}", 500

@app.route('/download_all_configs/<router_name>')
def download_all_configs(router_name):
    """Download all configurations for a specific router"""
    try:
        report_base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), BASE_DIR_NAME)
        
        # Find all configuration files for this router
        config_files = []
        for audit_dir in sorted(os.listdir(report_base_dir), reverse=True):
            audit_path = os.path.join(report_base_dir, audit_dir)
            if os.path.isdir(audit_path):
                for router_file in os.listdir(audit_path):
                    if (router_file.endswith('.conf') or router_file.endswith('.cfg') or 'running-config' in router_file) and                        (router_name in router_file or router_name == os.path.splitext(router_file)[0].replace('-running-config', '')):
                        config_files.append({
                            'audit_date': audit_dir,
                            'filename': router_file,
                            'filepath': os.path.join(audit_path, router_file)
                        })
        
        if not config_files:
            return "No configuration files found for this router", 404
        
        # Create a combined text file with all configurations
        combined_content = ""
        for config in sorted(config_files, key=lambda x: x['audit_date'], reverse=True):
            with open(config['filepath'], 'r') as f:
                file_content = f.read()
            
            combined_content += f"===== {config['filename']} ({config['audit_date']}) =====\n\n"
            combined_content += file_content
            combined_content += "\n\n" + "="*80 + "\n\n"
        
        # Create a response with the combined content
        response = make_response(combined_content)
        response.headers['Content-Type'] = 'text/plain'
        response.headers['Content-Disposition'] = f'attachment; filename={router_name}_all_configs.txt'
        
        return response
    except Exception as e:
        return f"Error downloading all configurations: {str(e)}", 500

@app.route('/download_config/<audit_date>/<filename>')
@app.route('/audit_progress_data')
def audit_progress_data():
    global current_audit_progress, audit_status, ui_logs, last_run_summary_data, current_run_failures, audit_paused
    global AUDIT_PROGRESS
    
    # Generate chart data for visualization
    chart_data = {}
    if last_run_summary_data and (last_run_summary_data.get('total_routers', 0) > 0 or audit_status != "Not Run"):
        succeeded_cleanly = last_run_summary_data.get('collected', 0) - last_run_summary_data.get('with_violations', 0)
        chart_data_values = [max(0, succeeded_cleanly), max(0, last_run_summary_data.get('with_violations', 0)), max(0, last_run_summary_data.get('failed_collection', 0)), max(0, last_run_summary_data.get('failed_ssh_auth', 0)), max(0, last_run_summary_data.get('failed_icmp', 0))]
        chart_data_labels = ['Collected (No Violations)', 'Collected (Violations)', 'Failed Collection', 'Failed SSH Auth', 'Failed ICMP']; chart_data_colors = ['#4CAF50', '#FFC107', '#FF9800', '#F44336', '#9E9E9E']
        filtered_labels = [label_item for i, label_item in enumerate(chart_data_labels) if chart_data_values[i] > 0]; filtered_values = [value_item for value_item in chart_data_values if value_item > 0]; filtered_colors = [color_item for i, color_item in enumerate(chart_data_colors) if chart_data_values[i] > 0]
        if filtered_values: chart_data = {"labels": filtered_labels, "values": filtered_values, "colors": filtered_colors}
        else: chart_data = {"labels": ["No Data Available"], "values": [1], "colors": ["#D3D3D3"]}
    else: chart_data = {"labels": ["Not Run Yet"], "values": [1], "colors": ["#D3D3D3"]}
    
    # Prepare enhanced progress data
    enhanced_progress = AUDIT_PROGRESS.copy()
    
    # Format timestamps for JSON serialization - with safe handling of missing keys
    if enhanced_progress.get('start_time'):
        enhanced_progress['start_time'] = enhanced_progress['start_time'].isoformat()
    if enhanced_progress.get('end_time'):
        enhanced_progress['end_time'] = enhanced_progress['end_time'].isoformat()
    if enhanced_progress.get('current_device_start_time'):
        enhanced_progress['current_device_start_time'] = enhanced_progress['current_device_start_time'].isoformat()
    if enhanced_progress.get('estimated_completion_time'):
        enhanced_progress['estimated_completion_time'] = enhanced_progress['estimated_completion_time'].isoformat()
    
    # Calculate additional metrics
    if enhanced_progress['total_devices'] > 0:
        enhanced_progress['percent_complete'] = (enhanced_progress['completed_devices'] / enhanced_progress['total_devices']) * 100
    else:
        enhanced_progress['percent_complete'] = 0
    
    # Calculate elapsed time
    if enhanced_progress['start_time']:
        start_time = datetime.fromisoformat(enhanced_progress['start_time'])
        if enhanced_progress['status'] == 'completed' and enhanced_progress['end_time']:
            end_time = datetime.fromisoformat(enhanced_progress['end_time'])
            elapsed_seconds = (end_time - start_time).total_seconds()
        else:
            elapsed_seconds = (datetime.now() - start_time).total_seconds()
        
        # Format as HH:MM:SS
        hours, remainder = divmod(int(elapsed_seconds), 3600)
        minutes, seconds = divmod(remainder, 60)
        enhanced_progress['elapsed_time'] = f"{hours:02}:{minutes:02}:{seconds:02}"
    else:
        enhanced_progress['elapsed_time'] = "00:00:00"
    
    # Count device statuses
    status_counts = {
        'success': enhanced_progress.get('overall_success_count', 0),
        'warning': enhanced_progress.get('overall_warning_count', 0),
        'failure': enhanced_progress.get('overall_failure_count', 0),
        'in_progress': 0
    }
    
    # Count in-progress devices from device_statuses
    for status in enhanced_progress['device_statuses'].values():
        if status not in ['success', 'warning', 'failure'] and status is not None:
            status_counts['in_progress'] += 1
    
    enhanced_progress['status_counts'] = status_counts
    
    # Prepare the response data
    serializable_failures = {k: str(v) if v is not None else None for k, v in current_run_failures.items()}
    data_to_send = {
        "progress": current_audit_progress,  # Legacy progress format
        "enhanced_progress": enhanced_progress,  # New enhanced progress format
        "overall_audit_status": audit_status, 
        "ui_logs": ui_logs,  # Send all logs instead of just the last 50
        "last_run_summary": last_run_summary_data, 
        "current_run_failures": serializable_failures, 
        "per_router_status": last_run_summary_data.get("per_router_status", {}), 
        "chart_data": chart_data, 
        "audit_paused": audit_paused
    }
    
    return jsonify(data_to_send)

def audit_runner_thread_wrapper(): # (No changes needed in this function itself)
    global audit_status, ui_logs, audit_paused, audit_pause_event, current_audit_progress 
    global last_successful_audit_completion_time, last_audit_start_time_attempt
    if not audit_lock.acquire(blocking=False): log_to_ui_and_console("Audit run attempt failed: Another audit is already in progress or finalizing."); return
    last_audit_start_time_attempt = datetime.now(); ui_logs.clear(); log_to_ui_and_console("Starting new audit run...")
    audit_paused = False; audit_pause_event.set()
    current_audit_progress = {"status_message": "Initializing...", "devices_processed_count": 0, "total_devices_to_process": 0, 
    "percentage_complete": 0, "current_phase": "Initialization", "current_device_hostname": "N/A"}
    try:
        load_app_config(); load_active_inventory()
        if not ACTIVE_INVENTORY_DATA or not ACTIVE_INVENTORY_DATA.get("routers"):
            log_to_ui_and_console(f"{Fore.RED}Cannot start audit: No routers in active inventory '{APP_CONFIG['ACTIVE_INVENTORY_FILE']}'.{Style.RESET_ALL}"); audit_status = "Failed: No routers in inventory"; current_audit_progress["status_message"] = audit_status; return
        current_audit_progress["total_devices_to_process"] = len(ACTIVE_INVENTORY_DATA.get("routers", {}))
        run_the_audit_logic()
        if audit_status == "Completed": last_successful_audit_completion_time = datetime.now()
    except Exception as e_runner:
        log_to_ui_and_console(f"{Fore.RED}Critical error in audit runner: {e_runner}{Style.RESET_ALL}"); audit_status = f"Failed Critically: {e_runner}"; current_audit_progress["status_message"] = audit_status
        import traceback; tb_str = traceback.format_exc(); log_to_ui_and_console(sanitize_log_message(tb_str)); ui_logs.append(strip_ansi(sanitize_log_message(tb_str)))
    finally:
        ui_logs = [log_entry for log_entry in ui_logs if " कार्य प्रगति पर है " not in log_entry]
        if audit_status == "Running": audit_status = "Failed: Interrupted"; current_audit_progress["status_message"] = audit_status
        elif audit_status == "Completed": current_audit_progress["status_message"] = "Audit Completed"; current_audit_progress["percentage_complete"] = 100;
        if not last_successful_audit_completion_time and audit_status == "Completed": last_successful_audit_completion_time = datetime.now()
        audit_paused = False; audit_pause_event.set()
        audit_lock.release(); log_to_ui_and_console(f"Audit ended with status: {audit_status}")

@app.route('/start_audit', methods=['GET', 'POST'])
def start_audit_route():
    global audit_status, audit_paused, audit_pause_event, audit_thread, last_audit_start_time_attempt
    global AUDIT_SHOULD_PAUSE, AUDIT_SHOULD_STOP
    
    if audit_status == "Running":
        flash("Audit already running", "warning")
        return redirect(url_for('index'))
    
    # Reset both pause mechanisms
    audit_paused = False
    audit_pause_event.set() # Ensure not paused
    AUDIT_SHOULD_PAUSE = False
    AUDIT_SHOULD_STOP = False
    
    if audit_thread and audit_thread.is_alive():
        flash("Audit thread already running (abnormal state). Please restart the application.", "danger")
        return redirect(url_for('index'))
    
    audit_thread = threading.Thread(target=audit_runner_thread_wrapper)
    audit_thread.daemon = True
    audit_thread.start()
    last_audit_start_time_attempt = datetime.now()
    
    flash("Audit started", "success")
    return redirect(url_for('index'))

@app.route('/pause_audit', methods=['POST'])
def pause_audit_route():
    global audit_paused, audit_status, audit_pause_event, current_audit_progress
    global AUDIT_SHOULD_PAUSE
    
    if audit_lock.locked() and audit_status == "Running":
        # Set both pause mechanisms
        audit_paused = True
        audit_pause_event.clear()
        AUDIT_SHOULD_PAUSE = True
        
        # Update progress messages
        previous_status_message = current_audit_progress.get("status_message", "Running")
        current_audit_progress["status_message"] = f"Pausing ({previous_status_message})..."
        
        log_to_ui_and_console(f"{Fore.YELLOW}Audit PAUSE signal sent.{Style.RESET_ALL}")
        flash("Audit pause signal sent.", "info")
    else:
        flash("Audit is not currently running or cannot be paused.", "warning")
    
    return redirect(url_for('index'))

@app.route('/resume_audit', methods=['POST'])
def resume_audit_route():
    global audit_paused, audit_status, audit_pause_event, current_audit_progress
    global AUDIT_SHOULD_PAUSE
    
    if audit_lock.locked() and audit_paused:
        # Reset both pause mechanisms
        audit_paused = False
        audit_pause_event.set()
        AUDIT_SHOULD_PAUSE = False
        
        # Update progress messages
        current_audit_progress["status_message"] = "Resuming..."
        
        log_to_ui_and_console(f"{Fore.GREEN}Audit RESUME signal sent.{Style.RESET_ALL}")
        flash("Audit resume signal sent.", "info")
    else:
        flash("Audit is not paused or cannot be resumed.", "warning")
    
    return redirect(url_for('index'))

@app.route('/reports/<path:folder_relative_path>')
def serve_report_file(folder_relative_path): # (No changes needed)
    base_report_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), BASE_DIR_NAME); target_path = os.path.normpath(os.path.join(base_report_dir, folder_relative_path))
    if not target_path.startswith(os.path.abspath(base_report_dir)): flash("Access denied (Path Traversal).", "danger"); return redirect(url_for('index'))
    actual_directory = os.path.dirname(target_path); filename_part = os.path.basename(target_path)
    if not os.path.exists(os.path.join(actual_directory, filename_part)): flash(f"Report file '{filename_part}' not found.", "danger"); return redirect(url_for('index'))
    return send_from_directory(actual_directory, filename_part, as_attachment=True)

@app.route('/view_report/<folder>/<filename>')
def view_json_report(folder, filename): # (No changes needed)
    base_report_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), BASE_DIR_NAME); safe_folder = secure_filename(folder); safe_filename = secure_filename(filename)
    if folder != safe_folder or filename != safe_filename: flash("Invalid characters in report path.", "danger"); return redirect(url_for('index'))
    file_path = os.path.normpath(os.path.join(base_report_dir, safe_folder, safe_filename))
    if not file_path.startswith(os.path.abspath(base_report_dir)): flash("Access denied (Path Traversal).", "danger"); return redirect(url_for('index'))
    try:
        with open(file_path, 'r', encoding='utf-8') as f_report_view: content = f_report_view.read()
        if filename.lower().endswith(".json"):
            try: data_json = json.loads(content); pretty_data = json.dumps(data_json, indent=2)
            except json.JSONDecodeError: pretty_data = f"Error: File '{filename}' is not valid JSON.\n\nRaw content:\n{content}"
        else: pretty_data = content
        return render_template(TEMPLATE_VIEW_JSON_NAME, filename=filename, data_content=pretty_data)
    except FileNotFoundError: flash(f"Report '{filename}' not found in '{folder}'.", "danger"); return redirect(url_for('index'))
    except Exception as e_view: flash(f"Error loading report '{filename}': {e_view}", "danger"); return redirect(url_for('index'))

@app.context_processor
def inject_now(): 
    # Inject both timestamp and port number for all templates
    return {
        'BROWSER_TIMESTAMP': datetime.now(timezone.utc),
        'APP_PORT': app.config.get('PORT', 5007)  # Default to 5007 if not set
    }

def interactive_shell_reader(sid: str, paramiko_channel: paramiko.channel.Channel): # (No changes needed)
    try:
        while paramiko_channel.active and sid in interactive_sessions:
            if paramiko_channel.recv_ready(): data = paramiko_channel.recv(4096).decode('utf-8', errors='replace'); socketio.emit('shell_output', {'output': data}, room=sid, namespace='/')
            elif paramiko_channel.exit_status_ready(): break
            time.sleep(0.05)
    except Exception as e: print(f"Error in shell_reader for {sid}: {e}"); socketio.emit('shell_error', {'error': f"Shell reader error: {str(e)}"}, room=sid, namespace='/')
    finally:
        print(f"Shell reader thread for {sid} ended.")
        if sid in interactive_sessions and interactive_sessions[sid].get('channel') == paramiko_channel:
            socketio.emit('shell_stopped', {'message': 'Shell session ended or error.'}, room=sid, namespace='/'); cleanup_interactive_session(sid)

def cleanup_interactive_session(sid: str): # (No changes needed)
    session_data = interactive_sessions.pop(sid, None)
    if session_data:
        print(f"Cleaning up interactive session for {sid}")
        if session_data.get('channel'):
            try: session_data['channel'].close()
            except: pass
        if session_data.get('client'):
            try: session_data['client'].close()
            except: pass

@socketio.on('connect')
def handle_connect():
    print(f"Client connected: {request.sid}")
    
@socketio.on('request_full_state')
def handle_full_state_request(data):
    """Handle client requests for full application state after reconnection.
    
    This ensures clients can recover state after disconnection or page refresh.
    """
    print(f"Client {request.sid} requested full state")
    try:
        # Prepare and send the complete application state
        global AUDIT_PROGRESS, ui_logs
        
        # Convert progress data for JSON serialization
        json_safe_progress = prepare_progress_for_json(AUDIT_PROGRESS)
        
        # Prepare full state response with all necessary data
        full_state = {
            'logs': ui_logs[-100:] if ui_logs else [],
            'progress': json_safe_progress,
            'audit_status': AUDIT_PROGRESS.get('status', 'idle'),
            'audit_paused': audit_paused if 'audit_paused' in globals() else False
        }
        
        # Return the state directly to the client that requested it
        return full_state
    except Exception as e:
        print(f"Error preparing full state: {e}")
        return {'error': 'Failed to prepare application state'}

    # Send current logs to the newly connected client
    emit('log_update', {'logs': ui_logs})

@socketio.on('disconnect')
def handle_socket_disconnect(): print(f"Socket.IO client disconnected: {request.sid}"); cleanup_interactive_session(request.sid) # (No changes needed)

@app.route('/clear_logs', methods=['POST'])
def clear_logs():
    """Clear the UI logs."""
    global ui_logs
    ui_logs.clear()
    # Notify all clients that logs have been cleared
    socketio.emit('log_update', {'logs': ui_logs}, namespace='/')
    return jsonify({'status': 'success', 'message': 'Logs cleared'})

@app.route('/get_logs')
def get_logs():
    global ui_logs
    return jsonify({'logs': ui_logs})

@app.route('/get_audit_progress')
def get_audit_progress():
    """API endpoint to retrieve the current audit progress data"""
    global AUDIT_PROGRESS, last_run_summary_data, chart_data, audit_paused, audit_status, current_run_failures
    
    try:
        # Calculate enhanced progress data (similar to audit_progress_data function)
        progress_percent = 0
        if AUDIT_PROGRESS.get('total_devices', 0) > 0:
            progress_percent = (AUDIT_PROGRESS.get('completed_devices', 0) / AUDIT_PROGRESS.get('total_devices', 0)) * 100
        
        enhanced_progress = {
            'percentage': progress_percent,
            'status': AUDIT_PROGRESS.get('status', 'idle'),
            'current_device': AUDIT_PROGRESS.get('current_device'),
            'total_devices': AUDIT_PROGRESS.get('total_devices', 0),
            'completed_devices': AUDIT_PROGRESS.get('completed_devices', 0),
            'start_time': AUDIT_PROGRESS.get('start_time'),
            'end_time': AUDIT_PROGRESS.get('end_time'),
            'current_device_start_time': AUDIT_PROGRESS.get('current_device_start_time'),
            'estimated_completion_time': AUDIT_PROGRESS.get('estimated_completion_time'),
            'status_counts': {
                'success': AUDIT_PROGRESS.get('overall_success_count', 0),
                'warning': AUDIT_PROGRESS.get('overall_warning_count', 0),
                'failure': AUDIT_PROGRESS.get('overall_failure_count', 0),
                'in_progress': 1 if AUDIT_PROGRESS.get('status') == 'running' else 0
            },
            'device_statuses': AUDIT_PROGRESS.get('device_statuses', {})
        }
        
        # Create response data with both raw progress and enhanced progress
        # Handle possible undefined globals with defaults
        _audit_status = 'idle'
        if 'audit_status' in globals():
            _audit_status = audit_status
            
        _last_run_summary = {}
        if 'last_run_summary_data' in globals() and last_run_summary_data is not None:
            _last_run_summary = last_run_summary_data
            
        _chart_data = {}
        if 'chart_data' in globals() and chart_data is not None:
            _chart_data = chart_data
            
        _audit_paused = False
        if 'audit_paused' in globals():
            _audit_paused = audit_paused
            
        # Handle current_run_failures safely with type checking
        serializable_failures = {}
        if 'current_run_failures' in globals() and current_run_failures is not None:
            # Check if it's a dictionary before trying to use .items()
            if isinstance(current_run_failures, dict):
                serializable_failures = {k: str(v) if v is not None else None 
                                      for k, v in current_run_failures.items()}
            elif isinstance(current_run_failures, list):
                # If it's a list, convert to a dictionary with indices as keys
                serializable_failures = {str(i): str(v) if v is not None else None
                                      for i, v in enumerate(current_run_failures)}
            else:
                # For any other type, just convert to string
                serializable_failures = {'data': str(current_run_failures)}
        
        data_to_send = {
            'progress': prepare_progress_for_json(AUDIT_PROGRESS),  # Legacy progress format
            'enhanced_progress': enhanced_progress,  # New enhanced progress format
            'overall_audit_status': _audit_status, 
            'last_run_summary': _last_run_summary, 
            'current_run_failures': serializable_failures, 
            'chart_data': _chart_data, 
            'audit_paused': _audit_paused
        }
        
        return jsonify(data_to_send)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"[ERROR] Exception in /get_audit_progress: {e}\n{tb}")
        return jsonify({'error': str(e)}), 500
@app.route('/log_client_error', methods=['POST'])
def log_client_error():
    """API endpoint for logging client-side errors.
    
    This endpoint receives error reports from the frontend Error Boundary system,
    logs them appropriately, and provides a central place for error monitoring.
    """
    try:
        # Get the error data from the request
        error_data = request.json
        
        if not error_data:
            return jsonify({'status': 'error', 'message': 'No error data provided'}), 400
        
        # Log the error with appropriate details
        error_id = error_data.get('id', f"error_{int(datetime.now().timestamp())}")
        component = error_data.get('componentName', 'Unknown Component')
        error_message = error_data.get('error', {}).get('message', 'Unknown error')
        stack = error_data.get('error', {}).get('stack', '')
        context = error_data.get('context', {})
        
        # Create a structured log entry
        log_entry = {
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'error_id': error_id,
            'component': component,
            'message': error_message,
            'stack': stack,
            'context': context,
            'user_agent': request.headers.get('User-Agent', 'Unknown'),
            'ip': request.remote_addr,
        }
        
        # Log this both to application logs and UI logs
        print(f"[CLIENT ERROR] {component}: {error_message}")
        log_to_ui_and_console(
            f"Client Error in {component}: {error_message}", 
            log_level="ERROR"
        )
        
        # Store client errors in a global list (with a maximum size to prevent memory issues)
        global client_errors
        if 'client_errors' not in globals():
            client_errors = []
            
        client_errors.append(log_entry)
        # Keep only the last 100 errors
        if len(client_errors) > 100:
            client_errors = client_errors[-100:]
        
        # Return success response
        return jsonify({
            'status': 'success',
            'message': 'Error logged successfully',
            'error_id': error_id
        })
    except Exception as e:
        print(f"[ERROR] Failed to log client error: {e}")
        return jsonify({'status': 'error', 'message': f'Failed to log error: {str(e)}'}), 500

@socketio.on('start_interactive_shell')
def handle_start_interactive_shell(data: Dict[str, Any]): # (No changes needed)
    sid = request.sid
    if sid in interactive_sessions: emit('shell_error', {'error': 'Shell already active.'}, room=sid); return
    print(f"Attempting interactive shell for: {sid}"); client = paramiko.SSHClient(); client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        jump_host = APP_CONFIG.get("JUMP_HOST"); jump_user = APP_CONFIG.get("JUMP_USERNAME"); jump_pass = APP_CONFIG.get("JUMP_PASSWORD")
        if not all([jump_host, jump_user]): emit('shell_error', {'error': "Jump host/user not configured."}, room=sid); return
        client.connect(jump_host, username=jump_user, password=jump_pass, timeout=10, allow_agent=False, look_for_keys=False)
        channel = client.invoke_shell(term='xterm-color', width=data.get('cols', 80), height=data.get('rows', 24))
        reader_thread = threading.Thread(target=interactive_shell_reader, args=(sid, channel), daemon=True)
        interactive_sessions[sid] = {'client': client, 'channel': channel, 'thread': reader_thread}; reader_thread.start()
        emit('shell_started', {'message': f'Connected to {jump_host}.'}, room=sid); print(f"Interactive shell started for {sid} on {jump_host}")
    except Exception as e: print(f"Failed to start shell for {sid}: {e}"); emit('shell_error', {'error': f"Failed to connect: {sanitize_log_message(str(e))}"}, room=sid); client.close()

@socketio.on('terminal_input')
def handle_terminal_input(data: Dict[str, str]): # (No changes needed)
    sid = request.sid
    if sid in interactive_sessions and interactive_sessions[sid].get('channel'):
        try: interactive_sessions[sid]['channel'].send(data['input'])
        except Exception as e: print(f"Error sending input for {sid}: {e}")
    else: emit('shell_error', {'error': 'Shell not active.'}, room=sid)

@socketio.on('resize_terminal')
def handle_resize_terminal(data: Dict[str, int]): # (No changes needed)
    sid = request.sid
    if sid in interactive_sessions and interactive_sessions[sid].get('channel'):
        try: interactive_sessions[sid]['channel'].resize_pty(width=data.get('cols', 80), height=data.get('rows', 24))
        except Exception as e: print(f"Error resizing PTY for {sid}: {e}")

@socketio.on('stop_interactive_shell')
def handle_stop_interactive_shell(): # (No changes needed)
    sid = request.sid; print(f"Received stop_interactive_shell for {sid}"); cleanup_interactive_session(sid)
    emit('shell_stopped', {'message': 'Shell disconnected by user.'}, room=sid)

if __name__ == '__main__':
    # Initialize all global state variables to ensure a clean start
    # Note: We don't need 'global' statements at main level
    
    # Reset audit status
    audit_status = "idle"
    
    # Reset all audit tracking data structures
    AUDIT_PROGRESS = {
        'status': 'idle',
        'status_message': 'Ready',
        'current_phase': 'Not started',
        'total_devices': 0,
        'completed_devices': 0,
        'current_device': None,
        'start_time': None,
        'end_time': None,
        'current_device_start_time': None,
        'estimated_completion_time': None,
        'device_times': {},
        'device_statuses': {},
        'overall_success_count': 0,
        'overall_warning_count': 0,
        'overall_failure_count': 0,
        'in_progress_count': 0
    }
    
    # Update current_audit_progress
    current_audit_progress = {
        'status': 'idle',
        'status_message': 'Ready',
        'current_device_hostname': 'None',
        'devices_processed_count': 0,
        'total_devices_to_process': 0,
        'percentage_complete': 0,
        'current_phase': 'Not started',
        'overall_success_count': 0,
        'overall_warning_count': 0,
        'overall_failure_count': 0,
        'device_status_class': 'status-idle'
    }
    
    # Reset summary data for statistics display
    last_run_summary_data.clear()
    last_run_summary_data.update({
        "total_routers": 0,
        "icmp_reachable": 0,
        "ssh_auth_ok": 0,
        "collected": 0,
        "with_violations": 0,
        "failed_icmp": 0,
        "failed_ssh_auth": 0,
        "failed_collection": 0,
        "per_router_status": {},
        "total_physical_line_issues": 0,
        "failure_category_explicit_telnet": 0,
        "failure_category_transport_all": 0,
        "failure_category_default_telnet": 0
    })
    
    # Reset failure tracking
    current_run_failures.clear()
    
    # Reset detailed reports manifest
    detailed_reports_manifest.clear()
    
    # Reset timestamps
    last_audit_start_time_attempt = None
    last_successful_audit_completion_time = None
    
    # Clear UI logs to start fresh
    ui_logs.clear()
    
    # Set up script environment
    script_dir = os.path.dirname(os.path.abspath(__file__))
    report_base_dir_main = os.path.join(script_dir, BASE_DIR_NAME)
    os.makedirs(report_base_dir_main, exist_ok=True)
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    
    if not list_inventory_files(): # Check if any inventory files exist
        log_to_ui_and_console("No inventory files found. Creating default YAML and CSV inventories.", console_only=True)
        create_default_inventory_if_missing() # This now creates both YAML and CSV defaults
        load_active_inventory()
    
    log_to_ui_and_console(f"Reports in: {report_base_dir_main}", console_only=True)
    log_to_ui_and_console(f"Inventories in: {app.config['UPLOAD_FOLDER']}", console_only=True)
    log_to_ui_and_console(f"Active inventory: {APP_CONFIG.get('ACTIVE_INVENTORY_FILE', 'N/A')} (Format: {APP_CONFIG.get('ACTIVE_INVENTORY_FORMAT','yaml').upper()})", console_only=True)
    log_to_ui_and_console(f"Jump Ping Path: {APP_CONFIG.get('JUMP_PING_PATH', '/bin/ping (default)')}", console_only=True)
    port = app.config['PORT']
    log_to_ui_and_console(f"Flask App with SocketIO running on http://0.0.0.0:{port}", console_only=True)
    log_to_ui_and_console(f"Access UI at: http://127.0.0.1:{port}", console_only=True)
    socketio.run(app, host='0.0.0.0', port=port, debug=True)